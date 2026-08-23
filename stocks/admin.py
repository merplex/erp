import json
import datetime # ✅ เพิ่มตัวนี้
from django.contrib import admin
from django.utils.html import format_html
from unfold.admin import ModelAdmin as UnfoldModelAdmin, TabularInline as UnfoldTabularInline, StackedInline as UnfoldStackedInline
from .models import ProductTag
from .models import *
from .models import (
    Product, ProductTag, ProductCategory, Supplier, 
    ProductBarcode, ProductSupplier,
    PurchaseOrder, PurchaseItem, PurchaseReceiptLog, PurchasePaymentLog,
    SalesOrder, SalesItem, SalesDeliveryLog, SalesPayment,
    ProductionOrder, ProductionMaterialUsage, ProductionLog,
    BOM, BOMIngredient, DocumentLock, StockPlanning, 
    StockAdjustment, Customer, CustomerProductContract, FinanceReport, 
    IncomeReport, ShipmentAccounting, InternationalPurchaseTracking,
    SalesReport  # 👈 เพิ่มตัวที่ทำพังเมื่อกี้เข้าไปแล้วครับ!
)
from .models import DocumentLock
# 1. เปลี่ยนชื่อที่ปรากฏบนหัวเอกสาร (Header สีน้ำเงิน)
admin.site.site_header = "Meebun ERP"

# 2. เปลี่ยนชื่อที่ปรากฏบน Browser Tab (Title)
admin.site.site_title = "Meebun ERP Admin"

# 3. เปลี่ยนชื่อหัวข้อหลักในหน้าแรก (Index Title)
admin.site.index_title = "ยินดีต้อนรับสู่ระบบจัดการข้อมูล"

# ตั้งค่า global สำหรับทุกหน้า
admin.ModelAdmin.list_per_page = 200
admin.ModelAdmin.show_full_result_count = False
from django.contrib import messages

from django.contrib.admin.widgets import AdminDateWidget
from django.contrib.admin import helpers  # <--- helpers ต้องดึงมาจาก admin ครับ
from django.utils.html import format_html
from django.core.exceptions import ValidationError
from django.forms import TextInput
from django.db import models # เพิ่มเพื่อรองรับ formfield_overrides
from django.db.models import Subquery, OuterRef, Q, Sum, F, DecimalField, ExpressionWrapper, Case, When, IntegerField, Value
from django.db.models.functions import TruncDate
from django.db.models.functions import Coalesce, Greatest
from django import forms # ✅ เพิ่มบรรทัดนี้ครับ ทำระบบ tag checkbox
from django.utils.safestring import mark_safe # ✅ ต้องมีบรรทัดนี้ครับ
# เพิ่มที่บรรทัดบนสุดของไฟล์ครับ
from django.http import HttpResponseRedirect
from django.template import Template, RequestContext 
from django.http import HttpResponse, HttpResponseRedirect
from django.template.response import TemplateResponse
from django.template.loader import render_to_string
from django.urls import reverse, path # ✅ 3บรรทัดนี้ สำหรับระบบล็อคเอกสาร
from django.contrib.contenttypes.models import ContentType
from django.contrib.admin.models import LogEntry
from django.core.paginator import Paginator
from datetime import timedelta
from django.utils import timezone
from decimal import Decimal
from unfold.contrib.filters.admin import RangeDateFilter as DjangoDateRangeFilter
from unfold.contrib.filters.admin import RangeDateTimeFilter as DjangoDateTimeRangeFilter
from unfold.contrib.filters.admin import (
    AutocompleteSelectMultipleFilter,
    MultipleChoicesDropdownFilter,
    BooleanRadioFilter,
)
from django.core.validators import EMPTY_VALUES
from django.forms import ValidationError as FilterValidationError
from unfold.utils import parse_datetime_str
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def _strip_html(text):
    if not isinstance(text, str):
        return '' if text is None else str(text)
    return re.sub(r'<[^>]+>', '', text).strip()


class RangeDateTimeFilter(DjangoDateTimeRangeFilter):
    """เหมือน unfold's RangeDateTimeFilter ทุกอย่าง ยกเว้นตอนกรอง:
    ถ้าเลือกวันที่แต่ไม่ได้กรอกเวลา ปกติ unfold จะไม่กรองฝั่งนั้นเลย (เงียบๆ ไม่มีผล)
    ที่นี่ถ้าไม่กรอกเวลา จะถือว่า "เริ่มต้นวัน" (00:01) / "สิ้นสุดวัน" (23:59) แทน
    ส่วนถ้าไม่เลือกวันที่เลย (ทั้งวันและเวลาว่าง) ยังคงไม่กรองฝั่งนั้นเหมือนเดิม
    (from ว่าง = ตั้งแต่รายการแรก, to ว่าง = ถึงรายการหลังสุด)"""

    def queryset(self, request, queryset):
        filters = {}

        date_value_from = self.used_parameters.get(f"{self.parameter_name}_from_0")
        time_value_from = self.used_parameters.get(f"{self.parameter_name}_from_1")
        if date_value_from not in EMPTY_VALUES and time_value_from in EMPTY_VALUES:
            time_value_from = "00:01"

        date_value_to = self.used_parameters.get(f"{self.parameter_name}_to_0")
        time_value_to = self.used_parameters.get(f"{self.parameter_name}_to_1")
        if date_value_to not in EMPTY_VALUES and time_value_to in EMPTY_VALUES:
            time_value_to = "23:59"

        if date_value_from not in EMPTY_VALUES and time_value_from not in EMPTY_VALUES:
            filters[f"{self.parameter_name}__gte"] = parse_datetime_str(f"{date_value_from} {time_value_from}")

        if date_value_to not in EMPTY_VALUES and time_value_to not in EMPTY_VALUES:
            filters[f"{self.parameter_name}__lte"] = parse_datetime_str(f"{date_value_to} {time_value_to}")

        try:
            return queryset.filter(**filters)
        except (ValueError, FilterValidationError):
            return None


class ExportToExcelMixin:
    """เพิ่ม action Export Excel ให้ ModelAdmin ใดก็ได้"""

    @admin.action(description="📊 Export เป็น Excel")
    def export_to_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(self.model._meta.verbose_name_plural or 'Export')[:31]

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 28

        # รวบรวม columns จาก list_display
        columns = []
        for field_name in (self.list_display or []):
            if field_name == 'action_checkbox':
                continue
            header = field_name.replace('_', ' ').title()
            # หา short_description จาก admin method ก่อน
            if hasattr(self, field_name):
                attr = getattr(self, field_name)
                if hasattr(attr, 'short_description'):
                    header = attr.short_description
            else:
                # ลองหาจาก model field
                try:
                    field = self.model._meta.get_field(field_name)
                    header = str(field.verbose_name).capitalize()
                except Exception:
                    # ลองหาจาก model method/property
                    if hasattr(self.model, field_name):
                        attr = getattr(self.model, field_name)
                        if hasattr(attr, 'short_description'):
                            header = attr.short_description
            columns.append((header, field_name))

        # เขียน header row
        for col_idx, (header, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # keyword ที่บ่งบอกว่า field นี้ควรเป็น text (ไม่แปลงเป็นตัวเลข)
        TEXT_FIELD_KEYWORDS = ('code', 'barcode', 'บาร์โค้ด', 'รหัส', 'เลขที่', 'เบอร์', 'phone', 'tax')

        # เขียนข้อมูล
        for row_idx, obj in enumerate(queryset, start=2):
            for col_idx, (header, field_name) in enumerate(columns, start=1):
                value = ''
                is_text_field = any(k in field_name.lower() or k in header.lower()
                                    for k in TEXT_FIELD_KEYWORDS)
                try:
                    if hasattr(self, field_name):
                        raw = getattr(self, field_name)(obj)
                    elif hasattr(obj, field_name):
                        raw = getattr(obj, field_name)
                        if callable(raw):
                            raw = raw()
                    else:
                        raw = ''
                    value = _strip_html(str(raw)) if raw is not None else ''
                    if not is_text_field:
                        # แปลงตัวเลขถ้าทำได้
                        clean = value.replace(',', '').replace('%', '').strip()
                        if clean:
                            try:
                                value = int(clean) if '.' not in clean else float(clean)
                            except ValueError:
                                pass
                except Exception:
                    value = ''
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # บังคับ format text สำหรับ barcode/code field
                if is_text_field and isinstance(value, str):
                    cell.number_format = '@'

        # ปรับความกว้าง column อัตโนมัติ
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

        ws.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        model_name = self.model._meta.model_name
        response['Content-Disposition'] = f'attachment; filename="{model_name}_export.xlsx"'
        wb.save(response)
        return response


class DocumentLockMixin:
    def change_view(self, request, object_id, form_url='', extra_context=None):
        content_type = ContentType.objects.get_for_model(self.model)
        
        # 1. เช็คว่ามีใครล็อกใบนี้อยู่ไหม
        lock = DocumentLock.objects.filter(content_type=content_type, object_id=object_id).first()
        
        if lock:
            # 2. ถ้ามีคนล็อกอยู่ และไม่ใช่เรา + ล็อกยังไม่หมดอายุ -> "ห้ามเข้า"
            if lock.user != request.user and not lock.is_expired():
                messages.error(
                    request, 
                    f"⛔ หยุดก่อน! ใบนี้กำลังถูกแก้ไขโดย {lock.user.get_full_name() or lock.user.username} "
                    f"กรุณารอประมาณ 10 นาที หรือติดต่อผู้ใช้คนดังกล่าวค่ะ"
                )
                # ดีดกลับไปหน้า List ทันที
                return HttpResponseRedirect(reverse(f'admin:{self.model._meta.app_label}_{self.model._meta.model_name}_changelist'))
            
            # 3. ถ้าเป็นเราเอง หรือล็อกมันหมดอายุแล้ว -> "ต่ออายุล็อก"
            lock.user = request.user
            lock.save()
        else:
            # 4. ถ้ายังไม่มีใครล็อก -> "สร้างล็อกใหม่"
            DocumentLock.objects.create(content_type=content_type, object_id=object_id, user=request.user)
            
        return super().change_view(request, object_id, form_url, extra_context)

    def save_model(self, request, obj, form, change):
        # เมื่อกด Save เสร็จสมบูรณ์ -> "ปลดล็อก" ให้คนอื่นเข้าต่อได้ทันที
        super().save_model(request, obj, form, change)
        content_type = ContentType.objects.get_for_model(self.model)
        DocumentLock.objects.filter(content_type=content_type, object_id=obj.pk).delete()

    def render_change_form(self, request, context, add=False, change=False, form_url='', obj=None):
        response = super().render_change_form(request, context, add, change, form_url, obj)
        if obj and change:
            content_type = ContentType.objects.get_for_model(self.model)
            script = (
                f'<script>'
                f'function _sendUnlockDoc(){{'
                f'var fd=new FormData();'
                f'fd.append("content_type_id","{content_type.pk}");'
                f'fd.append("object_id","{obj.pk}");'
                # แนบ CSRF token ไปด้วยเสมอ (ไม่พึ่ง @csrf_exempt อย่างเดียว เผื่อ middleware ไม่ยกเว้นให้จริง)
                f'var _m=document.cookie.match(/csrftoken=([^;]+)/);'
                f'var _csrf=_m?_m[1]:(document.querySelector("[name=csrfmiddlewaretoken]")||{{}}).value;'
                f'if(_csrf)fd.append("csrfmiddlewaretoken",_csrf);'
                f'navigator.sendBeacon("/admin/unlock-doc/",fd);'
                f'}}'
                # beforeunload ใช้ไม่ได้เสมอไป (มือถือ/LINE in-app browser/bfcache มักไม่ยิง)
                # เลยเสริม pagehide กับ visibilitychange ไว้ด้วยกันไม่ให้ lock ค้าง
                f'window.addEventListener("beforeunload",_sendUnlockDoc);'
                f'window.addEventListener("pagehide",_sendUnlockDoc);'
                f'document.addEventListener("visibilitychange",function(){{'
                f'if(document.visibilityState==="hidden")_sendUnlockDoc();'
                f'}});'
                f'</script>'
            )
            response.render()
            response.content = response.content.replace(b'</body>', script.encode() + b'</body>', 1)
        return response

class DetailedHistoryMixin:
    """
    ปกติ Django log ("ประวัติ"/History) จะบอกแค่ *ชื่อ field* ที่ถูกแก้ (เช่น
    "Changed latest_buy_price") ไม่บอกว่าแก้จากค่าอะไรเป็นค่าอะไร — mixin นี้
    เขียน change_message ใหม่ให้มีค่าเดิม → ค่าใหม่ครบ ทั้ง field บนฟอร์มหลัก
    และทุก inline (เช่น ราคา supplier ใน A4, บาร์โค้ด, ราคาสัญญาลูกค้า)
    """

    def _fmt_value(self, value):
        if value in (None, ''):
            return '(ว่าง)'
        if isinstance(value, bool):
            return 'ใช่' if value else 'ไม่ใช่'
        return str(value)

    def _field_diffs(self, form, field_names):
        lines = []
        for field_name in field_names:
            if field_name not in form.fields:
                continue
            old_val = form.initial.get(field_name)
            new_val = getattr(form.instance, field_name, None)
            # initial ของ FK field เก็บเป็น pk ดิบ ๆ แปลงเป็น object ก่อน เพื่อ str() อ่านง่าย
            field = form.fields[field_name]
            if hasattr(field, 'queryset') and old_val is not None and not hasattr(old_val, 'pk'):
                try:
                    old_val = field.queryset.model.objects.filter(pk=old_val).first() or old_val
                except Exception:
                    pass
            old_s, new_s = self._fmt_value(old_val), self._fmt_value(new_val)
            if old_s == new_s:
                continue
            label = field.label or field_name
            lines.append(f"{label}: {old_s} → {new_s}")
        return lines

    def construct_change_message(self, request, form, formsets, add=False):
        if add:
            return super().construct_change_message(request, form, formsets, add)

        messages = []
        if form.changed_data:
            diffs = self._field_diffs(form, form.changed_data)
            if diffs:
                messages.append("แก้ไข: " + "; ".join(diffs))

        for formset in (formsets or []):
            for obj in formset.new_objects:
                messages.append(f"เพิ่ม {obj._meta.verbose_name}: {obj}")

            if formset.changed_objects:
                form_by_pk = {f.instance.pk: f for f in formset.forms if f.instance.pk}
                for changed_object, changed_fields in formset.changed_objects:
                    matched_form = form_by_pk.get(changed_object.pk)
                    diffs = self._field_diffs(matched_form, changed_fields) if matched_form else []
                    detail = "; ".join(diffs) if diffs else ", ".join(changed_fields)
                    messages.append(f"แก้ไข {changed_object._meta.verbose_name} ({changed_object}): {detail}")

            for obj in formset.deleted_objects:
                messages.append(f"ลบ {obj._meta.verbose_name}: {obj}")

        if not messages:
            return super().construct_change_message(request, form, formsets, add)
        return "\n".join(messages)


class PurchaseOrderTagsFilter(AutocompleteSelectMultipleFilter):
    """
    Filter 'By tag ของสินค้าในใบ' สำหรับ PurchaseOrder/FinanceReport — ใช้ Exists()
    แทนการ .filter(items__product__tags__in=...) ตรงๆ เพราะ join แบบนั้นจะไปชน join
    เดิมที่ annotate(Sum('items__...')) ใช้อยู่แล้ว (ใน FinanceReportAdmin) ทำให้ยอดรวม
    เงินพองขึ้นถ้า item ตัวเดียวมีหลาย tag ตรงกับตัวกรองพร้อมกัน — Exists() ปลอดภัยกว่า
    เพราะเป็นแค่เงื่อนไข boolean ไม่เพิ่ม join ให้ query หลัก
    """
    def queryset(self, request, queryset):
        value = self.value()
        ids = value if isinstance(value, (list, tuple)) else ([value] if value else [])
        ids = [v for v in ids if v not in EMPTY_VALUES]
        if not ids:
            return queryset
        from django.db.models import Exists, OuterRef
        from .models import PurchaseItem
        sub = PurchaseItem.objects.filter(purchase_order=OuterRef('pk'), product__tags__id__in=ids)
        return queryset.filter(Exists(sub))


class SalesOrderTagsFilter(AutocompleteSelectMultipleFilter):
    """เหมือน PurchaseOrderTagsFilter แต่ฝั่งขาย (SalesOrder/IncomeReport)"""
    def queryset(self, request, queryset):
        value = self.value()
        ids = value if isinstance(value, (list, tuple)) else ([value] if value else [])
        ids = [v for v in ids if v not in EMPTY_VALUES]
        if not ids:
            return queryset
        from django.db.models import Exists, OuterRef
        from .models import SalesItem
        sub = SalesItem.objects.filter(sales_order=OuterRef('pk'), product__tags__id__in=ids)
        return queryset.filter(Exists(sub))


class ProductOnlyFilter(admin.SimpleListFilter):
    title = 'ประเภทรายการ' # หัวข้อบนแถบ Filter
    parameter_name = 'is_product'

    def lookups(self, request, model_admin):
        return (
            ('true', 'สินค้าเท่านั้น'),
            ('false', 'ไม่ใช่สินค้า (ค่าบริการ/อื่นๆ)'),
            ('all', 'แสดงทั้งหมด'),
        )

    def queryset(self, request, queryset):
        # ✅ กำหนด Logic การกรอง
        if self.value() == 'true':
            return queryset.filter(is_product=True)
        if self.value() == 'false':
            return queryset.filter(is_product=False)
        if self.value() == 'all':
            return queryset
        
        # 🎯 จุดสำคัญ: ถ้ายังไม่ได้เลือก (Default) ให้โชว์แค่สินค้า
        if self.value() is None:
            return queryset.filter(is_product=True)
        return queryset
    
class DatePeriodFilter(admin.SimpleListFilter):
    title = 'ช่วงเวลารายงาน'
    parameter_name = 'period'

    def lookups(self, request, model_admin):
        return (
            ('1year', 'ย้อนหลัง 1 ปี (Default)'),
            ('4months', 'ย้อนหลัง 4 เดือน'),
            ('1month', 'ย้อนหลัง 1 เดือน'),
        )

    def queryset(self, request, queryset):
        now = timezone.now()
        if self.value() == '1year':
            return queryset.filter(sales_items__sales_order__order_date__year=now.year)
        if self.value() == '4months':
            start_date = now - timedelta(days=120)
            return queryset.filter(sales_items__sales_order__order_date__gte=start_date)
        if self.value() == '1month':
            start_date = now - timedelta(days=30)
            return queryset.filter(sales_items__sales_order__order_date__gte=start_date)
        return queryset # Default จะไปจัดการใน get_queryset

# ✅ 1. Inline รายการสินค้า (แบบ Read-Only สำหรับหน้าการเงิน)
class PurchaseItemReadOnlyInline(UnfoldTabularInline):
    model = PurchaseItem
    extra = 0
    can_delete = False # ห้ามลบรายการ
    verbose_name = "🛒 รายการสินค้า (ตรวจสอบราคา)"
    verbose_name_plural = "🛒 รายการสินค้า (Read-Only)"
    
    # โชว์ครบ: สินค้า, จำนวน, ราคาต่อหน่วย (ที่ล็อกแล้ว), ราคารวมบรรทัด
    fields = ('product', 'quantity_ordered', 'unit_price', 'get_line_total')
    readonly_fields = ('product', 'quantity_ordered', 'unit_price', 'get_line_total')

    def get_line_total(self, obj):
        return f"{obj.total_price:,.2f}"
    get_line_total.short_description = "ราคารวม"

    def has_add_permission(self, request, obj=None): return False

# ✅ เพิ่มอันนี้เข้าไปครับ: ตารางแสดงรายการสินค้า (แบบดูได้อย่างเดียว)
class SalesItemReadOnlyInline(UnfoldTabularInline):
    model = SalesItem  # ชื่อ Model สินค้าฝั่งขาย (เช็คใน models.py ว่าชื่อนี้ไหม)
    extra = 0
    fields = ['product', 'quantity_ordered', 'sale_price', 'get_total_display', 'auto_produce']
    readonly_fields = ['product', 'quantity_ordered', 'sale_price', 'get_total_display', 'auto_produce']
    can_delete = False
    verbose_name = "📦 รายการสินค้าที่ขาย"
    verbose_name_plural = "รายการสินค้า"
    
    def has_add_permission(self, request, obj):
        return False
    
    def get_unit_price(self, obj):
        # ใช้ sale_price ตามที่เปรมบอก
        price = obj.product.sale_price if obj.product else 0 
        return f"{price:,.2f}"
    get_unit_price.short_description = "ราคาขาย (@)"

    # ✅ คำนวณยอดรวม (จำนวนที่สั่ง x ราคาขาย)
    def get_line_total(self, obj):
        price = obj.product.sale_price if obj.product else 0
        total = price * obj.quantity_ordered
        return f"{total:,.2f}"
    get_line_total.short_description = "รวมเงิน"

    def get_total_display(self, obj):
        # คำนวณ: จำนวน x ราคาขาย
        price = obj.sale_price or 0
        qty = obj.quantity_ordered or 0
        total = price * qty
        return f"{total:,.2f}"
    
    get_total_display.short_description = "ราคารวม"
    
# ✅ 2. Inline การจ่ายเงิน และการรับเงิน (บันทึกยอดได้เรื่อยๆ)
class PurchasePaymentInline(UnfoldTabularInline):
    model = PurchasePaymentLog
    extra = 1
    verbose_name = "💰 บันทึกการจ่ายเงิน"
    verbose_name_plural = "💰 ประวัติการจ่ายเงิน (Payments)"
    fields = ('amount', 'notes', 'payment_date', 'user')
    readonly_fields = ('user',)


class SalesPaymentInline(UnfoldTabularInline):
    model = SalesPayment
    extra = 1
    verbose_name = "💰 รายการรับเงิน"
    verbose_name_plural = "ประวัติการรับเงิน (กรอกเองกรณีแบ่งจ่าย / หรือกด Action หน้ารวมเพื่อรับเต็มจำนวน)"
    fields = ('payment_date', 'amount', 'remark', 'evidence')
    readonly_fields = ('get_status_from_logs',)

    def get_status_from_logs(self, obj):
        # ใช้ prefetch_related จาก get_queryset แทนการ query ใหม่ต่อแถว
        logs = obj.sales_order.delivery_logs.all()
        confirmed = any(getattr(log, 'is_revenue_confirmed', False) for log in logs)
        return "ยืนยันแล้วจากหน้า C6" if confirmed else "รอยืนยัน"
    get_status_from_logs.short_description = "สถานะรับเงิน"

    def has_change_permission(self, request, obj=None):
        # 🎯 ถ้าใบสั่งขายนี้มียอดที่คอนเฟิร์มใน C6 แล้ว ห้ามแก้หน้า C3
        if obj:
            from .models import SalesDeliveryLog # 👈 Import มาใช้ตรงๆ
            already_confirmed = SalesDeliveryLog.objects.filter(
                sales_order=obj, 
                is_revenue_confirmed=True
            ).exists()
            
            if already_confirmed:
                return False
        return True
    
# ---------------------------------------------------------
# 1. รายการสั่งซื้อ (ค้างรับ) -> ใช้ po_number และติดลบ
# ---------------------------------------------------------
class PendingPurchaseInline(UnfoldTabularInline):
    model = PurchaseItem
    fields = ['get_ref_no', 'quantity_ordered', 'quantity_received', 'get_pending']
    readonly_fields = fields
    extra = 0
    can_delete = False
    verbose_name = "🛒 รายการสั่งซื้อ (ค้างรับ)"
    verbose_name_plural = "🛒 รายการสั่งซื้อค้างรับ"

    def get_queryset(self, request):
        return super().get_queryset(request).filter(
            quantity_ordered__gt=F('quantity_received')
        ).exclude(
            purchase_order__status__in=['Received', 'Completed', 'Cancelled']
        )

    def get_ref_no(self, obj):
        # ✅ แก้จาก obj.order เป็น obj.purchase_order ตามโครงสร้างเปรม
        return obj.purchase_order.po_number 
    get_ref_no.short_description = "PO No."

    def get_pending(self, obj):
        diff = obj.quantity_ordered - obj.quantity_received
        return format_html('<b style="color:#dc3545;">-{}</b>', diff)
    get_pending.short_description = "ขาดรับ"

    def has_add_permission(self, request, obj=None): return False

# ---------------------------------------------------------
# 2. รายการผลิต (ค้างผลิต) -> ใช้ pd_number
# ---------------------------------------------------------
class PendingProductionInline(UnfoldTabularInline):
    model = ProductionOrder
    fields = ['pd_number', 'quantity_planned', 'quantity_actual', 'get_pending']
    readonly_fields = fields
    extra = 0
    can_delete = False
    verbose_name = "🔨 รายการผลิต (ค้างผลิต)"
    verbose_name_plural = "🔨 รายการผลิตค้างผลิต"

    def get_queryset(self, request):
        return super().get_queryset(request).filter(
            quantity_planned__gt=F('quantity_actual') # ✅ ยังไม่ครบจำนวน
        ).exclude(
            status__in=['Finished', 'Completed', 'Cancelled'] # ✅ และยังไม่จบงาน/ยกเลิก
        )

    def get_pending(self, obj):
        diff = obj.quantity_planned - obj.quantity_actual
        return format_html('<b style="color:#dc3545;">-{}</b>', diff)
    get_pending.short_description = "ขาดผลิต"

    def has_add_permission(self, request, obj=None): return False

# ---------------------------------------------------------
# 3. รายการขาย (ค้างส่ง) -> ใช้ so_number
# ---------------------------------------------------------
class PendingSaleInline(UnfoldTabularInline):
    model = SalesItem
    fields = ['sales_order_link','quantity_ordered', 'quantity_shipped', 'get_pending','order_status']
    readonly_fields = fields
    extra = 0
    can_delete = False
    verbose_name = "📦 รายการขาย (ค้างส่ง)"
    verbose_name_plural = "📦 รายการขายค้างส่ง"

    def get_queryset(self, request):
        # ✅ กรองเฉพาะ:
        # 1. ยอดที่สั่งซื้อต้อง "มากกว่า" ยอดที่ส่งไปแล้ว (ยังมีของค้างส่ง)
        # 2. สถานะใบสั่งขายต้องไม่ใช่ 'Shipped' (ส่งครบ), 'Completed' (ปิดงาน), หรือ 'Cancelled' (ยกเลิก)
        return super().get_queryset(request).filter(
            quantity_ordered__gt=F('quantity_shipped')
        ).exclude(
            sales_order__status__in=['Completed', 'Cancelled']
        )

    def get_pending(self, obj):
        diff = obj.quantity_ordered - obj.quantity_shipped
        return format_html('<b style="color:#dc3545;">-{}</b>', diff)
    get_pending.short_description = "ขาดส่ง"

    def has_add_permission(self, request, obj=None): return False


    # ✅ แถม: ฟังก์ชันโชว์สถานะของใบสั่งขายในตาราง
    def order_status(self, obj):
        status = obj.sales_order.status
        colors = {
            'Draft': '#6c757d',
            'Confirmed': '#007bff',
            'Partially Shipped': '#ffc107',
        }
        color = colors.get(status, '#000')
        return format_html('<b style="color: {};">{}</b>', color, status)
    order_status.short_description = "สถานะใบสั่ง"

    # ✅ แถม: ฟังก์ชันคลิกที่เลขที่ใบสั่งแล้วกระโดดไปหน้าแก้ไขได้เลย
    def sales_order_link(self, obj):
        from django.urls import reverse
        url = reverse("admin:stocks_salesorder_change", args=[obj.sales_order.id])
        return format_html('<a href="{}">{}</a>', url, obj.sales_order.so_number)
    sales_order_link.short_description = "เลขที่ใบสั่งขาย"

# --- Inlines ---
# ---------------------------------------------------------
# Inline สำหรับจัดการหลายบาร์โค้ดในหน้าเดียว
# ---------------------------------------------------------
class ProductBarcodeInline(UnfoldTabularInline):
    model = ProductBarcode
    extra = 1  # จะมีช่องว่างให้เติม 1 ช่องเสมอ และมีปุ่ม + เพิ่มได้เรื่อยๆ
    verbose_name = "บาร์โค้ดสินค้า"
    verbose_name_plural = "บาร์โค้ดทั้งหมดของสินค้านี้"

@admin.register(ProductBarcode)
class ProductBarcodeAdmin(UnfoldModelAdmin):
    search_fields = ['code', 'product__name', 'unit_name']
    list_display = ['code', 'product', 'unit_name', 'conversion_factor']
    autocomplete_fields = ['product']

    # 🎯 กรอง Autocomplete ของช่อง barcode_obj ให้เห็นเฉพาะบาร์โค้ดของสินค้า/วัตถุดิบที่กำลังเลือกอยู่
    # (JS ที่ส่ง material_id/product_id มา: bom_ingredient_barcode_filter.js)
    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        if 'autocomplete' in request.path:
            material_id = (request.GET.get('material_id') or request.GET.get('product_id') or '').strip()
            if material_id:
                queryset = queryset.filter(product_id=material_id)
        return queryset, use_distinct

class ProductSupplierInline(UnfoldTabularInline):
    model = ProductSupplier
    extra = 1

class SupplierProductInline(UnfoldTabularInline):
    model = ProductSupplier
    extra = 1
    autocomplete_fields = ['product']
    fields = ('product', 'supplier_sku', 'latest_buy_price')

from django import forms # อย่าลืม import forms ไว้ด้านบนนะครับ

class BOMIngredientForm(forms.ModelForm):
    class Meta:
        model = BOMIngredient
        fields = '__all__'
        widgets = {
            # 🎯 บังคับให้ช่อง Quantity รับทศนิยม 4 ตำแหน่ง และขยับทีละ 0.0001
            'quantity': forms.NumberInput(attrs={'step': '0.0001', 'style': 'width: 150px;'}),
        }


class BOMIngredientInline(UnfoldTabularInline):
    model = BOMIngredient
    form = BOMIngredientForm # ✅ เอา Form ที่เราสร้างมาใส่ตรงนี้ครับ
    fields = ('material', 'barcode_obj', 'quantity', 'get_unit_display')
    readonly_fields = ('get_unit_display',)
    autocomplete_fields = ['material', 'barcode_obj']
    extra = 1
    def get_unit_display(self, obj): return obj.get_unit
    get_unit_display.short_description = "หน่วย"

class PurchaseItemInline(UnfoldTabularInline):
    model = PurchaseItem
    # 🎯 เก็บความสามารถเดิมไว้: ช่วยให้ค้นหาชื่อสินค้าได้ไวขึ้น
    autocomplete_fields = ['product', 'barcode_obj']
    extra = 0

    # 🎯 กล่อง "ราคา/หน่วย" แสดงแค่ ~5 หลักก็พอ ไม่ต้องยืดเต็มคอลัมน์
    formfield_overrides = {
        models.DecimalField: {'widget': forms.NumberInput(attrs={'style': 'width: 90px;'})},
    }

    # 🎯 จัดเรียงคอลัมน์ใหม่ตามที่เปรมต้องการ
    fields = [
        'barcode_obj',
        'product',
        'quantity_unit',
        'quantity_ordered',
        'quantity_received',
        'get_pending',     # ✅ คอลัมน์ "ขาดรับ"
        'unit_price',
        'total_price'      # ✅ คอลัมน์ "ราคารวม"
    ]

    # 🎯 ป้องกันการแก้เลขที่ระบบควรคำนวณเอง (quantity_ordered คำนวณอัตโนมัติจาก quantity_unit x หน่วยบาร์โค้ด)
    readonly_fields = ['quantity_ordered', 'quantity_received', 'get_pending', 'total_price']

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "product":
            # 1. พยายามหา ID จากหลายๆ ช่องทาง (ป้องกันชื่อ ID เปลี่ยน)
            resolved = request.resolver_match
            object_id = None
            if resolved:
                object_id = resolved.kwargs.get('object_id') or resolved.kwargs.get('pk')

            if object_id:
                try:
                    from django.db.models import Q
                    # ✅ ใช้ self.parent_model แทนการระบุชื่อตรงๆ จะปลอดภัยกว่าค่ะ
                    parent_obj = self.parent_model.objects.get(pk=object_id)
                    
                    if parent_obj.supplier:
                        # ✅ กรองสินค้า: 
                        # - เป็นสินค้าที่ Supplier นี้ขาย (ผ่าน product_suppliers)
                        # - หรือ เป็นรายการที่ไม่ใช่สินค้า (is_product=False)
                        kwargs["queryset"] = Product.objects.filter(
                            Q(product_suppliers__supplier=parent_obj.supplier) | 
                            Q(is_product=False)
                        ).distinct()
                    
                except Exception as e:
                    # ถ้ามี Error ให้มันพ่นออกมาใน Console เปรมจะได้เห็นค่ะ
                    print(f"🚨 Filter Error: {e}")
            
            # 💡 ถ้าเป็นหน้า "เพิ่มใหม่" (Add Mode) ซึ่งไม่มี ID 
            # ปกติ Django จะโชว์หมด เพราะมันยังไม่รู้ว่าเปรมจะเลือก Supplier คนไหน
            # ถ้าเปรมอยากให้มันว่างไว้ก่อนจนกว่าจะเลือก ให้ใส่บรรทัดนี้ค่ะ (แต่ต้องกด Save รอบนึงก่อนนะ)
            # elif not object_id:
            #     kwargs["queryset"] = Product.objects.none()

        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    def get_pending(self, obj):
        # ตรวจสอบค่าว่างก่อนคำนวณป้องกัน Error
        qty_ordered = obj.quantity_ordered or 0
        qty_received = obj.quantity_received or 0
        
        diff = qty_ordered - qty_received
        
        if diff > 0:
            # ✅ แสดงยอดติดลบสีแดง (-X) สำหรับยอดที่ยังขาดรับ
            return format_html('<b style="color:#dc3545;">-{}</b>', diff)
        return 0
    
    get_pending.short_description = "ขาดรับ"

class PurchaseReceiptLogInline(UnfoldTabularInline):
    model = PurchaseReceiptLog
    extra = 1
    autocomplete_fields = ['barcode_obj']
    fields = ('barcode_obj', 'product', 'supplier_invoice', 'quantity_received', 'user', 'notes', 'received_date')
    readonly_fields = ('user', 'received_date')

    formfield_overrides = {
        models.CharField: {'widget': TextInput(attrs={'style': 'width: 120px;', 'placeholder': 'เลขใบส่งของ'})},
        models.TextField: {'widget': TextInput(attrs={'style': 'width: 200px;', 'placeholder': 'หมายเหตุ'})},
        # 🎯 กล่อง "รับสะสม" ให้ยาวพอสำหรับ 7 หลัก เท่ากับตอนเปิดรายการสั่งซื้อ
        models.PositiveIntegerField: {'widget': forms.NumberInput(attrs={'style': 'width: 110px;'})},
    }

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "product":
            resolved = request.resolver_match
            if resolved and 'object_id' in resolved.kwargs:
                po_id = resolved.kwargs['object_id']
                ordered_ids = list(
                    PurchaseItem.objects.filter(purchase_order_id=po_id)
                    .order_by('id').exclude(product=None)
                    .values_list('product_id', flat=True)
                )
                seen = set()
                unique_ids = [x for x in ordered_ids if not (x in seen or seen.add(x))]
                if unique_ids:
                    preserved = Case(*[When(pk=pk, then=pos) for pos, pk in enumerate(unique_ids)])
                    kwargs["queryset"] = Product.objects.filter(pk__in=unique_ids).order_by(preserved)
                else:
                    kwargs["queryset"] = Product.objects.none()
        formfield = super().formfield_for_foreignkey(db_field, request, **kwargs)
        if db_field.name == 'product':
            formfield.widget.attrs['style'] = 'width: 300px;'
        return formfield

    def _lock_reason(self, po):
        if po is None:
            return None
        if po.status == 'Cancelled':
            return "ใบสั่งซื้อถูกยกเลิก"
        if po.status == 'Completed':
            return "ใบสั่งซื้อปิดงานแล้ว"
        if po.payment_status == 'Paid':
            return "จ่ายเงินครบแล้ว"
        return None

    def has_add_permission(self, request, obj=None):
        if self._lock_reason(obj): return False
        return super().has_add_permission(request, obj)

    def has_change_permission(self, request, obj=None):
        if self._lock_reason(obj): return False
        return super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if self._lock_reason(obj): return False
        return super().has_delete_permission(request, obj)

class SalesItemInline(UnfoldTabularInline):
    model = SalesItem
    autocomplete_fields = ['barcode_obj', 'product', 'bom']
    extra = 1

    # 🎯 กล่อง "ราคาขาย" แสดงแค่ ~5 หลักก็พอ ไม่ต้องยืดเต็มคอลัมน์
    formfield_overrides = {
        models.DecimalField: {'widget': forms.NumberInput(attrs={'style': 'width: 90px;'})},
    }

    def _is_locked(self, so):
        return so is not None and so.status in ('Completed', 'Cancelled')

    def has_add_permission(self, request, obj=None):
        if self._is_locked(obj): return False
        return super().has_add_permission(request, obj)

    def has_change_permission(self, request, obj=None):
        if self._is_locked(obj): return False
        return super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if self._is_locked(obj): return False
        return super().has_delete_permission(request, obj)

    # product ต้องเอาออกจาก readonly_fields เพื่อให้เปรมเลือกเองได้กรณีไม่มีบาร์โค้ด
    readonly_fields = ('quantity_ordered', 'get_unit_name_display','get_total_display')

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('product', 'barcode_obj', 'bom')

    # 1. เรียงลำดับคอลัมน์จากซ้ายไปขวา
    fields = [
        'barcode_obj', 
        'product', 
        'quantity_unit',
        'get_unit_name_display',
        'quantity_ordered', 
        'sale_price',        # ✅ ใส่ตรงนี้เพื่อให้ "แก้ไขได้" (ห้ามใส่ใน readonly_fields)
        'get_total_display', # 🔒 ใส่ตรงนี้เพื่อโชว์ผลลัพธ์ (ต้องใส่ใน readonly_fields ด้วย)
        'bom',               # เลือกสูตรผลิต (ระบบเลือกให้อัตโนมัติในเบื้องต้น)
        'auto_produce',      # 🔘 Checkbox อยู่ท้ายสุดตามที่เปรมต้องการ
    ]
    def get_unit_name_display(self, obj):
        if obj.barcode_obj and obj.barcode_obj.conversion_factor > 1:
            return obj.barcode_obj.unit_name
        return "ชิ้น(ปกติ)"
    get_unit_name_display.short_description = "หน่วยขาย"

    # 3. สร้างฟังก์ชันคำนวณราคารวม (Quantity * Sale Price)
    def get_total_display(self, obj):
        if obj.quantity_ordered and obj.sale_price:
            total = obj.quantity_ordered * obj.sale_price
            return f"{total:,.2f}"
        return "0.00"
    get_total_display.short_description = "ราคารวม"

class SalesDeliveryLogForm(forms.ModelForm):
    """
    ใช้ text input สำหรับรหัสบาร์โค้ด แทน dropdown
    - แถวใหม่: กรอก barcode code → validate server-side → บันทึก FK
    - แถวที่ save แล้ว: แสดง code แบบ readonly ห้ามแก้
    """
    barcode_code = forms.CharField(
        label='บาร์โค้ด',
        required=False,
        widget=forms.TextInput(attrs={
            'placeholder': 'กรอกรหัสบาร์โค้ด',
            'autocomplete': 'off',
            'class': 'barcode-code-input',
            'style': 'width: 180px;',
        })
    )

    class Meta:
        model = SalesDeliveryLog
        exclude = ['barcode_obj']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance.pk and self.instance.barcode_obj:
            # row ที่ save แล้ว → แสดง code เดิม + readonly
            self.fields['barcode_code'].initial = self.instance.barcode_obj.code
            self.fields['barcode_code'].widget.attrs.update({
                'readonly': True,
                'style': 'width: 180px; background:#f3f4f6; color:#374151;',
            })

    def clean(self):
        cleaned = super().clean()
        # row ที่ save แล้ว → ใช้ barcode เดิม ไม่ต้องแปลง
        if self.instance.pk and self.instance.barcode_obj:
            self._resolved_barcode = self.instance.barcode_obj
            return cleaned

        code = (cleaned.get('barcode_code') or '').strip()
        if not code:
            self._resolved_barcode = None
            return cleaned

        try:
            barcode = ProductBarcode.objects.get(code=code)
        except ProductBarcode.DoesNotExist:
            self.add_error('barcode_code', 'ไม่พบบาร์โค้ดนี้ในระบบ')
            self._resolved_barcode = None
            return cleaned

        # ตรวจว่า barcode นี้อยู่ใน SO ด้วย (ถ้า so_id ถูก inject มา)
        so_id = getattr(self.__class__, '_so_id', None)
        if so_id:
            in_so = SalesItem.objects.filter(
                sales_order_id=so_id, barcode_obj=barcode
            ).exists()
            if not in_so:
                self.add_error('barcode_code', 'บาร์โค้ดนี้ไม่อยู่ในรายการสั่งขายนี้')
                self._resolved_barcode = None
                return cleaned

        self._resolved_barcode = barcode
        return cleaned

    def save(self, commit=True):
        instance = super().save(commit=False)
        barcode = getattr(self, '_resolved_barcode', None)
        if barcode:
            instance.barcode_obj = barcode
        elif self.instance.pk and self.instance.barcode_obj:
            instance.barcode_obj = self.instance.barcode_obj
        if commit:
            instance.save()
            self.save_m2m()
        return instance


class SalesDeliveryLogInline(UnfoldTabularInline):
    # 🎯 แถวใหม่สร้างผ่าน "checklist ส่งของ" ที่หัวเอกสารเท่านั้นแล้ว (ดู SalesOrderAdmin.ship_batch_view)
    # ไม่ใช้ปุ่ม "Add another" ของ inline นี้อีกต่อไป — เพราะ Django formset ปฏิบัติกับแถวที่เพิ่ม
    # มาแบบ dynamic (ไม่ได้มีอยู่ตั้งแต่หน้าโหลดครั้งแรก) เป็น "extra form" เสมอ ไม่สนใจค่า -id ที่
    # ส่งมาเลย ต่อให้ auto-save ผูก id ถูกต้องแค่ไหนก็ตาม พอกด Save ทั้งหน้า Django ก็ยังสร้างแถวใหม่
    # ซ้ำกับที่ auto-save สร้างไว้แล้วอยู่ดี (ไล่ debug กับเปรมจนเจอ root cause นี้) — extra=0 +
    # has_add_permission=False ตัดปัญหานี้ทั้งกระบวนตั้งแต่ต้นทาง ส่วนแก้ไข/ลบแถวที่มีอยู่แล้วยังทำ
    # ได้ปกติ (initial form ตัวจริง Django query instance จาก -id ถูกต้องอยู่แล้ว ไม่มีปัญหานี้)
    model = SalesDeliveryLog
    form = SalesDeliveryLogForm
    extra = 0
    fields = ('barcode_code', 'shipping_no', 'quantity_shipped', 'user', 'notes', 'shipped_date')
    readonly_fields = ('user',)
    formfield_overrides = {
        models.CharField: {'widget': TextInput(attrs={'style': 'width: 120px;', 'placeholder': 'เลขใบส่งของ'})},
        models.TextField: {'widget': TextInput(attrs={'style': 'width: 200px;', 'placeholder': 'หมายเหตุ'})},
        # 🎯 กล่อง "จำนวน" ให้ยาวพอสำหรับ 7 หลัก เท่ากับตอนเปิดรายการสั่งขาย
        models.PositiveIntegerField: {'widget': forms.NumberInput(attrs={'style': 'width: 110px;'})},
    }

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('barcode_obj', 'user')

    def get_formset(self, request, obj=None, **kwargs):
        """Inject so_id เข้า form class เพื่อ validate barcode ใน SO"""
        so_id = obj.pk if obj else None
        DynamicForm = type('SalesDeliveryLogForm', (SalesDeliveryLogForm,), {'_so_id': so_id})
        kwargs['form'] = DynamicForm
        return super().get_formset(request, obj, **kwargs)

    def _lock_reason(self, so):
        if so is None:
            return None
        if so.status == 'Cancelled':
            return "ใบสั่งขายถูกยกเลิก"
        if so.status == 'Completed':
            return "ใบสั่งขายปิดงานแล้ว"
        if so.delivery_logs.filter(
            Q(is_revenue_confirmed=True) | Q(is_dc_confirmed=True) | Q(is_rebate_confirmed=True)
        ).exists():
            return "มีรายการที่ยืนยันใน C6 แล้ว"
        return None

    def has_add_permission(self, request, obj=None):
        # เพิ่มแถวใหม่ผ่าน checklist ส่งของที่หัวเอกสารเท่านั้น (ดูคอมเมนต์บนคลาส)
        return False

    def has_change_permission(self, request, obj=None):
        if self._lock_reason(obj): return False
        return super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if self._lock_reason(obj): return False
        return super().has_delete_permission(request, obj)

    def get_barcode_display(self, obj):
        return obj.barcode_obj.code if obj and obj.barcode_obj else '-'
    get_barcode_display.short_description = 'บาร์โค้ด'

    def get_fields(self, request, obj=None):
        if self._lock_reason(obj):
            return ('get_barcode_display', 'shipping_no', 'quantity_shipped', 'user', 'notes', 'shipped_date')
        return ('barcode_code', 'shipping_no', 'quantity_shipped', 'user', 'notes', 'shipped_date')

    def get_readonly_fields(self, request, obj=None):
        if self._lock_reason(obj):
            return ('get_barcode_display', 'shipping_no', 'quantity_shipped', 'user', 'notes', 'shipped_date')
        return ('user',)

class ProductionLogInline(UnfoldTabularInline):
    model = ProductionLog
    extra = 1
    fields = ('quantity_finished', 'user','notes', 'finished_date')
    readonly_fields = ('user', 'finished_date')
    formfield_overrides = {
        models.CharField: {'widget': TextInput(attrs={'style': 'width: 120px;', 'placeholder': 'เลขใบส่งของ'})},
        models.TextField: {'widget': TextInput(attrs={'style': 'width: 200px;', 'placeholder': 'หมายเหตุ'})},
    }

# --- Helper ---
def color_diff(diff):
    color = "green" if diff >= 0 else "red"
    prefix = "+" if diff > 0 else ""
    return format_html('<span style="color: {}; font-weight: bold;">{}{}{}</span>', color, prefix, diff, " ชิ้น" if diff != 0 else "")

# --- Admin Registrations ---

@admin.register(Supplier)
class SupplierAdmin(DetailedHistoryMixin, DocumentLockMixin, UnfoldModelAdmin):
    list_display = ('company_name', 'contact_person', 'type')
    search_fields = ('company_name', 'contact_person', 'supplier_code')
    inlines = [SupplierProductInline]

class ProductBarcodeAdmin(ExportToExcelMixin, UnfoldModelAdmin):
    # 🎯 ตัวนี้แหละคือ "หัวใจ" ที่จะแก้ Error E039
    search_fields = ['code', 'product__name','product__tags__name']
    list_display = ('code', 'product', 'conversion_factor', 'unit_name', 'get_forecast_stock')
    list_filter = (
        ('product__tags', AutocompleteSelectMultipleFilter), # กรองตามกลุ่มสินค้าที่หน้า A4
    )
    list_filter_submit = True
    actions = ['export_to_excel']

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('product')


class ProductStockInline(UnfoldTabularInline):
    # โชว์เฉพาะคลังที่ไม่ใช่คลังหลัก (ดู get_inline_instances ใน WarehouseAdmin)
    # สต๊อกตรงนี้แก้ไข/เพิ่มลบโดยตรงไม่ได้ ต้องผ่านเอกสารโอนย้ายคลังเท่านั้น
    model = ProductStock
    extra = 0
    can_delete = False
    fields = ('code_display', 'name_display', 'quantity', 'buy_price_display', 'sale_price_display')
    readonly_fields = fields

    def has_add_permission(self, request, obj=None):
        return False

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('product').filter(quantity__gt=0)

    @admin.display(description="รหัส")
    def code_display(self, obj):
        return obj.product.latest_barcode

    @admin.display(description="ชื่อสินค้า")
    def name_display(self, obj):
        return obj.product.name

    @admin.display(description="ต้นทุน")
    def buy_price_display(self, obj):
        return obj.product.buy_price

    @admin.display(description="ราคาขาย")
    def sale_price_display(self, obj):
        return obj.product.sale_price


@admin.register(Warehouse)
class WarehouseAdmin(UnfoldModelAdmin):
    list_display = ('name', 'type', 'is_default')
    list_filter = ('type',)
    search_fields = ('name',)
    inlines = [ProductStockInline]

    def get_inline_instances(self, request, obj=None):
        # คลังหลักใช้ Product.stock_quantity ตรงๆ ไม่มีแถวใน ProductStock เลย จึงไม่ต้องโชว์ inline
        if obj is None or obj.is_default:
            return []
        return super().get_inline_instances(request, obj)


@admin.register(StockTransfer)
class StockTransferAdmin(UnfoldModelAdmin):
    list_display = ('transfer_number', 'transfer_date', 'product', 'quantity', 'from_warehouse', 'to_warehouse', 'created_by')
    list_filter = ('from_warehouse', 'to_warehouse', 'transfer_date')
    search_fields = ('transfer_number', 'product__name', 'product__barcodes__code')
    autocomplete_fields = ['product']
    readonly_fields = ('transfer_number',)
    # ปิด action "ลบที่เลือก" เพราะเป็น bulk queryset.delete() ที่ไม่เรียก StockTransfer.delete()
    # ต่อรายการ (จะไม่คืนสต๊อกให้ทำให้ยอดคลังเพี้ยน) ต้องลบทีละใบผ่านหน้ายืนยันลบเท่านั้น
    actions = None

    def get_readonly_fields(self, request, obj=None):
        ro = list(self.readonly_fields)
        if obj is not None:
            # ล็อกฟิลด์หลักหลังบันทึกแล้ว เพราะการแก้ไขจะไม่ไปปรับสต๊อกที่ตัดไปแล้วให้อัตโนมัติ
            ro += ['product', 'quantity', 'from_warehouse', 'to_warehouse']
        return ro

    def save_model(self, request, obj, form, change):
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)


def _normalize_history_date(d):
    """แปลงวันที่ (DateField หรือ DateTimeField หรือ None) ให้เป็น datetime แบบมี timezone เดียวกันหมด
    เพื่อให้เรียงลำดับ/แสดงผลในตารางประวัติได้โดยไม่พังตอนเทียบ date กับ datetime"""
    if d is None:
        return timezone.make_aware(datetime.datetime.min, timezone.get_default_timezone())
    if isinstance(d, datetime.datetime):
        return timezone.make_aware(d, timezone.get_default_timezone()) if timezone.is_naive(d) else d
    return timezone.make_aware(datetime.datetime.combine(d, datetime.time.min), timezone.get_default_timezone())


def build_product_history_rows(product):
    """
    รวมประวัติความเคลื่อนไหวทั้งหมดของสินค้าชิ้นนี้เป็นรายการเดียว:
    ซื้อเข้า (PO), ขายออก (SO), ผลิตได้/ถูกใช้ผลิต (PD), ยกเลิก PO/SO, และแก้ไขข้อมูลสินค้า
    เรียงจากล่าสุดไปเก่าสุด
    """
    rows = []

    # 1. ซื้อเข้า — รับของตาม PO
    receipts = list(
        PurchaseReceiptLog.objects.filter(product=product)
        .select_related('purchase_order', 'purchase_order__supplier', 'barcode_obj')
    )
    po_ids = [r.purchase_order_id for r in receipts]
    po_price_map = {
        pi.purchase_order_id: pi.unit_price
        for pi in PurchaseItem.objects.filter(purchase_order_id__in=po_ids, product=product)
    }
    for r in receipts:
        unit_price = po_price_map.get(r.purchase_order_id) or Decimal('0')
        # unit_price เป็นราคาต่อชิ้น (หน่วยหลัก) เสมอ — ต้องแปลง quantity_received เป็นชิ้นก่อนคูณ
        factor = getattr(r.barcode_obj, 'conversion_factor', 1) or 1
        qty_pieces = r.quantity_received * factor
        rows.append({
            'date': r.received_date,
            'type': 'purchase',
            'type_label': '🛒 ซื้อเข้า (รับของ)',
            'ref_number': r.purchase_order.po_number,
            'ref_url': reverse('admin:stocks_purchaseorder_change', args=[r.purchase_order_id]),
            'quantity': qty_pieces,
            'unit_price': unit_price,
            'total_value': unit_price * qty_pieces,
            'party': str(r.purchase_order.supplier) if r.purchase_order.supplier_id else '-',
            'note': r.notes,
        })

    # 2. ขายออก — ส่งของตาม SO
    deliveries = (
        SalesDeliveryLog.objects.filter(product=product)
        .select_related('sales_order', 'sales_order__customer')
    )
    for d in deliveries:
        unit_price = (d.shipment_value / d.quantity_shipped) if d.quantity_shipped else Decimal('0')
        rows.append({
            'date': d.shipped_date,
            'type': 'sale',
            'type_label': '📤 ขายออก (ส่งของ)',
            'ref_number': d.sales_order.so_number,
            'ref_url': reverse('admin:stocks_salesorder_change', args=[d.sales_order_id]),
            # ✅ ส่งของ = สต๊อกออก ต้องเป็นค่าลบ (เทมเพลตใช้เครื่องหมายตัดสีแดง/เขียว)
            'quantity': -d.quantity_shipped,
            'unit_price': unit_price,
            'total_value': -d.shipment_value,
            'party': str(d.sales_order.customer) if d.sales_order.customer_id else '-',
            'note': d.notes,
        })

    # 3. ผลิตได้ — สินค้าสำเร็จรูปเข้าสต็อกจากใบ PD (เฉพาะกรณีสินค้านี้คือสินค้าที่ผลิต)
    output_logs = (
        ProductionLog.objects.filter(production_order__product=product)
        .select_related('production_order', 'production_order__bom')
    )
    for log in output_logs:
        pd_order = log.production_order
        unit_cost = pd_order.bom.total_cost if pd_order.bom_id else (product.buy_price or Decimal('0'))
        rows.append({
            'date': log.finished_date,
            'type': 'production_out',
            'type_label': '🏭 ผลิตได้ (เข้าสต็อก)',
            'ref_number': pd_order.pd_number,
            'ref_url': reverse('admin:stocks_productionorder_change', args=[pd_order.pk]),
            'quantity': log.quantity_finished,
            'unit_price': unit_cost,
            'total_value': unit_cost * log.quantity_finished,
            'party': '-',
            'note': log.notes,
        })

    # 4. ถูกใช้ผลิต — สินค้านี้ถูกตัดสต็อกไปเป็นวัตถุดิบ/แพ็คเกจของใบ PD อื่น
    usage_logs = (
        ProductionLog.objects.filter(production_order__material_usages__raw_material=product)
        .select_related('production_order', 'production_order__product')
        .prefetch_related('production_order__material_usages')
        .distinct()
    )
    for log in usage_logs:
        pd_order = log.production_order
        if not pd_order.quantity_planned:
            continue
        usage = next(
            (u for u in pd_order.material_usages.all() if u.raw_material_id == product.pk), None
        )
        if not usage:
            continue
        ratio = Decimal(str(log.quantity_finished)) / Decimal(str(pd_order.quantity_planned))
        used_qty = usage.actual_qty_to_use * ratio
        unit_cost = product.buy_price or Decimal('0')
        rows.append({
            'date': log.finished_date,
            'type': 'production_use',
            'type_label': '🧪 ถูกใช้ผลิต (ตัดสต็อก)',
            'ref_number': pd_order.pd_number,
            'ref_url': reverse('admin:stocks_productionorder_change', args=[pd_order.pk]),
            'quantity': -used_qty,
            'unit_price': unit_cost,
            'total_value': -(used_qty * unit_cost),
            'party': '-',
            'note': f"ใช้ผลิต {pd_order.product.name}" if pd_order.product_id else '',
        })

    # 5. ยกเลิกใบสั่งซื้อ
    cancelled_pos = (
        PurchaseOrder.objects.filter(status='Cancelled', items__product=product)
        .distinct().select_related('supplier')
    )
    po_item_map = {
        pi.purchase_order_id: pi
        for pi in PurchaseItem.objects.filter(purchase_order__in=cancelled_pos, product=product)
    }
    for po in cancelled_pos:
        item = po_item_map.get(po.pk)
        qty = item.quantity_ordered if item else 0
        price = item.unit_price if item else Decimal('0')
        rows.append({
            'date': po.cancelled_at or po.order_date,
            'type': 'po_cancelled',
            'type_label': '❌ ยกเลิกใบสั่งซื้อ',
            'ref_number': po.po_number,
            'ref_url': reverse('admin:stocks_purchaseorder_change', args=[po.pk]),
            'quantity': qty,
            'unit_price': price,
            'total_value': qty * price,
            'party': str(po.supplier) if po.supplier_id else '-',
            'note': 'ยกเลิกใบสั่งซื้อ',
        })

    # 6. ยกเลิกใบสั่งขาย
    cancelled_sos = (
        SalesOrder.objects.filter(status='Cancelled', items__product=product)
        .distinct().select_related('customer')
    )
    so_item_map = {
        si.sales_order_id: si
        for si in SalesItem.objects.filter(sales_order__in=cancelled_sos, product=product)
    }
    for so in cancelled_sos:
        item = so_item_map.get(so.pk)
        qty = item.quantity_ordered if item else 0
        price = item.sale_price if item else Decimal('0')
        rows.append({
            'date': so.cancelled_at or so.order_date,
            'type': 'so_cancelled',
            'type_label': '❌ ยกเลิกใบสั่งขาย',
            'ref_number': so.so_number,
            'ref_url': reverse('admin:stocks_salesorder_change', args=[so.pk]),
            'quantity': qty,
            'unit_price': price,
            'total_value': qty * price,
            'party': str(so.customer) if so.customer_id else '-',
            'note': 'ยกเลิกใบสั่งขาย',
        })

    # 7. แก้ไขข้อมูลสินค้า (ของเดิม — ประวัติการเปลี่ยนแปลงรายละเอียดข้อมูล)
    content_type = ContentType.objects.get_for_model(Product)
    edits = LogEntry.objects.filter(
        content_type=content_type, object_id=str(product.pk)
    ).select_related('user')
    for e in edits:
        who = e.user.get_full_name() or e.user.username if e.user_id else '-'
        rows.append({
            'date': e.action_time,
            'type': 'edit',
            'type_label': '📝 แก้ไขข้อมูลสินค้า',
            'ref_number': '-',
            'ref_url': None,
            'quantity': None,
            'unit_price': None,
            'total_value': None,
            'party': who,
            'note': e.get_change_message(),
        })

    for row in rows:
        row['date'] = _normalize_history_date(row['date'])
    rows.sort(key=lambda r: r['date'], reverse=True)
    return rows


# ── C0. คาดการณ์ Stock — คำนวณจากอัตราการใช้จริงในอดีต (ยอดขาย + ยอดใช้ผลิต) ──
# หลักการ: แต่ละคอลัมน์ (weekly/2weekly/monthly/3monthly) ดูข้อมูลย้อนหลัง "4 เท่า" ของช่วงตัวเอง
# (weekly ย้อน 4 สัปดาห์, monthly ย้อน 4 เดือน, ...) แล้วหารด้วยจำนวนช่วงที่มีข้อมูลจริง — ถ้าประวัติ
# สินค้าไม่ยาวพอ (สินค้าใหม่) ก็หารด้วยจำนวนช่วงเท่าที่มีจริง หรือ "ขยายสัดส่วน" ขึ้นถ้ามีไม่ถึง 1 ช่วง
_FORECAST_MAX_LOOKBACK_DAYS = 90 * 4  # เผื่อคอลัมน์ 3monthly (ย้อน 4x = 360 วัน) ซึ่งเป็นช่วงยาวสุด

def _historical_usage_events(product, today):
    """คืน list ของ (date, จำนวนที่ถูกใช้เป็นชิ้น) รวมทั้งยอดขายและยอดใช้ผลิต ย้อนหลังสูงสุด 360 วัน"""
    cutoff = today - datetime.timedelta(days=_FORECAST_MAX_LOOKBACK_DAYS)
    events = []

    sales = SalesDeliveryLog.objects.filter(
        product=product, shipped_date__date__gte=cutoff
    ).select_related('barcode_obj')
    for s in sales:
        factor = getattr(s.barcode_obj, 'conversion_factor', 1) or 1
        events.append((s.shipped_date.date(), s.quantity_shipped * factor))

    # ยอดใช้ผลิต — ใช้ pattern เดียวกับ build_product_history_rows (usage_logs block ด้านบน)
    usage_logs = (
        ProductionLog.objects.filter(
            production_order__material_usages__raw_material=product,
            finished_date__date__gte=cutoff,
        )
        .select_related('production_order')
        .prefetch_related('production_order__material_usages')
        .distinct()
    )
    for log in usage_logs:
        po = log.production_order
        if not po.quantity_planned:
            continue
        usage = next((u for u in po.material_usages.all() if u.raw_material_id == product.pk), None)
        if not usage:
            continue
        ratio = Decimal(str(log.quantity_finished)) / Decimal(str(po.quantity_planned))
        events.append((log.finished_date.date(), usage.actual_qty_to_use * ratio))

    return events


def _forecast_for_period(events, today, period_days):
    """หาค่าคาดการณ์เฉลี่ยต่อ 1 ช่วง (period_days วัน) จากประวัติย้อนหลัง 4 เท่าของช่วงนั้น"""
    if not events:
        return 0
    lookback_target = period_days * 4
    earliest = min(e[0] for e in events)
    days_available = min((today - earliest).days, lookback_target)
    if days_available <= 0:
        return 0
    window_start = today - datetime.timedelta(days=days_available)
    total = sum(qty for d, qty in events if d >= window_start)
    periods_covered = Decimal(days_available) / Decimal(period_days)
    if periods_covered <= 0:
        return 0
    return int(round(total / periods_covered))


@admin.register(Product)
class ProductAdmin(DetailedHistoryMixin, ExportToExcelMixin, DocumentLockMixin, UnfoldModelAdmin):
    list_display = ('name', 'display_tags', 'get_latest_barcode', 'get_buy_price_display', 'get_production_cost', 'sale_price', 'stock_quantity', 'min_stock', 'unit','get_total_stock_value', 'has_bom', 'created_by')
    list_filter = (
        ('category', AutocompleteSelectMultipleFilter),
        ('is_product', BooleanRadioFilter),
        ('tags', AutocompleteSelectMultipleFilter),
        ('has_bom', BooleanRadioFilter),
        ('suppliers', AutocompleteSelectMultipleFilter),
    )
    list_filter_submit = True
    search_fields = ('name', 'barcodes__code','tags__name')
    inlines = [ProductBarcodeInline, ProductSupplierInline,PendingPurchaseInline, PendingProductionInline, PendingSaleInline]
    readonly_fields = ('created_by', 'updated_by', 'created_at', 'updated_at', 'auto_cost', 'buy_price', 'cost_source', 'sale_price')
    actions = ['export_to_excel', 'auto_fill_cost_price']
    fieldsets = (
        (None, {
            'fields': ('name', 'category', 'is_product', 'tags', 'has_bom', 'unit', 'stock_quantity', 'min_stock', 'production_lead_time', 'delivery_lead_time')
        }),
        ('💰 ต้นทุน & ราคาขาย', {
            'fields': (('auto_cost', 'manual_buy_price'), ('buy_price', 'cost_source'), ('sale_price',)),
            'description': (
                '<b>ต้นทุนอัตโนมัติ</b>: คำนวณจาก Supplier ราคาสูงสุด +15% (อ่านอย่างเดียว อัปเดตเองเมื่อบันทึก) &nbsp;|&nbsp; '
                '<b>ต้นทุนกำหนดเอง</b>: กรอกเพื่อ override — ถ้ามีค่า (&gt;0) ระบบจะใช้ค่านี้แทนอัตโนมัติเสมอ<br>'
                '<b>ต้นทุนที่ใช้จริง</b>: ถ้าไม่มีต้นทุนกำหนดเอง ระบบจะเทียบ "ต้นทุน BOM เฉลี่ย" (ถ้าติ๊กมี BOM) กับ "ต้นทุนอัตโนมัติ (Supplier+15%)" '
                'แล้วเลือกค่าที่ <b>สูงกว่า</b> มาใช้เป็นต้นทุนจริง (อ่านอย่างเดียว) — ดูที่มาได้จากช่อง "ที่มาต้นทุน" ด้านข้าง'
            ),
        }),
        ('ข้อมูลระบบ', {
            'classes': ('collapse',),
            'fields': ('created_by', 'updated_by', 'created_at', 'updated_at')
        }),
    )

    @admin.action(description="💰 คำนวณต้นทุน/ราคาขายอัตโนมัติ (อัพเดท)")
    def auto_fill_cost_price(self, request, queryset):
        updated_count = 0
        skipped_count = 0

        for obj in queryset.select_related('category').prefetch_related('product_suppliers'):
            if obj.recalc_cost_and_price():
                updated_count += 1
            else:
                skipped_count += 1

        self.message_user(
            request,
            f"คำนวณอัตโนมัติเสร็จ: อัปเดต {updated_count} รายการ, ข้าม {skipped_count} รายการ (ค่าที่คำนวณได้เท่ากับค่าเดิม หรือไม่มี category)",
            messages.SUCCESS,
        )

    # ✅ ใช้ตัวนี้แทน filter_horizontal หรือ filter_vertical ค่ะ
    autocomplete_fields = ['tags']

    # --- ให้การค้นหา ใช้ รูปแบบ และ หรือ ได้ ---
    def get_search_results(self, request, queryset, search_term):
        # 🎯 1. จัดการระบบ OR (|) ก่อน
        if '|' in search_term:
            import operator
            from django.db.models import Q
            from functools import reduce
            parts = [p.strip() for p in search_term.split('|') if p.strip()]
            q_objects = []
            for part in parts:
                q_part = Q()
                for field in self.search_fields:
                    q_part |= Q(**{f"{field}__icontains": part})
                q_objects.append(q_part)
            queryset = queryset.filter(reduce(operator.or_, q_objects)).distinct()
            use_distinct = False
        else:
            # ถ้าไม่มี | ให้ค้นหาปกติ
            queryset, use_distinct = super().get_search_results(request, queryset, search_term)

        # 🎯 2. จัดการระบบ Autocomplete สำหรับ PO (ล็อคตาม Supplier)
        if 'autocomplete' in request.path:
            from django.db.models import Q
            supplier_id = request.GET.get('supplier_id', '').strip()

            # ✅ วิธีหลัก: JS (purchase_order_supplier_filter.js) ส่ง supplier_id ของ supplier
            # ที่กำลังเลือกอยู่ในหน้ามาให้ตรงๆ ทำงานได้ทั้งหน้า "เพิ่มใหม่" (ยังไม่มี PO id)
            # และหน้าแก้ไข — ต่างจากวิธีเดิมที่ parse HTTP_REFERER ซึ่งใช้ไม่ได้เลยตอนเพิ่มใหม่
            if supplier_id:
                queryset = queryset.filter(
                    Q(product_suppliers__supplier_id=supplier_id) | Q(is_product=False)
                )
            else:
                # ↩️ fallback เดิม เผื่อ JS โหลดไม่ทัน/ถูกปิด — ใช้ได้เฉพาะหน้าแก้ไข PO ที่มีอยู่แล้ว
                referer = request.META.get('HTTP_REFERER', '')
                if 'purchaseorder' in referer:
                    import re
                    match = re.search(r'purchaseorder/(\d+)/change/', referer)
                    if match:
                        po_id = match.group(1)
                        from .models import PurchaseOrder
                        try:
                            po = PurchaseOrder.objects.get(pk=po_id)
                            if po.supplier:
                                queryset = queryset.filter(
                                    Q(product_suppliers__supplier=po.supplier) | Q(is_product=False)
                                )
                        except PurchaseOrder.DoesNotExist: pass

        return queryset, use_distinct


    # 🎯 2. ปรับ CSS สำหรับแนวตั้งโดยเฉพาะ (เน้นความคลีน)
    def render_change_form(self, request, context, add=False, change=False, form_url='', obj=None):
        return super().render_change_form(request, context, add, change, form_url, obj)

    # ✅ 3. ตัวโชว์ Tag ในหน้ารวมรายการสินค้า (โค้ดเปรมดีอยู่แล้วครับ)
    def display_tags(self, obj):
        tags = obj.tags.all()
        if not tags: return "-"
        html = "".join([
            f'<span style="background:{t.color}; color:white; padding:2px 8px; '
            f'border-radius:12px; margin-right:4px; font-size:11px; font-weight:bold; display:inline-block; margin-bottom:2px;">'
            f'{t.name}</span>' for t in tags
        ])
        return mark_safe(html)
    display_tags.short_description = "แท็ก"

    @admin.display(description='ราคาทุน (ใช้จริง)', ordering='buy_price')
    def get_buy_price_display(self, obj):
        price_str = f"{obj.buy_price:,.2f}"
        if obj.cost_source == 'bom':
            return format_html('{} <span style="color:#888;font-size:11px;">(จาก BOM)</span>', price_str)
        return price_str

    # 🛠️ จุดที่แก้เพื่อเลิกล่ม: ดัก Error การจัดรูปแบบตัวเลข
    def get_production_cost(self, obj):
        try:
            count = getattr(obj, 'bom_count', 0)
            avg_cost = getattr(obj, 'production_cost_avg', 0)

            # 🔥 ไม้ตาย: ล้างความเป็น SafeString ออกให้หมดก่อนแปลงเป็น float
            clean_str = str(avg_cost).replace(',', '').strip()
            # ถ้าเป็น HTML มา (มี <span...) ให้ตัดเอาเฉพาะตัวเลข
            if '<' in clean_str:
                import re
                clean_str = re.sub('<[^<]+?>', '', clean_str)
            
            try:
                price_val = float(clean_str)
            except:
                price_val = 0.0

            if count and count > 0:
                # จัดรูปแบบทศนิยมข้างนอก format_html เพื่อความปลอดภัย
                display_num = "{:,.2f}".format(price_val)
                return format_html('<b style="color: #28a745;">{}</b> <span style="color: #666;">({})</span>', display_num, count)
            
            if getattr(obj, 'has_bom', False):
                return mark_safe('<span style="color: #999;">0.00 (0)</span>')
        except Exception as e:
            return f"Err: {str(e)[:20]}"
        return "-"

    get_production_cost.short_description = "ต้นทุนBOMเฉลี่ย"

    def get_latest_barcode(self, obj):
        # ดึงจาก property ที่เราเขียนไว้ใน models
        return obj.latest_barcode
    get_latest_barcode.short_description = "บาร์โค้ด (ล่าสุด)"

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        queryset = queryset.select_related('created_by').prefetch_related(
            'tags', 'barcodes', 'bom_formulas__ingredients__material'
        ).annotate(
            _total_stock_value=ExpressionWrapper(
                F('stock_quantity') * F('sale_price'),
                output_field=DecimalField()
            )
        )
        return queryset
    # 3. สร้างฟังก์ชันแสดงผล (ใน ProductAdmin)
    @admin.display(description='มูลค่า', ordering='-_total_stock_value')
    def get_total_stock_value(self, obj):
        # ✅ ใช้ int() เพื่อปัดเศษทศนิยมทิ้ง และใช้ :, เพื่อใส่คอมมาคั่นหลักพัน
        value = obj._total_stock_value or 0
        return f"{int(value):,}"

    def save_model(self, request, obj, form, change):
        if not change:
            obj.created_by = request.user
        obj.updated_by = request.user
        super().save_model(request, obj, form, change)

    COST_SOURCE_LABELS = {'manual': 'กำหนดเอง', 'bom': 'จาก BOM', 'supplier': 'อัตโนมัติจาก Supplier'}

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        obj = form.instance

        if not obj.pk:
            return

        # อ่านค่าล่าสุดจาก DB (หลัง inline ทุกตัว save เสร็จแล้ว — เผื่อ signal ของ ProductSupplier คำนวณไปแล้วรอบนึง)
        obj.refresh_from_db()
        changes = obj.recalc_cost_and_price()

        auto_filled = []
        if 'auto_cost' in changes:
            auto_filled.append(f"ต้นทุนอัตโนมัติ = {changes['auto_cost']:,.2f}")
        if 'buy_price' in changes:
            src = self.COST_SOURCE_LABELS.get(obj.cost_source, 'อัตโนมัติ')
            auto_filled.append(f"ต้นทุนที่ใช้จริง ({src}) = {changes['buy_price']:,.2f}")
        if 'sale_price' in changes:
            auto_filled.append(f"ราคาขาย = {changes['sale_price']:,.2f}")

        if auto_filled:
            self.message_user(
                request,
                f"คำนวณอัตโนมัติ ({obj.name}): {', '.join(auto_filled)} บาท",
                messages.INFO,
            )

    # ✅ แทนที่หน้า "History" เดิม (ที่โชว์แค่การแก้ไขข้อมูล) ด้วยประวัติแบบเต็ม
    # รวมซื้อเข้า/ขายออก/ผลิต/ยกเลิก เข้าไปด้วย — เรียกจากปุ่ม History เดิมของ Django admin ได้เลย
    def history_view(self, request, object_id, extra_context=None):
        from django.core.exceptions import PermissionDenied
        from django.http import Http404
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        obj = self.get_object(request, object_id)
        if obj is None:
            raise Http404("ไม่พบสินค้านี้")

        rows = build_product_history_rows(obj)

        paginator = Paginator(rows, 50)
        page_number = request.GET.get('p') or 1
        page_obj = paginator.get_page(page_number)

        context = {
            **self.admin_site.each_context(request),
            'title': f"ประวัติสินค้า: {obj.name}",
            'object': obj,
            'opts': self.model._meta,
            'page_obj': page_obj,
            'rows': page_obj.object_list,
            'total_count': paginator.count,
        }
        context.update(extra_context or {})
        return TemplateResponse(request, 'admin/product_history.html', context)

    class Media:
        js = ('js/admin_sum_selected.js',) # เรียกไฟล์ JS มาใช้งาน

@admin.register(BOM)
class BOMAdmin(DocumentLockMixin, UnfoldModelAdmin):
    list_display = ('name', 'product', 'total_cost_display', 'unit', 'production_time', 'created_by')
    list_filter = (('product__category', AutocompleteSelectMultipleFilter),)
    list_filter_submit = True
    autocomplete_fields = ['product']
    search_fields = ['name', 'product__name', 'product__code', 'product__barcodes__code']
    inlines = [BOMIngredientInline]
    readonly_fields = ('created_by', 'updated_by')

    def total_cost_display(self, obj):
        try:
            return f"{float(obj.total_cost):,.2f}"
        except: return "0.00"
    
    def save_model(self, request, obj, form, change):
        if not change: obj.created_by = request.user
        obj.updated_by = request.user
        super().save_model(request, obj, form, change)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "product":
            kwargs["queryset"] = Product.objects.filter(has_bom=True)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    class Media:
        js = ('js/barcode_autofill_generic.js', 'js/bom_ingredient_barcode_filter.js', 'js/bom_name_barcode_autofill.js')

@admin.register(PurchaseOrder)
class PurchaseOrderAdmin(DetailedHistoryMixin, ExportToExcelMixin, DocumentLockMixin, UnfoldModelAdmin):
    list_display = ('po_number', 'supplier', 'order_date', 'status', 'get_diff')
    list_filter = (
        ('status', MultipleChoicesDropdownFilter),
        ('order_date', DjangoDateRangeFilter),
        ('supplier', AutocompleteSelectMultipleFilter),
        ('items__product__tags', PurchaseOrderTagsFilter),
    )
    list_filter_submit = True
    search_fields = ('po_number', 'invoice_no_supplier', 'items__product__name',
    'items__product__barcodes__code', 'supplier__company_name')
    autocomplete_fields = ['supplier']
    inlines = [PurchaseItemInline, PurchaseReceiptLogInline]
    date_hierarchy = 'order_date' # ✅ เพิ่มบรรทัดนี้ค่ะ
    readonly_fields = ('created_by', 'status')

    actions = ['mark_as_completed', 'export_to_excel']

    def get_urls(self):
        custom_urls = [
            path('<int:object_id>/print/', self.admin_site.admin_view(self.print_view), name='stocks_purchaseorder_print'),
            path('<int:object_id>/print-receipt/', self.admin_site.admin_view(self.print_receipt_view), name='stocks_purchaseorder_print_receipt'),
            path('<int:object_id>/unlock/', self.admin_site.admin_view(self.unlock_view), name='stocks_purchaseorder_unlock'),
        ]
        return custom_urls + super().get_urls()

    def unlock_view(self, request, object_id):
        # 🎯 ทำเป็น view แยกต่างหาก (ไม่ยัดปุ่มเข้าไปใน form ของ change_view) เพราะตอนสถานะ
        # Completed ฟอร์มทั้งหน้าจะถูก render แบบ readonly ล้วนๆ (has_change_permission เป็น False
        # ตอน GET) ไม่มี <input> จริงให้ submit เลย ถ้ายัดปุ่ม submit เข้าไปในฟอร์มนั้นจะโดน Django
        # เช็ค required fields (Customer, VAT ฯลฯ) แล้ว error "This field is required." ทั้งที่ไม่ได้
        # จะแก้ field พวกนั้นเลย — เลี่ยงปัญหานี้โดยไม่ผ่าน form validation ของ change_view เลย
        from django.core.exceptions import PermissionDenied
        from django.http import Http404
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        obj = self.get_object(request, object_id)
        if obj is None:
            raise Http404("ไม่พบใบสั่งซื้อนี้")
        if obj.status == 'Completed':
            obj.status = 'Partially Received'
            obj.save(update_fields=['status'])
            self.message_user(request,
                f"🔓 ปลดล็อคใบสั่งซื้อ {obj.po_number} แล้ว แก้ไขรายการรับของได้ตามปกติ — "
                f"ถ้าแก้เสร็จแล้วยอดรับยังครบเหมือนเดิม ระบบจะปิดงานให้อัตโนมัติ")
        return HttpResponseRedirect(reverse('admin:stocks_purchaseorder_change', args=[obj.pk]))

    def print_view(self, request, object_id):
        from django.core.exceptions import PermissionDenied
        from django.http import Http404
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        obj = self.get_object(request, object_id)
        if obj is None:
            raise Http404("ไม่พบใบสั่งซื้อนี้")
        context = {
            **self.admin_site.each_context(request),
            'obj': obj,
            'items': obj.items.all().order_by('id').select_related('product'),
            'title': f"ใบสั่งซื้อ {obj.po_number}",
        }
        return TemplateResponse(request, 'admin/purchase_order_print.html', context)

    def print_receipt_view(self, request, object_id):
        from django.core.exceptions import PermissionDenied
        from django.http import Http404
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        obj = self.get_object(request, object_id)
        if obj is None:
            raise Http404("ไม่พบใบสั่งซื้อนี้")
        context = {
            **self.admin_site.each_context(request),
            'obj': obj,
            'receipts': obj.receipt_logs.all().order_by('received_date', 'id').select_related('product'),
            'title': f"ใบรับสินค้า {obj.po_number}",
        }
        return TemplateResponse(request, 'admin/purchase_receipt_print.html', context)

    @admin.action(description="✅ เปลี่ยนสถานะเป็น: เสร็จงาน/ปิดงาน")
    def mark_as_completed(self, request, queryset):
        queryset.update(status='Completed')
        self.message_user(request, f"ปิดงานสำเร็จ {queryset.count()} รายการแล้วค่ะ")

    def response_change(self, request, obj):
        if "_complete_order" in request.POST:
            obj.status = 'Completed'
            obj.save()
            self.message_user(request, f"ปิดงานใบสั่งซื้อ {obj.po_number} เรียบร้อยแล้ว")
            return HttpResponseRedirect(".")
        return super().response_change(request, obj)

    def render_change_form(self, request, context, add=False, change=False, form_url='', obj=None):
        if obj and change:
            reasons = []
            if obj.status in ('Completed', 'Cancelled'):
                reasons.append(f"สถานะ '{obj.status}'")
            if obj.payment_status == 'Paid':
                reasons.append("จ่ายเงินครบแล้ว")
            if reasons:
                messages.warning(request,
                    f"🔒 เอกสารนี้ถูกล็อค ({', '.join(reasons)}) — ไม่สามารถเพิ่ม/แก้ไข/ลบรายการรับของได้")
        response = super().render_change_form(request, context, add, change, form_url, obj)
        if obj and change:
            items = PurchaseItem.objects.filter(purchase_order=obj).order_by('id').select_related('product')
            items_data = {}
            for item in items:
                if item.product_id:
                    pid = str(item.product_id)
                    remaining = max(0, item.quantity_ordered - item.quantity_received)
                    if pid not in items_data:
                        items_data[pid] = {'name': str(item.product), 'base_remaining': remaining}
                    else:
                        items_data[pid]['base_remaining'] += remaining
            safe_json = json.dumps({'type': 'receipt', 'form_prefix': 'receipt_logs',
                                    'select_field': 'product', 'qty_field': 'quantity_received',
                                    'items': items_data}).replace('</', '<\\/')
            smart_script = f'<script>window.SMART_INLINE_DATA={safe_json};</script>'
            # ⚠️ ต้องฉีดปุ่ม "เสร็จงาน" เข้า response.content ก่อน </body> โดยตรง (แบบเดียวกับ smart_script
            # ด้านบน) ห้ามฝังผ่าน context['title'] เพราะ Unfold ไม่การันตีว่า {{ title }} จะถูก echo ใน body
            # 🎯 #submit-row ของ Unfold อยู่ "นอก" <form> จริง (เป็น sticky bar แยกต่างหาก) ปุ่ม submit ที่ยัด
            # เข้าไปต้องมี attribute form="{model}_form" กำกับด้วย ไม่งั้นกดแล้วไม่มี <form> ให้ผูก (กดแล้ว
            # ไม่เกิดอะไรขึ้นเลย) — ใช้ได้กับปุ่ม "เสร็จงาน" เพราะตอนนั้นฟอร์มยัง render แบบแก้ไขได้ปกติ
            # แต่ใช้กับปุ่ม "ปลดล็อค" ไม่ได้ เพราะตอนสถานะ Completed ทั้งฟอร์ม render แบบ readonly ล้วนๆ
            # (ไม่มี input จริงให้ Customer/VAT ฯลฯ) submit ทั้งฟอร์มแล้วจะโดน required-field error แทน
            # เลยทำปุ่มปลดล็อคเป็นลิงก์ไปหา view แยกต่างหาก (unlock_view) ที่ไม่ผ่านการ validate ฟอร์มเลย
            form_id = f"{self.model._meta.model_name}_form"
            unlock_url = reverse('admin:stocks_purchaseorder_unlock', args=[obj.pk])
            unlock_btn_html = ''
            if obj.status == 'Completed':
                unlock_btn_html = f'<a href="{unlock_url}" style="display:inline-block; background: #f59e0b; color: white; height: 35px; line-height: 35px; margin-right: 10px; border-radius: 4px; border: none; cursor: pointer; padding: 0 20px; font-weight: bold; text-decoration: none;">🔓 ปลดล็อค (Unlock)</a>'
            complete_btn_script = f"""
                <script>
                    django.jQuery(document).ready(function() {{
                        var target = django.jQuery('#submit-row .flex-col-reverse');
                        if (!target.length) {{ target = django.jQuery('#submit-row'); }}
                        var btn = '<input type="submit" form="{form_id}" value="เสร็จงาน (Complete)" name="_complete_order" style="background: #28a745; color: white; height: 35px; margin-right: 10px; border-radius: 4px; border: none; cursor: pointer; padding: 0 20px; font-weight: bold;">';
                        target.prepend(btn);
                        target.prepend('{unlock_btn_html}');
                    }});
                </script>
            """
            # ปุ่มพิมพ์: แปะเข้าแถบหัวข้อ (h2) ของแต่ละ inline group โดยตรง (โครงสร้างเดียวกับที่ยืนยัน
            # แล้วจาก SalesOrderAdmin) มี fallback ไป submit-row ถ้าหา header ไม่เจอ
            print_po_url = reverse('admin:stocks_purchaseorder_print', args=[obj.pk])
            print_receipt_url = reverse('admin:stocks_purchaseorder_print_receipt', args=[obj.pk])
            print_btn_script = f"""
                <script>
                    django.jQuery(document).ready(function() {{
                        function addPrintBtn(headingText, url, bg, fullLabel) {{
                            var $header = django.jQuery('h2[id$="-heading"]:contains("' + headingText + '")').first();
                            if ($header.length) {{
                                var link = '<a href="' + url + '" target="_blank" class="ml-auto" style="display:inline-block; background:' + bg + '; color:white; height:26px; line-height:26px; border-radius:4px; border:none; cursor:pointer; padding:0 14px; font-weight:bold; font-size:13px; text-decoration:none;">🖨️ พิมพ์</a>';
                                $header.append(link);
                            }} else {{
                                var fallback = django.jQuery('#submit-row .flex-col-reverse');
                                if (!fallback.length) {{ fallback = django.jQuery('#submit-row'); }}
                                var link = '<a href="' + url + '" target="_blank" style="display:inline-block; background:' + bg + '; color:white; height:35px; line-height:35px; margin-right:10px; border-radius:4px; border:none; cursor:pointer; padding:0 20px; font-weight:bold; text-decoration:none;">🖨️ ' + fullLabel + '</a>';
                                fallback.prepend(link);
                            }}
                        }}
                        addPrintBtn('Purchase items', '{print_po_url}', '#17a2b8', 'พิมพ์ใบสั่งซื้อ');
                        addPrintBtn('Purchase receipt logs', '{print_receipt_url}', '#6f42c1', 'พิมพ์ใบรับสินค้า');
                    }});
                </script>
            """
            response.render()
            response.content = response.content.replace(
                b'</body>', (smart_script + complete_btn_script + print_btn_script).encode('utf-8') + b'</body>', 1
            )
        return response

    # ✅ แก้ไขตรงนี้: เพื่อให้บันทึก PurchaseReceiptLog ได้
    # ✅ ปรับโครงสร้างให้เหมือน SalesOrderAdmin เป๊ะๆ
    def save_formset(self, request, form, formset, change):
        if formset.model == PurchaseItem:
            # ส่วนของรายการสินค้า (Items) - ทำเหมือน SO
            instances = formset.save(commit=False)
            for instance in instances:
                instance.save()
            formset.save_m2m()
        else:
            # ✅ ส่วนของประวัติการรับของ (ReceiptLog) 
            # ใช้ท่าเดียวกับ SO คือ formset.save()
            # Django จะจัดการเรื่อง "ลบรายการที่ถูกติ๊ก Delete" ให้เอง 100% ครับ
            formset.save()

    # (ฟังก์ชันอื่นคงเดิมได้เลยค่ะ)
    def save_model(self, request, obj, form, change):
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

    def get_queryset(self, request):
        # .distinct() กัน PO ซ้ำแถวเวลา filter ผ่าน items__product__tags (join หลายชั้น)
        return super().get_queryset(request).select_related('supplier').prefetch_related('items', 'receipt_logs').distinct()

    def get_diff(self, obj):
        ordered = sum(i.quantity_ordered for i in obj.items.all())
        # ใช้ยอดสะสมระดับ PurchaseItem (เป็นชิ้นเสมอ) ไม่ใช่ผลรวมดิบจาก receipt_logs
        # เพราะ receipt_logs.quantity_received อาจกรอกเป็นหน่วยบาร์โค้ด (เช่น แพ็ค)
        received = sum(i.quantity_received for i in obj.items.all())
        return color_diff(received - ordered)

    class Media:
        js = ('js/admin_sum_selected.js', 'js/smart_delivery_inline.js', 'js/purchase_order_supplier_filter.js', 'js/purchase_item_price_autofill.js', 'js/barcode_autofill_generic.js')

@admin.register(SalesOrder)
class SalesOrderAdmin(DetailedHistoryMixin, ExportToExcelMixin, DocumentLockMixin, UnfoldModelAdmin):
    list_display = ('so_number', 'customer', 'order_date', 'status', 'vat_percent','get_diff')
    list_filter = (
        ('status', MultipleChoicesDropdownFilter),
        ('order_date', DjangoDateRangeFilter),
        ('customer', AutocompleteSelectMultipleFilter),
        ('items__product__tags', SalesOrderTagsFilter),
    )
    list_filter_submit = True
    search_fields = ('so_number', 'po_no_customer', 'customer__company_name',
        'items__product__barcodes__code')
    autocomplete_fields = ['customer']
    inlines = [SalesItemInline, SalesDeliveryLogInline, SalesPaymentInline]
    readonly_fields = ('created_by', 'status') # ล็อค status ให้ระบบจัดการออโต้
    date_hierarchy = 'order_date' # ✅ เพิ่มบรรทัดนี้ค่ะ
    actions = ['mark_as_completed', 'export_to_excel']

    def get_queryset(self, request):
        # .distinct() กัน SO ซ้ำแถวเวลา filter ผ่าน items__product__tags (join หลายชั้น)
        return super().get_queryset(request).select_related('customer').prefetch_related('items', 'delivery_logs').distinct()

    def get_urls(self):
        custom_urls = [
            path('<int:object_id>/print/', self.admin_site.admin_view(self.print_view), name='stocks_salesorder_print'),
            path('<int:object_id>/print-delivery/', self.admin_site.admin_view(self.print_delivery_view), name='stocks_salesorder_print_delivery'),
            path('<int:object_id>/unlock/', self.admin_site.admin_view(self.unlock_view), name='stocks_salesorder_unlock'),
            path('<int:object_id>/ship/', self.admin_site.admin_view(self.ship_batch_view), name='stocks_salesorder_ship'),
        ]
        return custom_urls + super().get_urls()

    def ship_batch_view(self, request, object_id):
        # 🎯 สร้าง/แก้ไข "รอบส่งของ" ทั้งชุดในคำขอเดียว — ไม่ผ่าน Django formset เลย (นั่นคือต้นตอ
        # บั๊กบันทึกซ้ำเดิม: formset ปฏิบัติกับแถวที่เพิ่มมาแบบ dynamic เป็น "extra form" เสมอ ไม่
        # สนใจค่า -id ที่ส่งมา) ใช้ ORM ธรรมดาสร้าง/แก้ SalesDeliveryLog ทีละแถวผ่าน .save() ปกติ
        # เพื่อให้ side-effect เดิมทำงานครบ (ตัดสต็อก, สะสมยอดใน SalesItem, คำนวณ DC/Rebate/
        # shipment_value ผ่าน sync_dc_rebate_from_contract())
        import datetime as dt_module
        from django.core.exceptions import PermissionDenied
        from django.http import Http404, HttpResponseBadRequest
        from django.utils.dateparse import parse_date
        from django.utils import timezone as tz

        if request.method != 'POST':
            return HttpResponseBadRequest('POST only')
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        obj = self.get_object(request, object_id)
        if obj is None:
            raise Http404("ไม่พบใบสั่งขายนี้")

        action = request.POST.get('action', '')

        def _parse_batch_date_only(raw):
            """วันที่ล้วนๆ ไม่มีเวลา — ใช้หา instance เดิม (WHERE) เท่านั้น"""
            raw = (raw or '').strip()
            return parse_date(raw) if raw else None

        def _combine_date_time(date_raw, hour_raw, minute_raw):
            """รวมช่องวันที่ + ชั่วโมง/นาที (เลือกจาก <select> แยก 2 ช่อง ไม่ใช้
            <input type="time"> เพราะโชว์ AM/PM ตาม locale ของเบราว์เซอร์/OS ผู้ใช้ — บังคับ
            24 ชม. เสมอด้วยการเลือกตัวเลขตรงๆ) ไม่กรอกมา default เป็น 10:00 ตามที่เปรมขอ"""
            d = _parse_batch_date_only(date_raw)
            if not d:
                return None
            try:
                h = int(hour_raw)
                assert 0 <= h <= 23
            except (TypeError, ValueError, AssertionError):
                h = 10
            try:
                mi = int(minute_raw)
                assert 0 <= mi <= 59
            except (TypeError, ValueError, AssertionError):
                mi = 0
            t = dt_module.time(h, mi)
            return tz.make_aware(dt_module.datetime.combine(d, t), tz.get_current_timezone())

        if action == 'create_batch':
            ship_date = _combine_date_time(
                request.POST.get('shipped_date'),
                request.POST.get('shipped_hour'),
                request.POST.get('shipped_minute'),
            )
            if not ship_date:
                messages.error(request, "กรุณากรอกวันที่ส่งของก่อน")
                return HttpResponseRedirect(reverse('admin:stocks_salesorder_change', args=[obj.pk]))

            created_count = 0
            for key, raw_qty in request.POST.items():
                m = re.match(r'^ship_qty_(\d+)$', key)
                if not m:
                    continue
                barcode_id = m.group(1)
                if request.POST.get(f'ship_checked_{barcode_id}') != 'on':
                    continue
                try:
                    qty = int(raw_qty)
                except (TypeError, ValueError):
                    continue
                if qty <= 0:
                    continue
                try:
                    barcode = ProductBarcode.objects.select_related('product').get(pk=barcode_id)
                except ProductBarcode.DoesNotExist:
                    continue
                if not SalesItem.objects.filter(sales_order=obj, barcode_obj=barcode).exists():
                    continue
                log = SalesDeliveryLog(
                    sales_order=obj,
                    barcode_obj=barcode,
                    product=barcode.product,
                    quantity_shipped=qty,
                    shipped_date=ship_date,
                    user=request.user,
                )
                log.save()
                created_count += 1

            if created_count == 0:
                messages.warning(request, "ไม่มีรายการที่ติ๊กไว้ (หรือจำนวนเป็น 0) — ไม่ได้บันทึกอะไรเลย")
                return HttpResponseRedirect(reverse('admin:stocks_salesorder_change', args=[obj.pk]))

            obj.update_status()
            messages.success(request, f"✅ บันทึกรอบส่งของ {created_count} รายการ วันที่ {ship_date:%d/%m/%Y} แล้ว")

            if request.POST.get('next') == 'print':
                print_url = reverse('admin:stocks_salesorder_print_delivery', args=[obj.pk])
                return HttpResponseRedirect(f"{print_url}?shipped_date={ship_date.isoformat()}")
            return HttpResponseRedirect(reverse('admin:stocks_salesorder_change', args=[obj.pk]))

        elif action == 'edit_batch_date':
            old_date = _parse_batch_date_only(request.POST.get('old_date'))
            new_date = _combine_date_time(
                request.POST.get('new_date'),
                request.POST.get('new_hour'),
                request.POST.get('new_minute'),
            )
            if not old_date or not new_date:
                messages.error(request, "วันที่ไม่ถูกต้อง")
                return HttpResponseRedirect(reverse('admin:stocks_salesorder_change', args=[obj.pk]))

            logs = SalesDeliveryLog.objects.filter(sales_order=obj, shipped_date__date=old_date)
            updated_count = 0
            for log in logs:
                log.shipped_date = new_date
                log.save()  # ผ่าน .save() ปกติ ไม่ bulk update — จะได้ recalc DC/Rebate/shipment_value ให้ด้วย
                updated_count += 1

            obj.update_status()
            messages.success(request, f"✅ เปลี่ยนวันที่ส่งของรอบนี้ ({updated_count} รายการ) เป็น {new_date:%d/%m/%Y} แล้ว")

            if request.POST.get('next') == 'print':
                print_url = reverse('admin:stocks_salesorder_print_delivery', args=[obj.pk])
                return HttpResponseRedirect(f"{print_url}?shipped_date={new_date.isoformat()}")
            return HttpResponseRedirect(reverse('admin:stocks_salesorder_change', args=[obj.pk]))

        return HttpResponseBadRequest('unknown action')

    def unlock_view(self, request, object_id):
        # 🎯 ทำเป็น view แยกต่างหาก (ไม่ยัดปุ่มเข้าไปใน form ของ change_view) เพราะตอนสถานะ
        # Completed ฟอร์มทั้งหน้าจะถูก render แบบ readonly ล้วนๆ (has_change_permission เป็น False
        # ตอน GET) ไม่มี <input> จริงให้ submit เลย ถ้ายัดปุ่ม submit เข้าไปในฟอร์มนั้นจะโดน Django
        # เช็ค required fields (Customer, VAT ฯลฯ) แล้ว error "This field is required." ทั้งที่ไม่ได้
        # จะแก้ field พวกนั้นเลย — เลี่ยงปัญหานี้โดยไม่ผ่าน form validation ของ change_view เลย
        from django.core.exceptions import PermissionDenied
        from django.http import Http404
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        obj = self.get_object(request, object_id)
        if obj is None:
            raise Http404("ไม่พบใบสั่งขายนี้")
        if obj.status == 'Completed':
            obj.status = 'Shipped'
            obj.save(update_fields=['status'])
            self.message_user(request,
                f"🔓 ปลดล็อคใบสั่งขาย {obj.so_number} แล้ว แก้ไขรายการส่งของได้ตามปกติ — "
                f"ถ้าแก้เสร็จแล้วยอดส่งยังครบเหมือนเดิม ระบบจะปิดงานให้อัตโนมัติ")
        return HttpResponseRedirect(reverse('admin:stocks_salesorder_change', args=[obj.pk]))

    def print_view(self, request, object_id):
        from django.core.exceptions import PermissionDenied
        from django.http import Http404
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        obj = self.get_object(request, object_id)
        if obj is None:
            raise Http404("ไม่พบใบสั่งขายนี้")
        context = {
            **self.admin_site.each_context(request),
            'obj': obj,
            'items': obj.items.all().order_by('id').select_related('product', 'barcode_obj'),
            'title': f"ใบสั่งขาย {obj.so_number}",
        }
        return TemplateResponse(request, 'admin/sales_order_print.html', context)

    def print_delivery_view(self, request, object_id):
        from django.core.exceptions import PermissionDenied
        from django.http import Http404
        from django.conf import settings
        from .utils import thai_baht_text
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        obj = self.get_object(request, object_id)
        if obj is None:
            raise Http404("ไม่พบใบสั่งขายนี้")
        deliveries = obj.delivery_logs.all().order_by('shipped_date', 'id').select_related('product', 'barcode_obj')
        # 🎯 กด "พิมพ์" จากรอบใดรอบหนึ่งใน shipment panel (ดู ship_batch_view) จะแนบ shipped_date
        # มาด้วย เพื่อพิมพ์เฉพาะรอบนั้น — ถ้าไม่มี query param นี้ (เช่น เข้าจากปุ่มพิมพ์เดิมที่หัว
        # inline) ยังคง behavior เดิมคือโชว์ประวัติการส่งของทั้งหมด
        shipped_date_param = request.GET.get('shipped_date', '').strip()
        batch_date = None
        if shipped_date_param:
            from django.utils.dateparse import parse_date, parse_datetime
            # รองรับทั้งรูปแบบวันที่ล้วน ("YYYY-MM-DD" จากลิงก์พิมพ์ของแต่ละรอบใน shipment panel)
            # และรูปแบบ datetime เต็ม (จาก ship_batch_view หลังบันทึก/แก้ไขรอบแล้ว redirect มาพิมพ์)
            batch_date = parse_date(shipped_date_param)
            if not batch_date:
                batch_dt = parse_datetime(shipped_date_param)
                batch_date = batch_dt.date() if batch_dt else None
            if batch_date:
                deliveries = deliveries.filter(shipped_date__date=batch_date)
        deliveries = list(deliveries)

        # 🎯 กลุ่มรายการตามบาร์โค้ด — เอกสารนี้เป็นใบส่งของ/ใบกำกับภาษี ต้องโชว์ทีละรายการสินค้า
        # (ไม่ใช่ทีละแถวประวัติการบันทึก) เผื่อกรณีมีหลายแถวของบาร์โค้ดเดียวกันในช่วงที่พิมพ์
        line_map = {}
        line_order = []
        for d in deliveries:
            key = d.barcode_obj_id or f'product-{d.product_id}'
            if key not in line_map:
                unit_name = (d.barcode_obj.unit_name if d.barcode_obj else None) or 'ชิ้น'
                line_map[key] = {
                    'code': d.barcode_obj.code if d.barcode_obj else '-',
                    'product_name': d.product.name if d.product else '-',
                    'unit_name': unit_name,
                    'qty': 0,
                    'value': Decimal('0'),
                }
                line_order.append(key)
            line_map[key]['qty'] += d.quantity_shipped
            line_map[key]['value'] += d.shipment_value

        line_items = []
        for i, key in enumerate(line_order, start=1):
            row = line_map[key]
            unit_price = (row['value'] / row['qty']).quantize(Decimal('0.01')) if row['qty'] else Decimal('0')
            line_items.append({
                'no': i,
                'code': row['code'],
                'product_name': row['product_name'],
                'unit_name': row['unit_name'],
                'qty': row['qty'],
                'unit_price': unit_price,
                'line_total': row['value'],
            })

        subtotal = sum((row['value'] for row in line_map.values()), Decimal('0'))
        vat_percent = obj.vat_percent or Decimal('0')
        vat_amount = (subtotal * vat_percent / Decimal('100')).quantize(Decimal('0.01'))
        grand_total = subtotal + vat_amount

        if batch_date:
            doc_date = batch_date
        elif deliveries:
            doc_date = deliveries[-1].shipped_date.date()
        else:
            doc_date = obj.order_date

        due_date = deliveries[-1].payment_due_date if deliveries else None

        salesperson = '-'
        if obj.created_by_id:
            salesperson = obj.created_by.get_full_name() or obj.created_by.username

        context = {
            **self.admin_site.each_context(request),
            'obj': obj,
            'deliveries': deliveries,
            'line_items': line_items,
            'subtotal': subtotal,
            'vat_percent': vat_percent,
            'vat_amount': vat_amount,
            'grand_total': grand_total,
            'amount_words': thai_baht_text(grand_total),
            'doc_number': obj.so_number,
            'doc_date': doc_date,
            'credit_days': obj.customer.payment_term if obj.customer_id else 0,
            'due_date': due_date,
            'salesperson': salesperson,
            'customer': obj.customer,
            'company_name': settings.COMPANY_NAME,
            'company_address': settings.COMPANY_ADDRESS,
            'company_tax_id': settings.COMPANY_TAX_ID,
            'company_phone': settings.COMPANY_PHONE,
            'company_mobile': settings.COMPANY_MOBILE,
            'copies': [
                {'label': 'ต้นฉบับ (เอกสารออกเป็นชุด)', 'page_no': 1},
                {'label': 'สำเนา (เอกสารออกเป็นชุด)', 'page_no': 2},
            ],
            'title': f"ใบส่งสินค้า {obj.so_number}",
        }
        return TemplateResponse(request, 'admin/sales_delivery_print.html', context)

    @admin.action(description="✅ เปลี่ยนสถานะเป็น: เสร็จงาน/ปิดงาน")
    def mark_as_completed(self, request, queryset):
        queryset.update(status='Completed')
        self.message_user(request, f"ปิดงานสำเร็จ {queryset.count()} รายการแล้วค่ะ")

    def response_change(self, request, obj):
        if "_complete_order" in request.POST:
            obj.status = 'Completed'
            obj.save()
            self.message_user(request, f"ปิดงานใบสั่งขาย {obj.so_number} เรียบร้อยแล้ว")
            return HttpResponseRedirect(".")
        return super().response_change(request, obj)

    def render_change_form(self, request, context, add=False, change=False, form_url='', obj=None):
        if obj and change:
            reasons = []
            if obj.status in ('Completed', 'Cancelled'):
                reasons.append(f"สถานะ '{obj.status}'")
            if obj.delivery_logs.filter(
                Q(is_revenue_confirmed=True) | Q(is_dc_confirmed=True) | Q(is_rebate_confirmed=True)
            ).exists():
                reasons.append("มีรายการที่ยืนยันใน C6 แล้ว")
            if reasons:
                messages.warning(request,
                    f"🔒 เอกสารนี้ถูกล็อค ({', '.join(reasons)}) — ไม่สามารถเพิ่ม/แก้ไข/ลบรายการส่งของได้")
        response = super().render_change_form(request, context, add, change, form_url, obj)
        if obj and change:
            # Shipment panel: checklist for creating new delivery batches + list of past batches
            # 🎯 group ตาม "วันที่" เท่านั้น (ไม่เอาเวลา) เพราะรายการเก่าก่อนเปลี่ยนมาใช้ checklist
            # (ผ่านระบบ auto-save เดิม) มี shipped_date เป็นเวลาสุ่มไม่ตรงกันเป๊ะ ถ้า group ด้วย
            # datetime ตรงๆ แต่ละแถวเก่าจะกลายเป็นคนละ "รอบ" หมด ยาวรกจนใช้งานไม่ได้
            batch_dates = list(
                SalesDeliveryLog.objects.filter(sales_order=obj)
                .annotate(_d=TruncDate('shipped_date'))
                .order_by('_d').values_list('_d', flat=True).distinct()
            )
            batches = [
                {
                    'date': d,
                    'date_iso': d.isoformat(),
                    # ⏰ default เวลาเป็น 10:00 เสมอ (ตามที่เปรมขอ) — เดิมเก็บแค่วันที่ (group ตาม
                    # วันที่ล้วนๆ กันรายการเก่าที่เวลาสุ่มไม่ตรงกันแตกเป็นคนละรอบ) ตอนแก้ไขย้อนหลัง
                    # เลยต้องมีเวลาให้กรอกด้วยจะได้ตรงกับตอนสร้างรอบใหม่ — ใช้ dropdown ชั่วโมง/นาที
                    # ล้วนๆ (ไม่ใช่ <input type="time"> หรือ "datetime-local") เพราะ input พวกนั้น
                    # โชว์ AM/PM ตาม locale ของเบราว์เซอร์/OS ผู้ใช้ คุมให้เป็น 24 ชม. เสมอไม่ได้
                    'hour': '10',
                    'minute': '00',
                }
                for d in batch_dates
            ]

            sales_items = SalesItem.objects.filter(sales_order=obj).select_related('barcode_obj', 'product')
            pending_map = {}
            for item in sales_items:
                if not item.barcode_obj_id:
                    continue
                factor = item.barcode_obj.conversion_factor or 1
                ordered_pieces = item.quantity_ordered or 0
                shipped_units = SalesDeliveryLog.objects.filter(
                    sales_order=obj, barcode_obj=item.barcode_obj
                ).aggregate(total=Sum('quantity_shipped'))['total'] or 0
                remaining_pieces = ordered_pieces - shipped_units * factor
                remaining = remaining_pieces // factor
                if remaining > 0:
                    key = item.barcode_obj_id
                    if key not in pending_map:
                        pending_map[key] = {
                            'barcode_id': item.barcode_obj_id,
                            'code': item.barcode_obj.code,
                            'product_name': item.product.name if item.product else '-',
                            'unit_name': item.barcode_obj.unit_name or 'ชิ้น',
                            'qty': remaining,
                        }
                    else:
                        pending_map[key]['qty'] += remaining

            shipment_panel_html = render_to_string('admin/sales_shipment_panel.html', {
                'batches': batches,
                'pending_items': list(pending_map.values()),
                'next_batch_no': len(batches) + 1,
                'ship_url': reverse('admin:stocks_salesorder_ship', args=[obj.pk]),
                'print_base_url': reverse('admin:stocks_salesorder_print_delivery', args=[obj.pk]),
                # ⏰ default วันที่ส่งของรอบใหม่ = วันนี้ เวลา 10:00 (ตามที่เปรมขอ)
                'default_new_date': timezone.now().date().isoformat(),
                'default_new_hour': '10',
                'default_new_minute': '00',
                'hour_options': [f"{h:02d}" for h in range(24)],
                'minute_options': [f"{m:02d}" for m in range(60)],
            }, request=request)
            shipment_panel_wrapped = f'<div id="sales-shipment-panel-holder" style="display:none;">{shipment_panel_html}</div>'
            move_panel_script = """
                <script>
                    django.jQuery(document).ready(function() {
                        var $holder = django.jQuery('#sales-shipment-panel-holder');
                        var $heading = django.jQuery('h2[id$="-heading"]:contains("Sales delivery logs")').first();
                        var $panel = $holder.children().first();
                        $panel.show();
                        if ($heading.length) {
                            $heading.closest('[id*="delivery_logs"]').before($panel);
                        } else {
                            django.jQuery('#content-main').prepend($panel);
                        }
                        $holder.remove();
                    });
                </script>
            """
            # 🎯 #submit-row ของ Unfold อยู่ "นอก" <form> จริง (เป็น sticky bar แยกต่างหาก) ปุ่ม submit ที่ยัด
            # เข้าไปต้องมี attribute form="{model}_form" กำกับด้วย ไม่งั้นกดแล้วไม่มี <form> ให้ผูก (กดแล้ว
            # ไม่เกิดอะไรขึ้นเลย) — ใช้ได้กับปุ่ม "เสร็จงาน" เพราะตอนนั้นฟอร์มยัง render แบบแก้ไขได้ปกติ
            # แต่ใช้กับปุ่ม "ปลดล็อค" ไม่ได้ เพราะตอนสถานะ Completed ทั้งฟอร์ม render แบบ readonly ล้วนๆ
            # (ไม่มี input จริงให้ Customer/VAT ฯลฯ) submit ทั้งฟอร์มแล้วจะโดน required-field error แทน
            # เลยทำปุ่มปลดล็อคเป็นลิงก์ไปหา view แยกต่างหาก (unlock_view) ที่ไม่ผ่านการ validate ฟอร์มเลย
            form_id = f"{self.model._meta.model_name}_form"
            unlock_url = reverse('admin:stocks_salesorder_unlock', args=[obj.pk])
            unlock_btn_html = ''
            if obj.status == 'Completed':
                unlock_btn_html = f'<a href="{unlock_url}" style="display:inline-block; background: #f59e0b; color: white; height: 35px; line-height: 35px; margin-right: 10px; border-radius: 4px; border: none; cursor: pointer; padding: 0 20px; font-weight: bold; text-decoration: none;">🔓 ปลดล็อค (Unlock)</a>'
            complete_btn_script = f"""
                <script>
                    django.jQuery(document).ready(function() {{
                        var target = django.jQuery('#submit-row .flex-col-reverse');
                        if (!target.length) {{ target = django.jQuery('#submit-row'); }}
                        var btn = '<input type="submit" form="{form_id}" value="เสร็จงาน (Complete)" name="_complete_order" style="background: #218838; color: white; height: 35px; margin-right: 10px; border-radius: 4px; border: none; cursor: pointer; padding: 0 20px; font-weight: bold;">';
                        target.prepend(btn);
                        target.prepend('{unlock_btn_html}');
                    }});
                </script>
            """
            # ปุ่มพิมพ์: แปะเข้าแถบหัวข้อ (h2) ของแต่ละ inline group โดยตรง (ตรวจสอบโครงสร้างจริงของ
            # Unfold แล้วว่า h2 มี id="<formset-prefix>-heading" และข้อความในนั้นคือ verbose_name_plural
            # ล้วนๆ) ถ้าหา header ไม่เจอ (เผื่อ Unfold เปลี่ยนโครงสร้างในอนาคต) ให้ fallback ไป submit-row
            print_so_url = reverse('admin:stocks_salesorder_print', args=[obj.pk])
            print_delivery_url = reverse('admin:stocks_salesorder_print_delivery', args=[obj.pk])
            print_btn_script = f"""
                <script>
                    django.jQuery(document).ready(function() {{
                        function addPrintBtn(headingText, url, bg, fullLabel) {{
                            var $header = django.jQuery('h2[id$="-heading"]:contains("' + headingText + '")').first();
                            if ($header.length) {{
                                var link = '<a href="' + url + '" target="_blank" class="ml-auto" style="display:inline-block; background:' + bg + '; color:white; height:26px; line-height:26px; border-radius:4px; border:none; cursor:pointer; padding:0 14px; font-weight:bold; font-size:13px; text-decoration:none;">🖨️ พิมพ์</a>';
                                $header.append(link);
                            }} else {{
                                var fallback = django.jQuery('#submit-row .flex-col-reverse');
                                if (!fallback.length) {{ fallback = django.jQuery('#submit-row'); }}
                                var link = '<a href="' + url + '" target="_blank" style="display:inline-block; background:' + bg + '; color:white; height:35px; line-height:35px; margin-right:10px; border-radius:4px; border:none; cursor:pointer; padding:0 20px; font-weight:bold; text-decoration:none;">🖨️ ' + fullLabel + '</a>';
                                fallback.prepend(link);
                            }}
                        }}
                        addPrintBtn('Sales items', '{print_so_url}', '#17a2b8', 'พิมพ์ใบสั่งขาย');
                        addPrintBtn('Sales delivery logs', '{print_delivery_url}', '#6f42c1', 'พิมพ์ใบส่งสินค้า');
                    }});
                </script>
            """
            response.render()
            response.content = response.content.replace(
                b'</body>', (shipment_panel_wrapped + move_panel_script + complete_btn_script + print_btn_script).encode('utf-8') + b'</body>', 1
            )
        return response

    def save_model(self, request, obj, form, change):
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

    def get_diff(self, obj):
        ordered = sum(i.quantity_ordered for i in obj.items.all())
        shipped = sum(l.quantity_shipped for l in obj.delivery_logs.all())
        return color_diff(shipped - ordered)
    get_diff.short_description = "สถานะส่งของ"

    # ❗ ต้องเพิ่มฟังก์ชันนี้เข้าไปด้วย ระบบถึงจะหาตัวสร้างใบ PD เจอครับ
    def create_auto_production_order(self, sales_item, user):
        from .models import ProductionOrder, BOM
        import datetime
        
        # 1. เช็กสถานะ "สินค้าผลิตเอง" (has_bom)
        if not getattr(sales_item.product, 'has_bom', False):
            return "NOT_MANUFACTURED" # คืนค่าบอกว่าตัวนี้ไม่ใช่สินค้าผลิต

        bom_to_use = sales_item.bom 
    
        if not bom_to_use:
        # ถ้าในบรรทัดนั้นไม่มี BOM จริงๆ ค่อยลองหาตัวล่าสุด (Backup plan)
            bom_to_use = BOM.objects.filter(product=sales_item.product).order_by('-id').first()

        # 2. เช็กว่ามีสูตร BOM ในระบบจริงไหม
        bom_obj = BOM.objects.filter(product=sales_item.product).first()
        if not bom_obj:
            return "NO_BOM_FORMULA" # คืนค่าบอกว่ายังไม่ได้ทำสูตร

        # 3. ตรวจสอบจำนวนสั่งซื้อ
        qty = getattr(sales_item, 'quantity_ordered', 0)
        if qty <= 0:
            return "ZERO_QTY"

        # 4. ถ้าผ่านเงื่อนไขทั้งหมด -> สร้างใบผลิต
        try:
            new_pd = ProductionOrder(
                product=sales_item.product,
                bom=bom_to_use,
                quantity_planned=qty,
                status='Draft',
                order_date=datetime.date.today(),
                created_by=user,
                notes=f"Auto PD จาก SO: {sales_item.sales_order.so_number}"
            )
            new_pd.save()
            return new_pd # คืนค่า object PD เมื่อสำเร็จ
        except Exception:
            return "SAVE_ERROR"
    
    @admin.action(description='⚡ เปิดใบสั่งผลิต (Auto PD)')
    def make_production_order(self, request, queryset):
        created_count = 0
        fail_list = []

        for so in queryset:
            for item in so.items.all():
                # เรียกใช้ฟังก์ชัน Engine ที่เราปรับปรุง
                new_pd = self.create_auto_production_order(item, request.user)
                
                if new_pd:
                    created_count += 1
                else:
                    fail_list.append(f"{item.product.name} ({so.so_number})")

        # สรุปผลบนแถบแจ้งเตือน
        if created_count > 0:
            self.message_user(request, f"✅ สร้างสำเร็จ {created_count} รายการ", messages.SUCCESS)
        
        if fail_list:
            msg = "⚠️ ข้ามรายการที่ไม่มี BOM: " + ", ".join(fail_list)
            self.message_user(request, msg, messages.WARNING)
            
    # ✅ ฟังก์ชันสร้างใบผลิตอัตโนมัติ
    def save_formset(self, request, form, formset, change):
        from django.contrib import messages
        from .models import SalesItem

        if formset.model == SalesItem:
            # เราต้องหาว่าอันไหนกำลังจะโดนลบ เพื่อตัดยอดออกจาก Planning
            deleted_count = 0
            for delete_form in formset.deleted_forms:
                if delete_form.instance.pk:
                    deleted_count += 1
            
            # บันทึกข้อมูลลงฐานข้อมูล (รวมถึงสั่งลบรายการที่ติ๊กไว้ด้วย)
            instances = formset.save(commit=False)
            
            # สั่งลบจริงในฐานข้อมูลสำหรับรายการที่ติ๊ก Delete
            for obj in formset.deleted_objects:
                obj.delete()
            
            # ตัวนับแยกประเภท
            count_success = 0
            count_not_manufactured = 0
            count_no_formula = 0
            
            for instance in instances:
                instance.save() # เซฟข้อมูลเบื้องต้นก่อน
                
                # ตรวจสอบ: ถ้าติ๊ก auto_produce และยังไม่ได้ถูกผลิต
                if getattr(instance, 'auto_produce', False) and not getattr(instance, 'is_produced', False):
                    
                    result = self.create_auto_production_order(instance, request.user)
                    
                    if isinstance(result, object) and not isinstance(result, str):
                        # ✅ กรณีสำเร็จ
                        instance.is_produced = True
                        instance.save()
                        count_success += 1
                    else:
                        # ❌ กรณีไม่สำเร็จ: ล้างติ๊กถูกออกทันที
                        instance.auto_produce = False
                        instance.save()
                        
                        # แยกประเภท Error เพื่อเก็บยอดสรุป
                        if result == "NOT_MANUFACTURED":
                            count_not_manufactured += 1
                        elif result == "NO_BOM_FORMULA":
                            count_no_formula += 1

            formset.save_m2m()

            # 📢 สรุปข้อความแจ้งเตือนแยกตามประเภท
            if count_success > 0:
                messages.success(request, f"✅ สร้างใบผลิตสำเร็จ {count_success} รายการ")
            
            if count_not_manufactured > 0:
                messages.warning(request, f"ℹ️ ไม่ใช่สินค้าผลิตเอง {count_not_manufactured} รายการ (ระบบล้างติ๊กออกให้แล้ว)")
            
            if count_no_formula > 0:
                messages.error(request, f"⚠️ ยังไม่ได้สร้างสูตร BOM {count_no_formula} รายการ (กรุณาไปสร้างสูตรก่อน)")
        else:
            formset.save()

    def get_confirmed_status(self, obj):
        # ไปแอบดูใน Log ว่ามีการติ๊กรับชำระเงินหรือยัง
        from .models import SalesDeliveryLog # 👈 Import มาใช้ตรงๆ
        
        # ใช้ obj (ที่เป็น SalesOrder) ไปหาใน Log
        confirmed = SalesDeliveryLog.objects.filter(
            sales_order=obj, 
            is_revenue_confirmed=True
        ).exists()
        
        if confirmed:
            return format_html('<span style="color:green;"><b>✔ ยืนยันยอดจากใบส่งของแล้ว</b></span>')
        return format_html('<span style="color:gray;">รอยืนยันยอด</span>')
        
    get_confirmed_status.short_description = "สถานะการตรวจสอบ"

    def has_change_permission(self, request, obj=None):
        if obj:
            # 🔒 เปลี่ยนจากเช็ก Confirmation เป็นเช็กสถานะใบสั่งขาย
            if obj.status == 'Completed':
                return False # ล็อคเฉพาะตอนกด "เสร็จงาน/ปิดงาน" เท่านั้น — ปุ่ม "ปลดล็อค" ไม่ผ่านเส้นทางนี้
                             # (เป็นลิงก์ไป unlock_view แยกต่างหาก ใช้ has_view_or_change_permission แทน)
        return super().has_change_permission(request, obj)
    class Media:
        # ⚠️ delivery_barcode_select2.js ต้องโหลดก่อน smart_delivery_inline.js เสมอ — ตัวมันดักจับ
        # submit event ของฟอร์มหลักไว้ก่อน (เพื่อรอ auto-save ที่ยังค้างอยู่ให้เสร็จก่อนค่อยปล่อยให้
        # submit จริง) ถ้าโหลดสลับกัน ตัวกันกด submit ซ้ำใน smart_delivery_inline.js จะบล็อคการ
        # re-submit ทีหลังของมันไปด้วย (เพราะ set flag "submitted" ไปแล้วตั้งแต่รอบแรก)
        js = ('js/admin_sum_selected.js', 'js/delivery_barcode_select2.js', 'js/smart_delivery_inline.js', 'js/sales_item_barcode_autofill.js')

class ProductionMaterialUsageInline(UnfoldTabularInline):
    model = ProductionMaterialUsage
    extra = 0
    fields = ['raw_material', 'get_projected_stock', 'planned_qty', 'actual_qty_to_use', 'used_so_far', 'auto_produce']
    readonly_fields = ['planned_qty', 'used_so_far', 'get_projected_stock'] # สองฟิลด์นี้ให้ระบบคำนวณเอง
    verbose_name = "ส่วนประกอบ/Package ตามสูตร"

    # ใช้สูตรเดียวกับ StockPlanningAdmin.get_available ("คาดการณ์ (Plan)")
    _ACTIVE_PO = ['Draft', 'Pending', 'Confirmed', 'Ordered', 'Paid', 'Loaded', 'Departed', 'Arrived', 'Received', 'Partially Received']
    _ACTIVE_SO = ['Draft', 'Confirmed', 'Shipped']
    _ACTIVE_PD = ['Draft', 'Started', 'Finished']

    def get_projected_stock(self, obj):
        if not obj or not obj.raw_material_id:
            return "-"
        material = obj.raw_material
        p_in = PurchaseItem.objects.filter(
            product=material, purchase_order__status__in=self._ACTIVE_PO
        ).aggregate(t=Sum(Greatest(F('quantity_ordered') - F('quantity_received'), Value(0))))['t'] or 0
        p_out = SalesItem.objects.filter(
            product=material, sales_order__status__in=self._ACTIVE_SO
        ).aggregate(t=Sum(Greatest(F('quantity_ordered') - F('quantity_shipped'), Value(0))))['t'] or 0
        p_receipt = ProductionOrder.objects.filter(
            product=material, status__in=self._ACTIVE_PD
        ).aggregate(t=Sum(Greatest(F('quantity_planned') - F('quantity_actual'), Value(0))))['t'] or 0
        # actual_qty_to_use/used_so_far เป็น DecimalField ต้องระบุ output_field ให้ Value(0) ชัดเจน
        # ไม่งั้น Django จะ error "mixed types: DecimalField, IntegerField"
        # ไม่นับรวมยอดจองของ "ใบสั่งผลิตนี้เอง" เพื่อให้เห็นสต็อกที่มีจริงก่อนหักของใบนี้
        p_usage_qs = ProductionMaterialUsage.objects.filter(
            raw_material=material, production_order__status__in=self._ACTIVE_PD
        )
        if obj.production_order_id:
            p_usage_qs = p_usage_qs.exclude(production_order_id=obj.production_order_id)
        p_usage = p_usage_qs.aggregate(t=Sum(Greatest(F('actual_qty_to_use') - F('used_so_far'), Value(0, output_field=DecimalField()))))['t'] or 0
        net = material.stock_quantity + int(p_in) - int(p_out) + (int(p_receipt) - int(p_usage))
        color = "#dc3545" if net < 0 else "#0d6efd"
        return format_html('<b style="color: {};">{}</b> {}', color, net, material.unit)
    get_projected_stock.short_description = "ยอดสต็อกคาดการณ์"

@admin.register(ProductionOrder)
class ProductionOrderAdmin(DocumentLockMixin, UnfoldModelAdmin):
    fields = ['product', 'bom', 'quantity_planned', 'quantity_actual', 'created_by','status', 'notes']
    list_display = ('pd_number', 'product', 'quantity_planned', 'quantity_actual', 'get_diff', 'status')
    list_filter = (
        ('status', MultipleChoicesDropdownFilter),
        ('order_date', DjangoDateRangeFilter),
        ('product', AutocompleteSelectMultipleFilter),
    )
    list_filter_submit = True
    search_fields = ('pd_number', 'product__name')
    autocomplete_fields = ['product']
    inlines = [ProductionMaterialUsageInline,ProductionLogInline]
    date_hierarchy = 'order_date' # ✅ เพิ่มบรรทัดนี้ค่ะ
    readonly_fields = ('pd_number','quantity_actual',  'created_by', 'status') 
    
    actions = ['mark_as_completed']

    class Media:
        js = (
            'js/filter_bom.js',
            'js/admin_sum_selected.js',
        ) # เรียกไฟล์ JS มาใช้งาน

    # ✅ ต้องมีฟังก์ชันนี้ และ Indent (ย่อหน้า) ให้ตรงกับฟังก์ชันอื่นในคลาสครับ
    @admin.action(description="✅ เปลี่ยนสถานะเป็น: เสร็จงาน/ปิดงาน")
    def mark_as_completed(self, request, queryset):
        queryset.update(status='Completed')
        self.message_user(request, f"ปิดงานผลิตสำเร็จ {queryset.count()} รายการแล้วค่ะ")

    def response_change(self, request, obj):
        if "_complete_order" in request.POST:
            obj.status = 'Completed'
            obj.save()
            self.message_user(request, f"ปิดงานผลิต {obj.pd_number} เรียบร้อยแล้ว")
            return HttpResponseRedirect(".")
        return super().response_change(request, obj)

    def get_urls(self):
        custom_urls = [
            path('<int:object_id>/print/', self.admin_site.admin_view(self.print_view), name='stocks_productionorder_print'),
        ]
        return custom_urls + super().get_urls()

    def print_view(self, request, object_id):
        from django.core.exceptions import PermissionDenied
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        obj = self.get_object(request, object_id)
        if obj is None:
            from django.http import Http404
            raise Http404("ไม่พบใบสั่งผลิตนี้")
        context = {
            **self.admin_site.each_context(request),
            'obj': obj,
            'materials': obj.material_usages.select_related('raw_material').all(),
            'title': f"ใบสั่งผลิต {obj.pd_number}",
        }
        return TemplateResponse(request, 'admin/production_order_print.html', context)

    def render_change_form(self, request, context, add=False, change=False, form_url='', obj=None):
        # ⚠️ ห้ามฝัง <script> ผ่าน context['title'] เพราะ Unfold (Tailwind theme) ไม่ได้การันตีว่า
        # {{ title }} จะถูก echo ออกมาใน body เหมือน Django admin เดิม (title ไปโผล่แค่ <title> บาง view)
        # ใช้วิธีเดียวกับ DocumentLockMixin คือฉีดสคริปต์เข้า response.content ก่อน </body> โดยตรง แน่นอนกว่า
        response = super().render_change_form(request, context, add, change, form_url, obj)
        if obj and change:
            print_url = reverse('admin:stocks_productionorder_print', args=[obj.pk])
            script = f"""
                <script>
                    django.jQuery(document).ready(function() {{
                        var target = django.jQuery('#submit-row .flex-col-reverse');
                        if (!target.length) {{ target = django.jQuery('#submit-row'); }}
                        var completeBtn = '<input type="submit" value="ปิดงานผลิต (Complete)" name="_complete_order" style="background: #28a745; color: white; height: 35px; margin-right: 10px; border-radius: 4px; border: none; cursor: pointer; padding: 0 20px; font-weight: bold;">';
                        var printBtn = '<a href="{print_url}" target="_blank" style="display:inline-block; background:#17a2b8; color:white; height:35px; line-height:35px; margin-right:10px; border-radius:4px; border:none; cursor:pointer; padding:0 20px; font-weight:bold; text-decoration:none;">🖨️ พิมพ์ใบสั่งผลิต</a>';
                        target.prepend(completeBtn);
                        target.prepend(printBtn);
                    }});
                </script>
            """
            response.render()
            response.content = response.content.replace(b'</body>', script.encode('utf-8') + b'</body>', 1)
        return response


    def save_model(self, request, obj, form, change):
        # ✅ 1. กรณีเปลี่ยน BOM (ให้ล้างของเก่า ดึงของใหม่)
        if change and 'bom' in form.changed_data:
            # ใช้ชื่อที่เปรมตั้งไว้ใน related_name
            obj.material_usages.all().delete() 
            
            super().save_model(request, obj, form, change)
            
            if hasattr(obj, 'load_materials_from_bom'):
                obj.load_materials_from_bom()
            return

        # ✅ 2. กรณีสร้างใบใหม่
        elif not change and obj.bom:
            super().save_model(request, obj, form, change)
            if hasattr(obj, 'load_materials_from_bom'):
                obj.load_materials_from_bom()
            return

        super().save_model(request, obj, form, change)

    def get_diff(self, obj):
        planned = obj.quantity_planned
        actual = obj.quantity_actual
        return color_diff(actual - planned)
    get_diff.short_description = "สถานะผลิต"

    def create_child_production_order(self, usage, user):
        """สร้างใบสั่งผลิตต่อเนื่องสำหรับวัตถุดิบที่มี BOM ของตัวเอง (cascade จากใบสั่งผลิตนี้)"""
        import math
        material = usage.raw_material
        if not material.has_bom:
            return "NOT_MANUFACTURED"
        if not BOM.objects.filter(product=material).exists():
            return "NO_BOM_FORMULA"
        qty = usage.actual_qty_to_use
        if qty <= 0:
            return "ZERO_QTY"
        try:
            # ต่อท้ายหมายเหตุเดิมของใบแม่ (เช่น "Auto PD จาก SO: ...") ไม่ทับของเดิม
            # เพื่อให้ไล่ที่มาย้อนกลับได้จากใบลูกใบเดียว เว้นแต่ผู้ใช้ลบหมายเหตุเดิมออกเองก่อน save
            parent = usage.production_order
            ref_line = f"Auto PD จาก PD: {parent.pd_number}"
            notes = f"{parent.notes}\n{ref_line}" if parent.notes else ref_line
            new_pd = ProductionOrder(
                product=material,
                quantity_planned=math.ceil(qty),
                status='Draft',
                order_date=datetime.date.today(),
                created_by=user,
                notes=notes
            )
            new_pd.save()
            return new_pd
        except Exception:
            return "SAVE_ERROR"

    def save_formset(self, request, form, formset, change):
        from .models import ProductionLog, ProductionMaterialUsage
        from django.db.models import Sum

        # ✅ 1. เคลียร์รายการที่ติ๊ก Delete (แบบปลอดภัย)
        # ใช้ deleted_forms แทน deleted_objects เพื่อป้องกัน AttributeError
        for delete_form in formset.deleted_forms:
            if delete_form.instance.pk:
                delete_form.instance.delete()

        # ✅ 2. บันทึก/แก้ไข รายการที่เหลือ
        instances = formset.save(commit=False)
        cascade_success = 0
        cascade_fail = []
        for instance in instances:
            # ถ้าเป็น Log และยังไม่มีคนบันทึก ให้ใส่ชื่อ user คนปัจจุบัน
            if isinstance(instance, ProductionLog):
                if not instance.user_id:
                    instance.user = request.user
            instance.save()

            # ติ๊ก "ผลิตทันที (Auto PD)" ในรายการวัตถุดิบ -> สร้างใบสั่งผลิตต่อเนื่องให้วัตถุดิบตัวนี้
            if isinstance(instance, ProductionMaterialUsage) and instance.auto_produce and not instance.is_produced:
                result = self.create_child_production_order(instance, request.user)
                if isinstance(result, ProductionOrder):
                    instance.is_produced = True
                    instance.save()
                    cascade_success += 1
                else:
                    instance.auto_produce = False
                    instance.save()
                    cascade_fail.append(f"{instance.raw_material.name} ({result})")
        formset.save_m2m()

        # ✅ 3. เฉพาะกรณีเป็นตาราง ProductionLog ให้คำนวณยอดสะสมใหม่
        if formset.model == ProductionLog:
            obj = formset.instance
            total_finished = obj.production_logs.aggregate(s=Sum('quantity_finished'))['s'] or 0
            obj.quantity_actual = total_finished
            obj.save()

        if cascade_success:
            self.message_user(request, f"✅ สร้างใบสั่งผลิตต่อเนื่องสำเร็จ {cascade_success} รายการ", messages.SUCCESS)
        if cascade_fail:
            self.message_user(request, "⚠️ ข้ามรายการที่สร้างใบสั่งผลิตต่อเนื่องไม่ได้ (ไม่มี BOM/ไม่ใช่สินค้าผลิตเอง): " + ", ".join(cascade_fail), messages.WARNING)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        # 🎯 ดึง obj (ข้อมูลใบผลิตนี้) ออกมาจาก kwargs (Django จะใส่มาให้เอง)
        obj = kwargs.get('obj')

        # 1. กรองสินค้า: เอาเฉพาะที่มี BOM
        if db_field.name == "product":
            from .models import Product
            kwargs["queryset"] = Product.objects.filter(has_bom=True)

        # 2. กรองสูตร BOM: ให้ตรงกับสินค้าในใบนี้
        if db_field.name == "bom":
            if obj and hasattr(obj, 'product') and obj.product:
                kwargs["queryset"] = BOM.objects.filter(product=obj.product)
            elif 'product' in request.GET:
                kwargs["queryset"] = BOM.objects.filter(product_id=request.GET.get('product'))
        
        # 🎯 ห้ามใส่ obj ลงใน super() นะครับ! 
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    


class BuyPriceRangeFilter(admin.SimpleListFilter):
    title = 'ช่วงราคาทุน'
    parameter_name = 'price_range'

    def lookups(self, request, model_admin):
        return [
            ('0-100', '0 - 100 บาท'),
            ('101-500', '101 - 500 บาท'),
            ('501-1000', '501 - 1,000 บาท'),
            ('1001-plus', 'มากกว่า 1,000 บาท'),
        ]

    def queryset(self, request, queryset):
        if self.value() == '0-100': return queryset.filter(buy_price__lte=100)
        if self.value() == '101-500': return queryset.filter(buy_price__gt=100, buy_price__lte=500)
        if self.value() == '501-1000': return queryset.filter(buy_price__gt=500, buy_price__lte=1000)
        if self.value() == '1001-plus': return queryset.filter(buy_price__gt=1000)
        return queryset

class AdvanceOrderRuleForm(forms.ModelForm):
    class Meta:
        model = AdvanceOrderRule
        fields = '__all__'
        widgets = {
            'order_type': forms.RadioSelect,
        }


@admin.register(AdvanceOrderRule)
class AdvanceOrderRuleAdmin(UnfoldModelAdmin):
    form = AdvanceOrderRuleForm
    list_display = ('fo_number', 'order_type', 'product', 'quantity', 'frequency_days', 'next_run_date', 'end_date')
    list_filter = (('order_type', MultipleChoicesDropdownFilter),)
    search_fields = ('fo_number', 'product__name')
    readonly_fields = ('fo_number',)
    # barcode_obj ไม่ใช้ autocomplete — ให้ JS (advance_order_rule_form.js) คุมตัวเลือกทั้งหมด
    # จำกัดเฉพาะบาร์โค้ดของสินค้าที่เลือกเท่านั้น (ผ่าน /api/product-barcodes/)
    autocomplete_fields = ['product', 'bom', 'supplier']
    fields = ('fo_number', 'order_type', 'product', 'barcode_obj', 'bom', 'supplier', 'quantity', 'frequency_days', 'end_date', 'next_run_date')

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)
        product_id = request.GET.get('product')
        if product_id:
            initial['product'] = product_id
            initial['order_type'] = 'PURCHASE'
        return initial

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'barcode_obj':
            # จำกัดตัวเลือกให้เห็นเฉพาะบาร์โค้ดของสินค้าที่เลือกไว้แล้วเท่านั้น (แก้ไข/prefill จาก ?product=)
            # กรณีหน้าเพิ่มใหม่ที่ยังไม่รู้สินค้า ให้ว่างไว้ก่อน แล้วให้ JS เติมตัวเลือกให้เองหลังเลือกสินค้า
            resolved = request.resolver_match
            object_id = resolved.kwargs.get('object_id') if resolved else None
            product_id = request.GET.get('product')
            if object_id:
                obj = self.get_object(request, object_id)
                product_id = obj.product_id if obj else None
            if product_id:
                kwargs['queryset'] = ProductBarcode.objects.filter(product_id=product_id)
            else:
                kwargs['queryset'] = ProductBarcode.objects.none()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def save_model(self, request, obj, form, change):
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

    class Media:
        js = ('js/advance_order_rule_form.js',)


@admin.register(StockForecast)
class StockForecastAdmin(UnfoldModelAdmin):
    list_display = ('name', 'category', 'stock_quantity', 'get_weekly', 'get_2weekly', 'get_monthly', 'get_3monthly', 'get_add_button')
    list_filter = (
        ('category', AutocompleteSelectMultipleFilter),
        ('suppliers', AutocompleteSelectMultipleFilter),
        ('tags', AutocompleteSelectMultipleFilter),
        BuyPriceRangeFilter,
    )
    list_filter_submit = True
    search_fields = ('name', 'barcodes__code', 'tags__name')
    list_select_related = ('category',)

    def get_queryset(self, request):
        # หน้านี้เป็นการคาดการณ์การใช้สินค้า จึงกรองเฉพาะ "สินค้าที่มีสต๊อกจริง" เสมอ (ไม่มีตัวเลือกปิด)
        return super().get_queryset(request).filter(is_product=True).distinct()

    def _forecast(self, obj, period_days):
        # cache เหตุการณ์ย้อนหลังไว้ที่ object เพราะ list_display เรียกทีละคอลัมน์ต่อแถวเดิม
        # (ไม่งั้นจะ query ยอดขาย/ยอดใช้ผลิตซ้ำ 4 รอบต่อ 1 แถว)
        if not hasattr(obj, '_forecast_events_cache'):
            obj._forecast_events_cache = _historical_usage_events(obj, datetime.date.today())
        return _forecast_for_period(obj._forecast_events_cache, datetime.date.today(), period_days)

    @admin.display(description="weekly")
    def get_weekly(self, obj):
        return self._forecast(obj, 7)

    @admin.display(description="2weekly")
    def get_2weekly(self, obj):
        return self._forecast(obj, 14)

    @admin.display(description="monthly")
    def get_monthly(self, obj):
        return self._forecast(obj, 30)

    @admin.display(description="3monthly")
    def get_3monthly(self, obj):
        return self._forecast(obj, 90)

    @admin.display(description="+")
    def get_add_button(self, obj):
        url = reverse('admin:stocks_advanceorderrule_add') + f'?product={obj.pk}'
        return format_html('<a href="{}" style="font-weight:bold; font-size:16px;">+</a>', url)


@admin.register(StockPlanning)
class StockPlanningAdmin(ExportToExcelMixin, UnfoldModelAdmin):
    list_display = ('name', 'category', 'stock_quantity', 'min_stock', 'get_pending_in', 'get_pending_out', 'get_pending_prod', 'get_available', 'buy_price', 'get_total_inventory_value')
    list_filter = (
        ('category', AutocompleteSelectMultipleFilter),
        ('suppliers', AutocompleteSelectMultipleFilter),
        ('tags', AutocompleteSelectMultipleFilter),
        ProductOnlyFilter,
        BuyPriceRangeFilter,
    )
    list_filter_submit = True
    search_fields = ('name', 'barcodes__code', 'tags__name')
    actions = ['export_to_excel']

    list_select_related = ('category',)

    _ACTIVE_PO = ['Draft', 'Pending', 'Confirmed', 'Ordered', 'Paid', 'Loaded', 'Departed', 'Arrived', 'Received', 'Partially Received']
    _ACTIVE_SO = ['Draft', 'Confirmed', 'Shipped']
    _ACTIVE_PD = ['Draft', 'Started', 'Finished']

    def get_queryset(self, request):
        from .models import PurchaseItem, SalesItem, ProductionOrder, ProductionMaterialUsage
        qs = super().get_queryset(request)

        pending_in_sq = PurchaseItem.objects.filter(
            product=OuterRef('pk'),
            purchase_order__status__in=self._ACTIVE_PO,
        ).values('product').annotate(
            t=Sum(Greatest(F('quantity_ordered') - F('quantity_received'), Value(0)))
        ).values('t')

        pending_out_sq = SalesItem.objects.filter(
            product=OuterRef('pk'),
            sales_order__status__in=self._ACTIVE_SO,
        ).values('product').annotate(
            t=Sum(Greatest(F('quantity_ordered') - F('quantity_shipped'), Value(0)))
        ).values('t')

        pending_receipt_sq = ProductionOrder.objects.filter(
            product=OuterRef('pk'),
            status__in=self._ACTIVE_PD,
        ).values('product').annotate(
            t=Sum(Greatest(F('quantity_planned') - F('quantity_actual'), Value(0)))
        ).values('t')

        pending_usage_sq = ProductionMaterialUsage.objects.filter(
            raw_material=OuterRef('pk'),
            production_order__status__in=self._ACTIVE_PD,
        ).values('raw_material').annotate(
            # actual_qty_to_use/used_so_far เป็น DecimalField ต้องระบุ output_field ให้ Value(0) ชัดเจน
            # ไม่งั้น Django จะ error "mixed types: DecimalField, IntegerField"
            t=Sum(Greatest(F('actual_qty_to_use') - F('used_so_far'), Value(0, output_field=DecimalField())))
        ).values('t')

        # .distinct() กัน row ซ้ำเวลา filter ผ่าน tags/suppliers (M2M join)
        return qs.annotate(
            _pending_in=Coalesce(Subquery(pending_in_sq, output_field=DecimalField()), Value(0), output_field=DecimalField()),
            _pending_out=Coalesce(Subquery(pending_out_sq, output_field=DecimalField()), Value(0), output_field=DecimalField()),
            _pending_receipt=Coalesce(Subquery(pending_receipt_sq, output_field=DecimalField()), Value(0), output_field=DecimalField()),
            _pending_usage=Coalesce(Subquery(pending_usage_sq, output_field=DecimalField()), Value(0), output_field=DecimalField()),
        ).distinct()

    def _net(self, obj):
        p_in = int(obj._pending_in or 0)
        p_out = int(obj._pending_out or 0)
        p_receipt = int(obj._pending_receipt or 0)
        p_usage = int(obj._pending_usage or 0)
        return obj.stock_quantity + p_in - p_out + (p_receipt - p_usage)

    def get_pending_in(self, obj):
        return int(obj._pending_in or 0)
    get_pending_in.short_description = "แผนรับ (PO)"

    def get_pending_out(self, obj):
        return int(obj._pending_out or 0)
    get_pending_out.short_description = "แผนส่ง (SO)"

    def get_pending_prod(self, obj):
        net = int((obj._pending_receipt or 0) - (obj._pending_usage or 0))
        if net == 0:
            return 0
        color = "#28a745" if net > 0 else "#dc3545"
        return format_html('<span style="color: {};">{}</span>', color, net)
    get_pending_prod.short_description = "แผนผลิต (PD)"

    def get_available(self, obj):
        total = self._net(obj)
        color = "red" if total < 0 else "blue"
        return format_html('<b style="color: {};">{}</b>', color, total)
    get_available.short_description = "คาดการณ์ (Plan)"

    def get_total_inventory_value(self, obj):
        available_total = float(self._net(obj))
        total_value = available_total * float(obj.buy_price or 0)
        color = "#fd7e14" if total_value < 0 else "#212529"
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color,
            "{:,.2f}".format(total_value),
        )
    get_total_inventory_value.short_description = "มูลค่ารวม"

    def get_urls(self):
        custom_urls = [
            path('timeline/', self.admin_site.admin_view(self.timeline_view), name='stocks_stockplanning_timeline'),
            path('timeline/print/', self.admin_site.admin_view(self.timeline_print_view), name='stocks_stockplanning_timeline_print'),
        ]
        return custom_urls + super().get_urls()

    def timeline_view(self, request):
        data = self._timeline_data(request)
        context = {
            **self.admin_site.each_context(request),
            'title': '📅 Timeline Stock',
            'opts': self.model._meta,
            **data,
        }
        return TemplateResponse(request, 'admin/stock_timeline.html', context)

    def timeline_print_view(self, request):
        """หน้าพิมพ์ Timeline stock: ตารางล้วนๆ ไม่มี chrome ของแอดมิน (เหมือนหน้าพิมพ์ใบสั่งซื้อ/ขาย)"""
        data = self._timeline_data(request)
        return TemplateResponse(request, 'admin/stock_timeline_print.html', data)

    def _timeline_data(self, request):
        """
        คำนวณข้อมูล Timeline stock (rows + gridlines) ใช้ร่วมกันทั้งหน้าเว็บปกติ (timeline_view)
        และหน้าพิมพ์ (timeline_print_view) เพื่อให้ตารางตรงกันเป๊ะๆ ไม่ต้องคำนวณซ้ำสองที่

        กราฟ Gantt แบบ % position แสดงสต็อกปัจจุบันที่จุดเริ่ม แล้ว plot รายการรับ/ส่ง/ผลิต
        ที่ยัง pending อยู่ ตามวันที่ประมาณการ (PO: order_date + delivery_lead_time, SO: order_date,
        PD รับ: order_date + production_lead_time, PD เบิกวัตถุดิบ: order_date ของใบสั่งผลิต)
        จนถึงยอดคาดการณ์สุดท้าย ณ วันสิ้นสุดช่วงที่เลือก — เหตุการณ์ที่เกินช่วงที่แสดงจะไม่ถูกนับเลย
        เพื่อให้ยอดคาดการณ์ตรงกับสิ่งที่เห็นบนหน้าจอ
        """
        from django.core.paginator import Paginator
        from .models import PurchaseItem, SalesItem, ProductionOrder, ProductionMaterialUsage

        today = datetime.date.today()
        try:
            start_date = datetime.date.fromisoformat(request.GET.get('start', ''))
        except ValueError:
            start_date = today
        try:
            end_date = datetime.date.fromisoformat(request.GET.get('end', ''))
        except ValueError:
            end_date = start_date + timedelta(days=60)
        if end_date <= start_date:
            end_date = start_date + timedelta(days=60)
        total_days = max((end_date - start_date).days, 1)

        q = request.GET.get('q', '').strip()
        # ⚠️ ใช้ queryset เปล่า ๆ (ไม่เรียก self.get_queryset ของ get_queryset ด้านบน) เพราะอันนั้นมี
        # subquery annotate _pending_in/out/receipt/usage ที่หนักและไม่ได้ใช้ในหน้านี้เลย —
        # หน้านี้คำนวณ event เองแยกด้านล่างอยู่แล้ว ตัดออกให้เบาและเร็วขึ้น
        qs = self.model.objects.select_related('category').order_by('name')

        # Filters (เหมือนหน้า Forecast): category / ประเภทรายการ / supplier / ช่วงราคาทุน
        category_id = request.GET.get('category', '').strip()
        if category_id:
            qs = qs.filter(category_id=category_id)

        is_product = request.GET.get('is_product', '').strip()
        if is_product == 'false':
            qs = qs.filter(is_product=False)
        elif is_product == 'all':
            pass
        else:
            qs = qs.filter(is_product=True)  # default เหมือน ProductOnlyFilter

        supplier_id = request.GET.get('supplier', '').strip()
        if supplier_id:
            qs = qs.filter(suppliers__id=supplier_id).distinct()

        tag_ids = [t for t in request.GET.getlist('tag') if t.strip()]
        if tag_ids:
            qs = qs.filter(tags__id__in=tag_ids).distinct()

        price_range = request.GET.get('price_range', '').strip()
        if price_range == '0-100':
            qs = qs.filter(buy_price__lte=100)
        elif price_range == '101-500':
            qs = qs.filter(buy_price__gt=100, buy_price__lte=500)
        elif price_range == '501-1000':
            qs = qs.filter(buy_price__gt=500, buy_price__lte=1000)
        elif price_range == '1001-plus':
            qs = qs.filter(buy_price__gt=1000)

        if q:
            qs, _ = self.get_search_results(request, qs, q)

        paginator = Paginator(qs, 200)
        page_obj = paginator.get_page(request.GET.get('page', 1))
        products = list(page_obj.object_list)
        product_ids = [p.pk for p in products]
        products_by_id = {p.pk: p for p in products}

        events_by_product = {pid: [] for pid in product_ids}

        po_rows = PurchaseItem.objects.filter(
            product_id__in=product_ids,
            purchase_order__status__in=self._ACTIVE_PO,
        ).annotate(
            remaining=Greatest(F('quantity_ordered') - F('quantity_received'), Value(0))
        ).filter(remaining__gt=0).values(
            'product_id', 'remaining', 'purchase_order_id', 'purchase_order__order_date', 'purchase_order__po_number'
        )

        so_rows = SalesItem.objects.filter(
            product_id__in=product_ids,
            sales_order__status__in=self._ACTIVE_SO,
        ).annotate(
            remaining=Greatest(F('quantity_ordered') - F('quantity_shipped'), Value(0))
        ).filter(remaining__gt=0).values(
            'product_id', 'remaining', 'sales_order_id', 'sales_order__order_date', 'sales_order__so_number'
        )

        pd_rows = ProductionOrder.objects.filter(
            product_id__in=product_ids,
            status__in=self._ACTIVE_PD,
        ).annotate(
            remaining=Greatest(F('quantity_planned') - F('quantity_actual'), Value(0))
        ).filter(remaining__gt=0).values('id', 'product_id', 'remaining', 'order_date', 'pd_number')

        usage_rows = ProductionMaterialUsage.objects.filter(
            raw_material_id__in=product_ids,
            production_order__status__in=self._ACTIVE_PD,
        ).annotate(
            remaining=Greatest(F('actual_qty_to_use') - F('used_so_far'), Value(0, output_field=DecimalField()))
        ).filter(remaining__gt=0).values(
            'raw_material_id', 'remaining', 'production_order_id',
            'production_order__order_date', 'production_order__pd_number',
        )

        # เก็บ (วันที่, +/-จำนวน, เอกสารอ้างอิง {label, url}) ไว้โยงกลับไปดูใบสั่งซื้อ/ขาย/ผลิตที่เกี่ยวข้องได้ (คลิกได้)
        for row in po_rows:
            lead = products_by_id[row['product_id']].delivery_lead_time or 0
            event_date = row['purchase_order__order_date'] + timedelta(days=lead)
            ref = {
                'label': f"PO {row['purchase_order__po_number']}",
                'url': f"/admin/stocks/purchaseorder/{row['purchase_order_id']}/change/",
            }
            events_by_product[row['product_id']].append((event_date, int(row['remaining']), ref))

        for row in so_rows:
            ref = {
                'label': f"SO {row['sales_order__so_number']}",
                'url': f"/admin/stocks/salesorder/{row['sales_order_id']}/change/",
            }
            events_by_product[row['product_id']].append((row['sales_order__order_date'], -int(row['remaining']), ref))

        for row in pd_rows:
            lead = products_by_id[row['product_id']].production_lead_time or 0
            event_date = row['order_date'] + timedelta(days=lead)
            ref = {
                'label': f"PD {row['pd_number']}",
                'url': f"/admin/stocks/productionorder/{row['id']}/change/",
            }
            events_by_product[row['product_id']].append((event_date, int(row['remaining']), ref))

        for row in usage_rows:
            ref = {
                'label': f"PD {row['production_order__pd_number']}",
                'url': f"/admin/stocks/productionorder/{row['production_order_id']}/change/",
            }
            events_by_product[row['raw_material_id']].append(
                (row['production_order__order_date'], -int(row['remaining']), ref)
            )

        # FO: กฎสั่งซื้อ/สั่งผลิตล่วงหน้า (B7) ที่ยังไม่เกิดเป็น PO/PD จริง — โชว์แค่ "รอบถัดไป" ของแต่ละกฎ
        # เป็นยอดคาดการณ์บนกราฟ Timeline เท่านั้น ไม่นับรวมในยอด "คาดการณ์ (Plan)" ของหน้า C1/รายการหลัก
        fo_rules = AdvanceOrderRule.objects.filter(product_id__in=product_ids).exclude(
            end_date__isnull=False, end_date__lt=F('next_run_date')
        )
        for rule in fo_rules:
            product_obj = products_by_id[rule.product_id]
            lead = (product_obj.delivery_lead_time if rule.order_type == 'PURCHASE' else product_obj.production_lead_time) or 0
            event_date = rule.next_run_date + timedelta(days=lead)
            ref = {
                'label': f"FO {rule.fo_number}",
                'url': f"/admin/stocks/advanceorderrule/{rule.pk}/change/",
            }
            events_by_product[rule.product_id].append((event_date, int(rule.quantity), ref))

        def fmt_qty(n):
            """ย่อเลขจำนวนมาก: 12,500 -> 12.5K, 3,200,000 -> 3.2M (ทศนิยม 1 ตำแหน่ง) — เอาไว้ให้ label ในกราฟไม่ยาวเกิน"""
            n = int(n)
            a = abs(n)
            if a >= 1_000_000:
                s = f"{a / 1_000_000:.1f}M"
            elif a >= 1_000:
                s = f"{a / 1_000:.1f}K"
            else:
                s = str(a)
            return f"-{s}" if n < 0 else s

        def make_point(pct, balance, delta=None, event_date=None, refs=None):
            return {
                'pct': round(pct, 2),
                'date': event_date,
                'balance': balance,
                'delta': delta,
                'balance_fmt': fmt_qty(balance),
                'balance_full': f"{balance:,}",
                'delta_fmt': fmt_qty(delta) if delta is not None else None,
                'delta_full': f"{delta:+,}" if delta is not None else None,
                'refs': refs or [],
                'refs_text': ', '.join(r['label'] for r in refs) if refs else None,
            }

        rows = []
        for product in products:
            # เหตุการณ์ที่เกินวันสิ้นสุดช่วง → ตัดทิ้งไปเลย ไม่นับ (ยอดคาดการณ์ต้องตรงกับที่เห็นในกราฟ)
            # เหตุการณ์ที่ค้าง (วันที่ผ่านมาแล้วแต่ยัง pending) → นับที่จุดเริ่ม (start_date)
            # เก็บเลขที่เอกสาร (PO/SO/PD) ของทุกรายการที่ถูก net รวมกันในวันเดียวกันไว้ด้วย ไว้โชว์ตอน hover
            daily_delta = {}
            daily_refs = {}
            for event_date, delta, ref in events_by_product.get(product.pk, []):
                if event_date > end_date:
                    continue
                clamped = max(event_date, start_date)
                daily_delta[clamped] = daily_delta.get(clamped, 0) + delta
                daily_refs.setdefault(clamped, []).append(ref)

            running = product.stock_quantity
            points = [make_point(0.0, running)]
            for event_date in sorted(daily_delta):
                delta = daily_delta[event_date]
                running += delta
                pct = (event_date - start_date).days / total_days * 100
                points.append(make_point(pct, running, delta, event_date, daily_refs.get(event_date)))

            if points[-1]['pct'] < 100:
                points.append(make_point(100.0, running))

            rows.append({'product': product, 'points': points})

        querystring = request.GET.copy()
        querystring.pop('page', None)

        # เส้นแบ่งแนวตั้งตามวันที่: เส้นบางทุกวัน + เส้นเข้ม (dashed) ทุก 5 วันไว้จับตำแหน่งง่ายๆ
        # คำนวณครั้งเดียวสำหรับทั้งตาราง (ไม่ผูกกับแต่ละแถว) เพื่อไม่ให้ช้าตอน 200 แถว
        strong_every = 5
        day_step = 1 if total_days <= 120 else strong_every
        gridlines = []
        d = 0
        while d <= total_days:
            gridlines.append({
                'pct': round(d / total_days * 100, 2),
                'date': start_date + timedelta(days=d),
                'strong': (d % strong_every == 0),
            })
            d += day_step
        # ⚠️ วันสุดท้ายของช่วง (d == total_days) จะได้ pct=100 เป๊ะเสมอจาก loop ข้างบนอยู่แล้ว แต่ไม่ได้
        # แปลว่ามันจะ strong=True เสมอ (เช่น total_days=29 → 29 % 5 != 0) ถ้าปล่อยแบบนั้นป้ายวันที่สุดท้าย
        # จะหายไปเฉยๆ (ดูว่างๆ ตรงขอบขวา) ต้องบังคับให้ตัวสุดท้ายเป็น strong เสมอ ไม่ว่าจะ append ใหม่หรือมีอยู่แล้ว
        if gridlines[-1]['pct'] < 100:
            gridlines.append({'pct': 100.0, 'date': end_date, 'strong': True})
        else:
            gridlines[-1]['strong'] = True
        # ป้ายวันที่ตัวสุดท้าย ต้องชิดขอบขวา ไม่งั้นโดน overflow:hidden ตัดขาด (ตัวอื่นชิดซ้ายของเส้นอยู่แล้วโดย default)
        strong_ticks = [g for g in gridlines if g['strong']]
        if strong_ticks:
            strong_ticks[-1]['tick_align'] = 'end'

        # สัดส่วนความกว้างแต่ละช่วงระหว่างป้ายวันที่ (สำหรับ grid-template-columns ในหน้าพิมพ์ —
        # ใช้ layout แบบ grid ธรรมดาแทน position:absolute เพราะ Chrome มีบั๊กไม่วาดข้อความที่ position:absolute
        # ซ้ำใน <thead> ที่ repeat ข้ามหน้ากระดาษ (เส้น grid ที่ไม่มีตัวอักษรวาดซ้ำได้ปกติ แต่ป้ายวันที่ที่มีตัวอักษรไม่วาด)
        # ⚠️ คอลัมน์ของแต่ละวันที่ต้องเป็นช่วง "หลัง" ตัวมันเอง (ไปจนถึงวันที่ถัดไป) ไม่ใช่ช่วง "ก่อน" ตัวมันเอง
        # เพราะ label ถูก left-align ไว้ที่ต้นคอลัมน์ (ให้อยู่หลังเส้นเหมือนจุดข้อมูลในแถว) — ถ้าจับคู่ผิดฝั่ง
        # ป้ายวันที่จะไปโผล่ตรงตำแหน่งวันก่อนหน้าแทน ยิ่งวันหลังๆ ยิ่งเห็นเพี้ยนชัด (แถมคอลัมน์แรกจะกว้าง=0 ด้วย
        # เพราะวันแรกอยู่ที่ 0% อยู่แล้ว)
        tick_cells = []
        for _i, _t in enumerate(strong_ticks):
            _next_pct = strong_ticks[_i + 1]['pct'] if _i + 1 < len(strong_ticks) else 100.0
            tick_cells.append({'date': _t['date'], 'width_pct': round(_next_pct - _t['pct'], 4)})

        # สรุป filter ที่ใช้อยู่เป็นข้อความ (ไว้โชว์เป็นหัวกระดาษตอนพิมพ์ แทนฟอร์ม filter ที่กดเลือกได้จริง)
        filter_summary_parts = [f"ช่วงเวลา {start_date:%d/%m/%Y} - {end_date:%d/%m/%Y}"]
        if q:
            filter_summary_parts.append(f'ค้นหา "{q}"')
        if category_id:
            cat = ProductCategory.objects.filter(pk=category_id).first()
            if cat:
                filter_summary_parts.append(f"หมวดหมู่ {cat.name}")
        if supplier_id:
            sup = Supplier.objects.filter(pk=supplier_id).first()
            if sup:
                filter_summary_parts.append(f"Supplier {sup.company_name}")
        if tag_ids:
            tag_names = list(ProductTag.objects.filter(pk__in=tag_ids).values_list('name', flat=True))
            if tag_names:
                filter_summary_parts.append(f"Tag {', '.join(tag_names)}")
        if is_product == 'false':
            filter_summary_parts.append("ไม่ใช่สินค้า")
        elif is_product == 'all':
            filter_summary_parts.append("สินค้า+อื่นๆ ทั้งหมด")
        if price_range:
            filter_summary_parts.append(f"ราคาทุน {price_range}")

        return {
            'rows': rows,
            'start_date': start_date,
            'end_date': end_date,
            'gridlines': gridlines,
            'tick_cells': tick_cells,
            'q': q,
            'page_obj': page_obj,
            'querystring': querystring.urlencode(),
            'categories': ProductCategory.objects.order_by('name'),
            'suppliers': Supplier.objects.order_by('company_name'),
            'tags': ProductTag.objects.order_by('name'),
            'f_category': category_id,
            'f_is_product': is_product,
            'f_supplier': supplier_id,
            'f_tag': tag_ids,
            'f_price_range': price_range,
            'filter_summary': ' | '.join(filter_summary_parts),
        }

    class Media:
        js = ('js/admin_sum_selected.js', 'js/stock_view_toggle.js') # เรียกไฟล์ JS มาใช้งาน

@admin.register(ProductCategory)
class ProductCategoryAdmin(UnfoldModelAdmin):
    list_display = ('name', 'products_link')
    search_fields = ('name',)
    readonly_fields = ('products_link',)
    fields = ('name', 'products_link')

    def get_queryset(self, request):
        from django.db.models import Count
        return super().get_queryset(request).annotate(num_products=Count('product'))

    def products_link(self, obj):
        # ✅ ลิงก์ไปหน้ารายการสินค้าที่ filter ตาม category แทนการฝัง inline ทั้งหมด
        # (category ที่มีสินค้าเป็นร้อยรายการทำให้หน้า change ช้ามาก ทั้งตอนโหลดและตอน save
        # เพราะ Django ต้อง build/validate formset ของทุกแถวเสมอ ไม่ว่าจะแก้อะไรก็ตาม)
        if not obj or not obj.pk:
            return "-"
        count = getattr(obj, 'num_products', None)
        if count is None:
            count = obj.product_set.count()
        url = f"/admin/stocks/product/?category__id__exact={obj.pk}"
        return format_html('<a class="button" href="{}">📦 ดูสินค้าในกลุ่มนี้ ({} รายการ)</a>', url, count)
    products_link.short_description = "สินค้าในกลุ่มนี้"

    def changelist_view(self, request, extra_context=None):
        from django.db.models import Count
        # ✅ รวมร่าง: ดึง Tag พร้อมนับจำนวนสินค้า เรียงจากใช้บ่อยสุด (-num_products) และใหม่สุด (-id)
        tags = ProductTag.objects.annotate(num_products=Count('products')).order_by('-num_products', '-id')
        
        # ส่วนหัวของกล่อง Tag Cloud
        tag_html = '<div style="margin-bottom: 20px; padding: 20px; background: #fff; border: 1px solid #ddd; border-radius: 10px; line-height: 2.5; box-shadow: 0 2px 4px rgba(0,0,0,0.05);">'
        tag_html += '<h3 style="margin:0 0 15px 0; color:#333; font-size:18px; border-bottom: 2px solid #eee; padding-bottom: 10px;">🏷️ แท็กยอดนิยม & แท็กมาใหม่ (คลิกเพื่อดูสินค้า)</h3>'
        
        if tags.exists():
            for tag in tags:
                count = tag.num_products
                # ✅ คำนวณขนาด Font: ยิ่งใช้เยอะ ยิ่งตัวใหญ่ (Max 24px, Min 13px)
                # min(count, 10) เพื่อไม่ให้ตัวใหญ่เกินไปจนล้นจอ
                font_size = 13 + (min(count, 10) * 1.1) 
                
                # ลิงก์ไปหน้ารายการสินค้าแบบ Filter Tag ID ทันที
                url = f"/admin/stocks/product/?tags__id__exact={tag.id}"
                
                tag_html += f'''
                    <a href="{url}" style="display: inline-block; margin: 5px 10px; padding: 5px 18px; 
                    background: {tag.color}; color: white; border-radius: 25px; text-decoration: none; 
                    font-weight: bold; font-size: {font_size}px; box-shadow: 2px 2px 5px rgba(0,0,0,0.1); 
                    transition: transform 0.2s; border: 1px solid rgba(0,0,0,0.1);">
                    #{tag.name} <span style="font-size: 11px; opacity: 0.85;">({count})</span>
                    </a>'''
        else:
            tag_html += '<p style="color:#999; padding: 10px;">ยังไม่มีการสร้างแท็กสินค้าในระบบ</p>'
        
        tag_html += '</div>'
        
        # ส่งค่าไปยัง Template ของ Django Admin
        extra_context = extra_context or {}
        extra_context['tag_cloud'] = mark_safe(tag_html)
        return super().changelist_view(request, extra_context=extra_context)

# ---------------------------------------------------------
# Register ProductTag เพื่อให้เมนูโผล่ในหน้า Admin
# ---------------------------------------------------------
@admin.register(ProductTag)
class ProductTagAdmin(UnfoldModelAdmin):
    # แสดงชื่อแท็กและตัวอย่างสีในหน้า List
    list_display = ('display_name_with_count', 'color')
    search_fields = ('name',)
    readonly_fields = ('products_link',)

    def get_queryset(self, request):
        # ใช้ annotate นับจำนวนสินค้าที่เชื่อมกับ Tag นี้
        qs = super().get_queryset(request)
        return qs.annotate(product_count=models.Count('products'))

    def display_name_with_count(self, obj):
        # เอาชื่อ Tag มาต่อด้วยจำนวนสินค้า
        return f"{obj.name} ({obj.product_count})"
    display_name_with_count.short_description = "ชื่อแท็ก (จำนวนสินค้า)"

    def products_link(self, obj):
        # ✅ ลิงก์ไปหน้ารายการสินค้าที่ filter ตาม tag แทนการฝัง inline ทั้งหมด
        # (แท็กที่มีสินค้าเยอะๆ ทำให้หน้า change ช้ามาก ทั้งตอนโหลดและตอน save เหมือนที่แก้ไปแล้ว
        # ใน ProductCategoryAdmin เพราะ Django ต้อง build/validate formset ของทุกแถวเสมอ)
        if not obj or not obj.pk:
            return "-"
        count = getattr(obj, 'product_count', None)
        if count is None:
            count = obj.products.count()
        url = f"/admin/stocks/product/?tags__id__exact={obj.pk}"
        return format_html('<a class="button" href="{}">📦 ดูสินค้าที่ใช้แท็กนี้ ({} รายการ)</a>', url, count)
    products_link.short_description = "สินค้าที่ใช้แท็กนี้"

    def color_display(self, obj):
        # แสดงเป็นกล่องสีสวยๆ ให้เห็นในหน้า Admin เลยครับ
        return format_html(
            '<span style="background: {}; color: white; padding: 3px 10px; border-radius: 12px; font-weight: bold;">{}</span>',
            obj.color, obj.name
        )
    color_display.short_description = "ตัวอย่างสี"

    # ✅ แถม: เพิ่มฟังก์ชันนับจำนวนสินค้าในหน้า List ให้ดูง่ายๆ ค่ะ
    def get_product_count(self, obj):
        return obj.product_set.count()
    get_product_count.short_description = "จำนวนสินค้าที่ใช้"

class PaymentDateForm(forms.Form):
    payment_date = forms.DateField(
        label="ระบุวันที่ชำระเงิน",
        initial=timezone.now,
        widget=AdminDateWidget()
    )

@admin.action(description="🎯 ปิดยอด: กรณีพิเศษ/รับไม่ครบ (SETTLED)")    
def settle_income_special(modeladmin, request, queryset):
    if 'apply' in request.POST:
        # ตัดจบสถานะอย่างเดียว ไม่สร้างบันทึกการเงินเพิ่ม
        count = queryset.update(status='COMPLETED', payment_status='SETTLED')
        modeladmin.message_user(request, f"ตัดจบรายการรายรับสำเร็จ {count} รายการ (SETTLED)", messages.SUCCESS)
        return None

    return TemplateResponse(request, "admin/settle_confirmation.html", {
        **modeladmin.admin_site.each_context(request),
        'title': "ยืนยันปิดยอดรายรับกรณีพิเศษ (SETTLED)",
        'queryset': queryset,
        'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
        'action_name': 'settle_income_special', # ต้องตรงกับชื่อฟังก์ชัน,
        'mode': 'income'
    })
settle_income_special.short_description = "🎯 ปิดยอดกรณีพิเศษ (SETTLED)"

# ✅ ปุ่มใหม่สำหรับฝั่งรายจ่าย (Purchase)
def settle_purchase_special(modeladmin, request, queryset):
    if 'apply' in request.POST:
        count = queryset.update(status='COMPLETED', payment_status='SETTLED')
        modeladmin.message_user(request, f"ตัดจบรายการรายจ่ายสำเร็จ {count} รายการ (SETTLED)", messages.SUCCESS)
        return None

    return TemplateResponse(request, "admin/settle_confirmation.html", {
        **modeladmin.admin_site.each_context(request),
        'title': "ยืนยันปิดยอดรายจ่ายกรณีพิเศษ (SETTLED)",
        'queryset': queryset,
        'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
        'action_name': 'settle_purchase_special'
    })
settle_purchase_special.short_description = "🎯 ปิดยอดกรณีพิเศษ (SETTLED)"

# ✅ Action: ปิดงาน Finance แบบมีหน้ายืนยัน (Confirmation Page)
@admin.action(description='💰 ชำระครบ/ปิดยอด (Settle Payment)')
def settle_and_close_orders(modeladmin, request, queryset):
    # ... (Logic ปิดงาน) ...
    if 'apply' in request.POST:
        form = PaymentDateForm(request.POST)
        if form.is_valid():
            pay_date = form.cleaned_data['payment_date']
            updated_count = 0
            
            for obj in queryset:
                balance = obj.balance_due
                # สร้างรายการจ่ายเงิน (ตามยอดที่ค้าง)
                if balance > 0:
                    if isinstance(obj, PurchaseOrder):
                        PurchasePaymentLog.objects.create(purchase_order=obj, amount=balance, payment_date=pay_date, notes="Auto Settle")
                        obj.refresh_from_db()
                    elif isinstance(obj, SalesOrder): # รองรับทั้ง SalesOrder และ IncomeReport
                        SalesPayment.objects.create(order=obj, amount=balance, payment_date=pay_date, remark="Auto Settle")
                        obj.refresh_from_db()
                    updated_count += 1
                
                # บังคับอัปเดตสถานะการเงินเป็น "Paid"
                if obj.balance_due <= 0:
                    obj.payment_status = 'Paid'
                else:
                    obj.payment_status = 'Partial' # เพิ่มบรรทัดนี้เผื่อปิดยอดไม่หมดค่ะ
                
                obj.save(update_fields=['payment_status'])
            
            modeladmin.message_user(request, f"✅ บันทึกการชำระเงินเรียบร้อย {updated_count} รายการ", messages.SUCCESS)
            return HttpResponseRedirect(request.get_full_path())
            
    else:
        form = PaymentDateForm()

    # HTML Template สำหรับหน้าเลือกวันที่
    html_template = """
    {% extends "admin/base_site.html" %}
    {% load i18n admin_urls static admin_modify %}
    {% block extrahead %}{{ block.super }}<script src="{% url 'admin:jsi18n' %}"></script>{{ media }}{% endblock %}
    {% block content %}
    <div style="max-width: 600px; margin: 20px auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
        <h2 style="color: #007bff;">💰 ยืนยันการชำระเงินและปิดยอด ({{ queryset.count }} รายการ)</h2>
        <p>ระบบจะสร้างรายการชำระเงิน <b>"เต็มจำนวนคงเหลือ"</b> และเปลี่ยนสถานะเป็น <b>Paid</b> ให้อัตโนมัติ</p>
        <form method="post">{% csrf_token %}
            {% for obj in queryset %}<input type="hidden" name="{{ action_checkbox_name }}" value="{{ obj.pk }}">{% endfor %}
            <input type="hidden" name="action" value="settle_and_close_orders">
            <input type="hidden" name="apply" value="1">
            <div style="margin: 20px 0;">{{ form.as_p }}</div>
            <button type="submit" style="background: #007bff; color: white; border: none; padding: 10px 20px; font-size: 16px; border-radius: 4px; cursor: pointer;">✅ ยืนยัน (Confirm)</button>
            <a href="#" onclick="window.history.back();" style="margin-left: 10px; color: #666;">ยกเลิก</a>
        </form>
    </div>
    {% endblock %}
    """
    
    context = {
        'queryset': queryset, 'form': form, 'media': form.media, 
        'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME, 'opts': modeladmin.model._meta,
    }
    return HttpResponse(Template(html_template).render(RequestContext(request, context)))

@admin.register(FinanceReport)
class FinanceReportAdmin(ExportToExcelMixin, DocumentLockMixin, UnfoldModelAdmin):
    # หน้ารวม: ดูง่ายๆ ว่าใบไหนค้างจ่าย
    search_fields = ('po_number', 'supplier__company_name')
    actions = [settle_and_close_orders, settle_purchase_special, 'calculate_finance_totals', 'export_to_excel']

    @admin.action(description="📝 สรุปยอดเงินรายจ่ายที่เลือก")
    def calculate_finance_totals(self, request, queryset):
        grand_total = 0
        paid_total = 0
        total_balance_due = 0 # ✅ ใช้ชื่อให้ตรงกับ Balance Due ในหน้าจอ

        for obj in queryset:
            grand_total += float(obj.grand_total or 0)
            paid_total += float(obj.total_paid or 0)
            total_balance_due += float(obj.balance_due or 0)

        summary_message = (
            f"📊 สรุปรายจ่าย {queryset.count()} รายการ:  |  "
            f"💰 ยอดจ่ายสุทธิรวม: {grand_total:,.2f} บาท  |  "
            f"✅ จ่ายแล้วรวม: {paid_total:,.2f} บาท  |  "
            f"❗️ ค้างจ่ายรวม (Balance Due): {total_balance_due:,.2f} บาท"
        )
        self.message_user(request, summary_message, messages.SUCCESS)

    # จัดหน้าตาฟอร์ม
    # ✅ 1. เปลี่ยน list_display ให้โชว์ Payment Status แทน
    list_display = ('po_number', 'get_invoice_no_supplier', 'get_supplier_truncated', 'get_grand_total_list', 'get_balance_due_list', 'payment_status')
    # ✅ 2. ตัวกรอง ก็ต้องกรองตามการจ่ายเงิน
    list_filter = (
        ('order_date', DjangoDateRangeFilter),
        ('payment_status', MultipleChoicesDropdownFilter),
        ('supplier', AutocompleteSelectMultipleFilter),
        ('items__product__tags', PurchaseOrderTagsFilter),
    )
    list_filter_submit = True
    # ✅ 3. ในหน้า Detail ก็เปลี่ยน fields
    fieldsets = (
        ('📊 สรุปยอดเงิน', {
            'fields': (
                ('get_total_items_display', 'get_subtotal_display'), 
                ('vat_percent', 'get_vat_amount_display'), 
                ('get_grand_total_display', 'get_total_paid_display', 'get_balance_due_display')
            ),
            'classes': ('wide',), 
        }),
        ('📝 ข้อมูลเอกสาร', {
            'fields': ('po_number', 'invoice_no_supplier', 'supplier', 'order_date', 'status', 'payment_status')
        }),
    )

    readonly_fields = (
        'po_number', 'supplier', 'order_date',
        'get_total_items_display', 'get_subtotal_display',
        'get_vat_amount_display', 'get_grand_total_display',
        'get_total_paid_display', 'get_balance_due_display',
        'status', 'payment_status'
    )

    inlines = [PurchaseItemReadOnlyInline, PurchasePaymentInline]

    def get_invoice_no_supplier(self, obj):
        val = obj.invoice_no_supplier or '-'
        return format_html('<span style="font-size:12px;color:#555;">{}</span>', val)
    get_invoice_no_supplier.short_description = "Invoice ผู้ขาย"

    def get_supplier_truncated(self, obj):
        name = str(obj.supplier) if obj.supplier else '-'
        return format_html(
            '<span style="display:inline-block;max-width:140px;overflow:hidden;'
            'text-overflow:ellipsis;white-space:nowrap;" title="{}">{}</span>',
            name, name
        )
    get_supplier_truncated.short_description = "Supplier"

    def save_formset(self, request, form, formset, change):
        # 1. บันทึกข้อมูลที่กรอกในตารางก่อน
        formset.save()
    
        # 2. ดึงใบสั่งซื้อใบนี้ออกมา
        obj = formset.instance
    
        # 3. เช็คว่าถ้าเป็นการเซฟตาราง "บันทึกการจ่ายเงิน" ให้คำนวณสถานะใหม่
        if formset.model == PurchasePaymentLog:
            from django.db.models import Sum
        
        # ✅ ท่าไม้ตาย: ไม่ต้องง้อ _set แต่สั่งไปที่ Model PurchasePaymentLog โดยตรงเลย
        # กรองเอาเฉพาะรายการที่ฟิลด์ 'order' ตรงกับใบนี้
            paid_data = PurchasePaymentLog.objects.filter(purchase_order=obj).aggregate(Sum('amount'))
            paid = paid_data['amount__sum'] or 0
        
            # ยอดสุทธิที่ต้องจ่าย
            total = obj.grand_total

            update_fields = ['payment_status']
            if paid <= 0:
                obj.payment_status = 'Unpaid'
            elif paid < total:
                obj.payment_status = 'Partial'
            else:
                obj.payment_status = 'Paid'
                latest = PurchasePaymentLog.objects.filter(purchase_order=obj).order_by('-payment_date').first()
                if latest and not obj.paid_date:
                    obj.paid_date = latest.payment_date
                    update_fields.append('paid_date')

            obj.save(update_fields=update_fields)
    
    def get_total_items_display(self, obj):
        return f"{sum(i.quantity_ordered for i in obj.items.all()):,}"
    get_total_items_display.short_description = "📦 รวมจำนวนสินค้า"

    def get_subtotal_display(self, obj):
        # ✅ แก้ไข: จัดรูปแบบตัวเลขก่อนส่งเข้า HTML
        return format_html('<span style="font-size:14px;">{}</span>', f"{obj.total_items_price:,.2f}")
    get_subtotal_display.short_description = "💵 ราคารวม (ก่อน VAT)"

    def get_vat_amount_display(self, obj):
        return f"{obj.vat_amount:,.2f}"
    get_vat_amount_display.short_description = "ภาษีมูลค่าเพิ่ม (VAT)"

    def get_grand_total_display(self, obj):
        # ✅ แก้ไข: จัดรูปแบบตัวเลขก่อนส่งเข้า HTML
        return format_html('<b style="color:#007bff;">{}</b>', f"{obj.grand_total:,.2f}")
    get_grand_total_display.short_description = "💰 ยอดสุทธิ"

    def get_total_paid_display(self, obj):
        # ✅ แก้ไข
        return format_html('<b style="color:#28a745;">{}</b>', f"{obj.total_paid:,.2f}")
    get_total_paid_display.short_description = "✅ จ่ายแล้ว"

    def get_balance_due_display(self, obj):
        # ✅ แก้ไข
        balance = obj.balance_due
        color = "red" if balance > 0 else "green"
        text = f"{balance:,.2f}"
        return format_html('<b style="color:{};">{}</b>', color, text)
    

    def get_queryset(self, request):
        from django.db.models import Sum, F, ExpressionWrapper, DecimalField as DField
        return super().get_queryset(request).select_related('supplier').annotate(
            _total_items_price=Sum(
                ExpressionWrapper(F('items__quantity_ordered') * F('items__unit_price'), output_field=DField())
            ),
            _total_paid=Sum('payment_logs__amount'),
        )

    # --- List Display Functions (หน้ารวม) ---
    def get_grand_total_list(self, obj):
        subtotal = getattr(obj, '_total_items_price', None) or 0
        vat_p = obj.vat_percent or 0
        total = subtotal + (subtotal * vat_p / 100)
        return f"{total:,.2f}"
    get_grand_total_list.short_description = "💰 ยอดสุทธิ"

    def get_balance_due_list(self, obj):
        subtotal = getattr(obj, '_total_items_price', None) or 0
        vat_p = obj.vat_percent or 0
        grand_total = subtotal + (subtotal * vat_p / 100)
        paid = getattr(obj, '_total_paid', None) or 0
        bal = grand_total - paid
        if bal <= 0:
            return format_html('<span style="color:green; font-weight:bold;">{}</span>', "0.00")
        return format_html('<span style="color:red; font-weight:bold;">-{}</span>', f"{bal:,.2f}")
    get_balance_due_list.short_description = "ค้างจ่าย"

    class Media:
        js = ('js/admin_sum_selected.js',)

# 2. หน้า Admin ของ Income Report
@admin.register(IncomeReport)
class IncomeReportAdmin(ExportToExcelMixin, DocumentLockMixin, UnfoldModelAdmin):
    # ✅ ปรับ list_display ให้เอาตัวที่มีสีมาโชว์เลย จะได้ดูง่ายๆ
    list_display = ('so_number', 'get_po_no_customer', 'get_customer_truncated', 'get_grand_total_display', 'get_balance_due_display', 'payment_status')
    list_filter = (
        ('order_date', DjangoDateRangeFilter),
        ('payment_status', MultipleChoicesDropdownFilter),
        ('status', MultipleChoicesDropdownFilter),
        ('customer', AutocompleteSelectMultipleFilter),
        ('items__product__tags', SalesOrderTagsFilter),
    )
    list_filter_submit = True
    search_fields = ('so_number', 'customer__company_name')
    actions = [settle_and_close_orders, settle_income_special, 'calculate_income_totals', 'export_to_excel']

    def get_queryset(self, request):
        from django.db.models import Sum, F, ExpressionWrapper, DecimalField as DField, Q
        return super().get_queryset(request).select_related('customer').annotate(
            _total_items_price=Sum(
                ExpressionWrapper(F('items__quantity_ordered') * F('items__sale_price'), output_field=DField())
            ),
            _total_paid=Sum('payments__amount', filter=Q(payments__amount__gt=0)),
        )

    @admin.display(description="ค้างรับ")
    def get_balance_due_display(self, obj):
        subtotal = getattr(obj, '_total_items_price', None)
        if subtotal is None:
            subtotal = getattr(obj, 'total_items_price', 0) or 0
        else:
            subtotal = subtotal or 0

        paid = getattr(obj, '_total_paid', None)
        if paid is None:
            paid = self.get_revenue_only_paid(obj)
        else:
            paid = paid or 0

        vat_p = obj.vat_percent or 0
        grand_total = subtotal + (subtotal * vat_p / 100)
        balance = grand_total - paid

        color = "red" if balance > 0 else "green"
        formatted_balance = f"{float(balance):,.2f}"
        return format_html('<b style="color:{};">{}</b>', color, formatted_balance)
    get_balance_due_display.short_description = "ค้างรับ"

    @admin.action(description="📝 สรุปยอดเงินรายรับที่เลือก")
    def calculate_income_totals(self, request, queryset):
        grand_total = 0
        paid_total = 0
        total_balance_due = 0 # ✅ เปลี่ยนชื่อจาก balance_total เป็นชื่อนี้ให้อ่านง่าย
        
        for obj in queryset:
            grand_total += float(obj.grand_total or 0)
            paid_total += float(obj.total_paid or 0)
            # ✅ เรียกใช้ฟังก์ชันคำนวณที่เปรมมีอยู่แล้วใน Admin
            total_balance_due += float(self.calculate_balance_due(obj) or 0)

        summary_message = (
            f"💰 สรุปรายรับ {queryset.count()} รายการ: | "
            f"ยอดสุทธิ: {grand_total:,.2f} | "
            f"รับเงินแล้ว: {paid_total:,.2f} | "
            f"⚠️ ค้างรับ (Balance Due): {total_balance_due:,.2f}" # ✅ ใช้คำให้ตรงกับหน้าจอ
        )
        self.message_user(request, summary_message, messages.SUCCESS)

    fieldsets = (
        ('📊 สรุปยอดเงิน (Income Summary)', {
            'fields': (
                ('get_total_items_display', 'get_subtotal_display'), 
                ('get_vat_percent_display','get_vat_amount_display'), 
                ('get_grand_total_display', 'get_total_paid_display', 'get_balance_due_display')
            ),
        }),
        ('📝 ข้อมูลเอกสาร', {
            'fields': ('so_number', 'po_no_customer', 'customer', 'order_date', 'status', 'payment_status')
        }),
    )

    readonly_fields = (
        'so_number', 'customer', 'order_date', 'status', 'payment_status',
        'get_total_items_display', 'get_subtotal_display', 'get_vat_percent_display',
        'get_vat_amount_display', 'get_grand_total_display',
        'get_total_paid_display', 'get_balance_due_display'
    )

    inlines = [SalesItemReadOnlyInline, SalesPaymentInline]

    def get_po_no_customer(self, obj):
        val = obj.po_no_customer or '-'
        return format_html('<span style="font-size:12px;color:#555;">{}</span>', val)
    get_po_no_customer.short_description = "PO ลูกค้า"

    def get_customer_truncated(self, obj):
        name = str(obj.customer) if obj.customer else '-'
        return format_html(
            '<span style="display:inline-block;max-width:140px;overflow:hidden;'
            'text-overflow:ellipsis;white-space:nowrap;" title="{}">{}</span>',
            name, name
        )
    get_customer_truncated.short_description = "Customer"

    # --- Methods ที่ปรับปรุงใหม่ (ใช้ได้ทั้ง List และ Detail) ---
    def save_formset(self, request, form, formset, change):
        # 1. เซฟรายการรับเงิน
        formset.save()
        
        # 2. คำนวณสถานะ
        obj = formset.instance
        # ฝั่งขายใช้ SalesPayment (เปรมต้องเช็ค related_name ใน model นะคะ)
        # ถ้าไม่มีใช้ salespayment_set
        paid = sum(p.amount for p in obj.payments.all())
        total = obj.grand_total
        
        if paid <= 0:
            obj.payment_status = 'Unpaid'
        elif paid < total:
            obj.payment_status = 'Partial' # จ่ายบางส่วน
        else:
            obj.payment_status = 'Paid'
            
        obj.save(update_fields=['payment_status'])

    def get_total_items_display(self, obj):
        # ใช้ Sum จาก django.db.models (ซึ่งในไฟล์ admin ของเปรมยังไม่ได้ import ไว้ด้านบน)
        from django.db.models import Sum
        
        # ดึงจาก related_name='items' ที่ตั้งไว้ใน SalesItem
        result = obj.items.aggregate(total_qty=Sum('quantity_ordered'))
        total = result['total_qty'] or 0
        
        if total > 0:
            return f"{total:,} ชิ้น"
        return "0 ชิ้น"
    
    # ✅ จุดที่ 3: ชื่อตรงนี้ก็ต้องตรงกัน
    get_total_items_display.short_description = "📦 รวมจำนวนสินค้า"

    def get_subtotal_display(self, obj):
        # ✅ จัดรูปแบบด้วย f-string ให้เสร็จก่อน แล้วค่อยส่งเข้า format_html
        value = f"{obj.total_items_price:,.2f}"
        return format_html('<span style="font-size:14px;">{}</span>', value)
    get_subtotal_display.short_description = "💵 ก่อน VAT"

    # 1. ฟังก์ชันดึง % VAT มาโชว์ (อ่านอย่างเดียว)
    def get_vat_percent_display(self, obj):
        return f"{obj.vat_percent}%"
    get_vat_percent_display.short_description = "อัตราภาษี (%)"

    # 2. ฟังก์ชันคำนวณยอดเงิน VAT (ดึงค่าจากแม่มาคำนวณ)
    def get_vat_amount_display(self, obj):
        # คำนวณ: (ราคาก่อน VAT * % VAT) / 100
        subtotal = getattr(obj, 'total_items_price', 0) # สมมติว่าเปรมมี property นี้ใน SalesOrder
        vat_p = obj.vat_percent or 0
        vat_amt = (subtotal * vat_p) / 100
        return f"{vat_amt:,.2f}"
    get_vat_amount_display.short_description = "ภาษีมูลค่าเพิ่ม (VAT)"

    def get_grand_total_display(self, obj):
        subtotal = getattr(obj, '_total_items_price', None)
        if subtotal is None:
            subtotal = getattr(obj, 'total_items_price', 0) or 0
        else:
            subtotal = subtotal or 0
        vat_p = obj.vat_percent or 0
        total = subtotal + ((subtotal * vat_p) / 100)
        formatted_total = f"{total:,.2f}"
        return format_html('<b style="color:#007bff;">{}</b>', formatted_total)
    get_grand_total_display.short_description = "💰 ยอดสุทธิ"

    # 🎯 1. ฟังก์ชันช่วยคำนวณยอดรับ (เฉพาะยอดบวก) เพื่อไม่ให้ค่าใช้จ่ายมาดึงยอด Paid ลง
    def get_revenue_only_paid(self, obj):
        from django.db.models import Sum
        # ✅ ใช้ .payments เพราะเป็น related_name ที่เปรมตั้งไว้
        total = obj.payments.filter(amount__gt=0).aggregate(Sum('amount'))['amount__sum'] or 0
        return total

    # 🎯 2. แก้ไขการแสดงผลยอดรับเงิน
    def get_total_paid_display(self, obj):
        # เปลี่ยนจาก obj.total_paid มาใช้ฟังก์ชันใหม่ที่เราสร้าง
        paid_amount = self.get_revenue_only_paid(obj)
        value = f"{paid_amount:,.2f}"
        return format_html('<b style="color:#28a745;">{}</b>', value)
    get_total_paid_display.short_description = "✅ รับแล้ว"

    # 🎯 3. แก้ไขการคำนวณยอดคงค้าง (Balance Due)
    def calculate_balance_due(self, obj):
        subtotal = getattr(obj, 'total_items_price', 0) or 0
        vat_p = getattr(obj, 'vat_percent', 0) or 0
        # ยอดรวมภาษี (Grand Total)
        grand_total = subtotal + (subtotal * vat_p / 100)
        
        # หักเฉพาะยอดที่รับเงินมาจริง (ยอดบวก) ไม่เอา DC/Rebate มาลบซ้ำที่นี่
        paid = self.get_revenue_only_paid(obj)
        
        return grand_total - paid

    # 🎯 4. (แถม) ฟังก์ชันดูยอดที่โดนหักไป (DC + Rebate) เพื่อความโปร่งใส
    def get_total_deductions_display(self, obj):
        from django.db.models import Sum
        # ✅ ใช้ .payments และกรองเฉพาะยอดติดลบ (DC/Rebate)
        total = obj.payments.filter(amount__lt=0).aggregate(Sum('amount'))['amount__sum'] or 0
        return format_html('<span style="color:#6c757d;">{:,.2f}</span>', total)
    get_total_deductions_display.short_description = "➖ ยอดหักสะสม"

    class Media:
        js = ('js/admin_sum_selected.js',) # เรียกไฟล์ JS มาใช้งาน

#@admin.register(ShipmentPaymentReport)
class ShipmentPaymentReportAdmin(UnfoldModelAdmin):
    # ✅ โชว์มูลค่าที่ส่ง และวันที่จะได้รับเงินของยอดนั้นๆ
    list_display = ['payment_due_date', 'get_so_number', 'get_customer', 'quantity_shipped', 'get_shipment_value_display', 'get_dc_display','get_rebate_display', 'get_total_with_vat_display']
    search_fields = ['sales_order__so_number', 'sales_order__customer__company_name']
    list_filter = ['payment_due_date', 'sales_order__customer']
    list_filter_submit = True
    ordering = ['payment_due_date']

    actions = ['calculate_selected_totals']
    
    list_display = [
        'payment_due_date', 'get_so_number', 'get_customer', 
        'quantity_shipped', 'get_shipment_value_display', 
        'get_dc_display', 'get_rebate_display', 'get_total_with_vat_display'
    ]

    # --- 🧮 ฟังก์ชันคำนวณยอดรวมสำหรับรายการที่เลือก ---
    @admin.action(description="📝 สรุปยอดรวมรายการที่เลือก")
    def calculate_selected_totals(self, request, queryset):
        # 1. สั่งให้ฐานข้อมูลคำนวณ Sum ทุกคอลัมน์พร้อมกัน
        totals = queryset.aggregate(
            total_qty=Sum('quantity_shipped'),
            total_value=Sum('shipment_value'),
            total_dc=Sum('dc_amount'),
            total_rebate=Sum('rebate_amount'),
        )

        # 2. คำนวณยอดสุทธิ (คิด VAT 7%)
        net_before_vat = (totals['total_value'] or 0) - (totals['total_dc'] or 0) - (totals['total_rebate'] or 0)
        total_with_vat = float(net_before_vat) * 1.07 # หรือใช้ logic VAT จาก SO ของเปรม

        # 3. สร้างข้อความสรุป
        summary_message = (
            f"📊 สรุปยอดรวม {queryset.count()} รายการที่เลือก:  |  "
            f"📦 จำนวนรวม: {totals['total_qty'] or 0:,} ชิ้น  |  "
            f"💰 ยอดรวมสินค้า: {totals['total_value'] or 0:,.2f} บาท  |  "
            f"🔻 หัก DC: {totals['total_dc'] or 0:,.2f} บาท  |  "
            f"🔻 หัก Rebate: {totals['total_rebate'] or 0:,.2f} บาท  |  "
            f"✅ ยอดรับสุทธิ (รวม VAT): {total_with_vat:,.2f} บาท"
        )

        # 4. ส่งข้อความไปโชว์ที่หน้าจอ
        self.message_user(request, summary_message, messages.SUCCESS)

    def get_dc_display(self, obj):
        # โชว์ตัวเลขคลีนๆ มีคอมม่าและทศนิยม 2 ตำแหน่ง
        return f"{obj.dc_amount:,.2f}"
    get_dc_display.short_description = "หัก DC"
    get_dc_display.admin_order_field = 'dc_amount'

    def get_rebate_display(self, obj):
        # โชว์ตัวเลขคลีนๆ มีคอมม่าและทศนิยม 2 ตำแหน่ง
        return f"{obj.rebate_amount:,.2f}"
    get_rebate_display.short_description = "หัก Rebate"
    get_rebate_display.admin_order_field = 'rebate_amount'

    def get_so_number(self, obj):
        return obj.sales_order.so_number
    get_so_number.short_description = "เลขที่ SO"
    get_so_number.admin_order_field = 'sales_order__so_number' # ทำให้กดเรียงลำดับได้ด้วยค่ะ

    def get_customer(self, obj):
        # ดึงชื่อลูกค้าจาก SalesOrder -> Customer
        return obj.sales_order.customer.company_name
    get_customer.short_description = 'ลูกค้า' # ชื่อหัวตาราง
    get_customer.admin_order_field = 'sales_order__customer__company_name'

    def get_shipment_value_display(self, obj):
        return f"{obj.shipment_value:,.2f}"
    get_shipment_value_display.short_description = 'มูลค่าสินค้า (ก่อน VAT)'

    def get_total_with_vat_display(self, obj):
        # ดึงค่าจาก property ใน model มาโชว์
        return f"{obj.total_with_vat:,.2f}"
    get_total_with_vat_display.short_description = 'ยอดรวมสุทธิ (รวม VAT)'

    def has_add_permission(self, request):
        return False
    

class CustomerProductContractInline(UnfoldTabularInline):
    model = CustomerProductContract
    autocomplete_fields = ['barcode']
    extra = 3
    fields = ['barcode', 'product', 'barcode_unit_info', 'contract_price', 'dc_percent', 'rebate_percent']
    readonly_fields = ['product', 'barcode_unit_info']
    validate_min = False

    # 🎯 บังคับ step ให้ตรงกับ decimal_places ของแต่ละ field (ไม่งั้น widget default ของ Unfold
    # เติมทศนิยมเกิน เช่น พิมพ์ 9.13 กลายเป็น 9.130 แล้วชน validation "ทศนิยมไม่เกิน N ตำแหน่ง")
    # contract_price รับ 4 ตำแหน่ง ส่วนเปอร์เซ็นต์ยังคง 2 ตำแหน่งเหมือนเดิม
    def formfield_for_dbfield(self, db_field, request, **kwargs):
        if db_field.name == 'contract_price':
            kwargs['widget'] = forms.NumberInput(attrs={'step': '0.0001'})
        elif isinstance(db_field, models.DecimalField):
            kwargs['widget'] = forms.NumberInput(attrs={'step': '0.01'})
        return super().formfield_for_dbfield(db_field, request, **kwargs)

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)
        customer_obj = obj  # Customer instance

        class FormsetWithDupCheck(formset):
            def clean(self):
                super().clean()
                seen_barcodes = set()
                for form in self.forms:
                    if not form.cleaned_data or form.cleaned_data.get('DELETE'):
                        continue
                    barcode = form.cleaned_data.get('barcode')
                    if not barcode:
                        continue
                    # ซ้ำใน inline เดียวกัน
                    if barcode.pk in seen_barcodes:
                        form.add_error('barcode', f'บาร์โค้ด {barcode.code} ซ้ำกันในตารางนี้')
                    seen_barcodes.add(barcode.pk)
                    # ซ้ำกับ DB (customer อื่นหรือ record อื่น)
                    instance_pk = form.instance.pk
                    qs = CustomerProductContract.objects.filter(
                        customer=customer_obj, barcode=barcode
                    )
                    if instance_pk:
                        qs = qs.exclude(pk=instance_pk)
                    if qs.exists():
                        form.add_error('barcode', f'ลูกค้านี้มีราคาสัญญาของ {barcode.code} อยู่แล้ว')

        return FormsetWithDupCheck

    def barcode_unit_info(self, obj):
        if obj and obj.barcode_id:
            b = obj.barcode
            unit = b.unit_name or 'ชิ้น'
            factor = b.conversion_factor or 1
            return format_html('<span style="color:#6b7280;font-size:12px;">{} ({} ชิ้น/หน่วย)</span>', unit, factor)
        return '-'
    barcode_unit_info.short_description = 'หน่วย'

    class Media:
        css = {'all': ('css/contract_admin.css',)}
        js = ('js/contract_barcode_autofill.js',)

@admin.register(Customer)
class CustomerAdmin(DetailedHistoryMixin, DocumentLockMixin, UnfoldModelAdmin):
    list_display = ('company_name', 'contact_person', 'phone', 'tax_id')
    search_fields = ('company_name', 'contact_person', 'phone', 'tax_id', 'buyer_code')
    inlines = [CustomerProductContractInline]

# --- 3. ส่วนหน้าจัดการสัญญาโดยเฉพาะ (T2. ราคาสัญญา&DC/Rebate) ---
@admin.register(CustomerProductContract)
class CustomerProductContractAdmin(DetailedHistoryMixin, DocumentLockMixin, UnfoldModelAdmin):
    list_display = ['customer', 'barcode_display', 'product', 'contract_price', 'dc_percent', 'rebate_percent', 'display_product_tags']
    readonly_fields = ['display_product_tags', 'product', 'barcode_unit_detail']
    # ไม่ใช้ list_editable → ไม่มี spinner, คอลัมน์แคบลง
    list_filter = [
        ('customer', AutocompleteSelectMultipleFilter),
        ('product__tags', AutocompleteSelectMultipleFilter),
    ]
    list_filter_submit = True
    fields = ['customer', 'barcode', 'product', 'barcode_unit_detail', 'display_product_tags', 'contract_price', 'dc_percent', 'rebate_percent']

    # contract_price รับทศนิยม 4 ตำแหน่ง ส่วนเปอร์เซ็นต์ยังคง 2 ตำแหน่งเหมือนเดิม
    def formfield_for_dbfield(self, db_field, request, **kwargs):
        if db_field.name == 'contract_price':
            kwargs['widget'] = forms.NumberInput(attrs={'step': '0.0001'})
        elif isinstance(db_field, models.DecimalField):
            kwargs['widget'] = forms.NumberInput(attrs={'step': '0.01'})
        return super().formfield_for_dbfield(db_field, request, **kwargs)

    search_fields = [
        'customer__company_name',
        'product__name',
        'barcode__code',
        'product__barcodes__code',
        'product__tags__name',
    ]
    autocomplete_fields = ['customer', 'barcode', 'product_tag_link']

    def save_model(self, request, obj, form, change):
        # auto-set product จาก barcode
        if obj.barcode_id and obj.barcode:
            obj.product = obj.barcode.product
        super().save_model(request, obj, form, change)

    def barcode_display(self, obj):
        return obj.barcode.code if obj.barcode else '-'
    barcode_display.short_description = 'บาร์โค้ด'

    def barcode_unit_detail(self, obj):
        if obj and obj.barcode_id:
            b = obj.barcode
            unit = b.unit_name or 'ชิ้น'
            factor = b.conversion_factor or 1
            return format_html(
                '<span style="color:#374151;">หน่วย: <b>{}</b> &nbsp;|&nbsp; {} ชิ้น/หน่วย</span>',
                unit, factor
            )
        return '-'
    barcode_unit_detail.short_description = 'ข้อมูลหน่วย'

    class Media:
        css = {'all': ('css/contract_admin.css',)}
        js = ('js/contract_barcode_autofill.js',)
    
@admin.register(StockAdjustment)
class StockAdjustmentAdmin(UnfoldModelAdmin):
    list_display = ['created_at', 'product', 'adjustment_type', 'quantity', 'adjustment_value', 'reason']
    list_filter = [
        ('adjustment_type', MultipleChoicesDropdownFilter),
        ('product', AutocompleteSelectMultipleFilter),
    ]
    list_filter_submit = True
    autocomplete_fields = ['product']
    search_fields = ['product__name', 'reason']

@admin.register(SalesReport)
class SalesReportAdmin(ExportToExcelMixin, UnfoldModelAdmin):
    list_display = (
        'name', 'get_total_qty', 'get_total_revenue', 
        'get_total_cost_buy', 'get_total_cost_bom', 'get_profit_margin'
    )
    list_filter = (
        ('sales_items__sales_order__delivery_logs__shipped_date', RangeDateTimeFilter),
        ('category', AutocompleteSelectMultipleFilter),
        ('tags', AutocompleteSelectMultipleFilter),
        ('sales_items__sales_order__customer', AutocompleteSelectMultipleFilter), # Path: salesitem -> sales_order -> customer
    )
    list_filter_submit = True
    search_fields = ('name', 'barcodes__code', 'sales_items__sales_order__customer__company_name') # Path: customer__company_name

    # --- ให้การค้นหา ใช้ รูปแบบ และ หรือ ได้ ---
    def get_search_results(self, request, queryset, search_term):
        # ถ้าคนหาใช้เครื่องหมาย | ให้แยกคำแล้วใช้ Logic OR
        if '|' in search_term:
            import operator
            from django.db.models import Q
            from functools import reduce

            parts = [p.strip() for p in search_term.split('|') if p.strip()]
            # สร้าง Query แบบ (field1 OR field2) OR (field1 OR field2)
            q_objects = []
            for part in parts:
                q_part = Q()
                for field in self.search_fields:
                    q_part |= Q(**{f"{field}__icontains": part})
                q_objects.append(q_part)
            
            queryset = queryset.filter(reduce(operator.or_, q_objects)).distinct()
            return queryset, False
        
        # ถ้าไม่มี | ก็ให้ทำงานแบบปกติ (AND)
        return super().get_search_results(request, queryset, search_term)

    actions = ['calculate_selected_totals', 'export_to_excel']

    @admin.action(description="📝 สรุปยอดรวมรายการที่เลือก")
    def calculate_selected_totals(self, request, queryset):
        from django.db.models import Sum

        # ดึงผลรวมจากตัวแปรที่เราคำนวณไว้ใน get_queryset (total_qty และ total_sales_val)
        # เนื่องจากเป็นค่าจากการ annotate เราสามารถใช้ Sum() ซ้ำใน aggregate ได้เลยครับ
        totals = queryset.aggregate(
            sum_qty=Sum('total_qty'),
            sum_revenue=Sum('total_sales_val')
        )

        total_qty = totals['sum_qty'] or 0
        total_revenue = totals['sum_revenue'] or 0
        count = queryset.count()

        # แสดงผลเป็นแถบข้อความสีฟ้า (Info Message) ด้านบน
        self.message_user(
            request,
            f"📊 สรุปข้อมูลที่เลือก ({count} รายการ): "
            f"ส่งสำเร็จรวม: {total_qty:,.0f} ชิ้น | "
            f"ยอดขายรวม: ฿{total_revenue:,.2f}",
            messages.INFO
        )

    def get_queryset(self, request):
        # 1. ตั้งต้นที่สินค้า (Proxy Model)
        qs = super().get_queryset(request)
        
        period = request.GET.get('period', '1year')
        now = timezone.now()
        
        # 2. สร้างเงื่อนไขการกรอง (เน้นที่ยอดส่งสำเร็จเท่านั้น)
        # กรองสถานะใบสั่งซื้อที่ยอมรับได้
        date_query = Q(sales_order__status__in=['Shipped', 'Completed', 'ปิดงาน/ครบถ้วน', 'ส่งบางส่วน'])
        
        if period == '1year':
            date_query &= Q(sales_order__order_date__year=now.year)
        elif period == '4months':
            date_query &= Q(sales_order__order_date__gte=now - timedelta(days=120))
        elif period == '1month':
            date_query &= Q(sales_order__order_date__gte=now - timedelta(days=30))

        # 3. ใช้ Subquery เพื่อคำนวณยอด "ส่งสำเร็จ" (quantity_shipped) โดยเฉพาะ
        # วิธีนี้จะดึงยอด 700 มาโชว์ (ไม่ใช่ 2,100 และไม่เบิ้ลเป็น 6,300)
        shipped_subquery = SalesItem.objects.filter(
            product=OuterRef('pk'),
            **{f"{k}": v for k, v in date_query.children} # ส่งเงื่อนไข Shipped และวันที่เข้าไป
        ).values('product').annotate(
            total=Sum('quantity_shipped') # 🎯 เปลี่ยนจาก ordered เป็น shipped ตรงนี้ครับ!
        ).values('total')

        revenue_subquery = SalesItem.objects.filter(
            product=OuterRef('pk'),
            **{f"{k}": v for k, v in date_query.children}
        ).values('product').annotate(
            # sale_price = ราคาต่อหน่วยบาร์โค้ด แต่ quantity_shipped สะสมเป็นชิ้นเสมอ
            # (ดู SalesDeliveryLog.save) จึงต้องหารด้วย conversion_factor ก่อนคูณราคา
            total=Sum(
                F('sale_price') * F('quantity_shipped') / Coalesce(F('barcode_obj__conversion_factor'), Value(1)), # 🎯 คำนวณรายได้จากยอดส่งจริงเท่านั้น
                output_field=DecimalField()
            )
        ).values('total')

        # ต้นทุน BOM: เฉพาะรายการที่มี bom ถูก assign เท่านั้น
        # ใช้ BOM ที่เลือกในแต่ละรายการ ไม่ใช่ค่าเฉลี่ย
        bom_cost_subquery = SalesItem.objects.filter(
            product=OuterRef('pk'),
            bom__isnull=False,
            **{f"{k}": v for k, v in date_query.children}
        ).annotate(
            item_bom_cost=Subquery(
                BOMIngredient.objects.filter(
                    bom=OuterRef('bom')
                ).values('bom').annotate(
                    cost=Sum(
                        F('material__buy_price') * F('quantity') * Coalesce(F('barcode_obj__conversion_factor'), Value(1)),
                        output_field=DecimalField()
                    )
                ).values('cost')[:1]
            )
        ).values('product').annotate(
            total=Sum(F('item_bom_cost') * F('quantity_shipped'), output_field=DecimalField())
        ).values('total')

        # 4. เอาค่าที่บวกได้มาแปะในรายงาน
        return qs.annotate(
            total_qty=Subquery(shipped_subquery),
            total_sales_val=Subquery(revenue_subquery),
            total_bom_cost=Subquery(bom_cost_subquery)
        ).filter(total_qty__gt=0) # 🎯 โชว์เฉพาะสินค้าที่ "ส่งสำเร็จ" จริงๆ ในรอบนั้นๆ
    
    # 🎯 หัวใจหลัก: คำนวณยอดรวมของทั้งหน้า (Grand Total)
    def changelist_view(self, request, extra_context=None):
        response = super().changelist_view(request, extra_context)
        
        try:
            # ดึงข้อมูลมาคำนวณ
            cl = response.context_data['cl']
            qs = cl.get_queryset(request)
            
            aggregates = qs.aggregate(
                g_qty=Sum('total_qty'),
                g_rev=Sum('total_sales_val'),
                g_buy_cost=Sum(F('buy_price') * F('total_qty'), output_field=DecimalField()),
                g_bom_cost=Sum('total_bom_cost')
            )

            g_rev = aggregates['g_rev'] or 0
            g_buy_cost = aggregates['g_buy_cost'] or 0
            g_bom_cost = aggregates['g_bom_cost'] or 0
            g_profit = g_rev - g_buy_cost

            summary = {
                "qty": "{:,.0f}".format(aggregates['g_qty'] or 0),
                "rev": "{:,.2f}".format(g_rev),
                "buy": "{:,.2f}".format(g_buy_cost),
                "bom": "{:,.2f}".format(g_bom_cost),
                "profit": "{:,.2f}".format(g_profit)
            }
            
            # ✅ ใช้ปีกกาคู่ {{ }} สำหรับส่วนที่เป็น JavaScript แท้ๆ 
            # และใช้ {variable} สำหรับส่วนที่ดึงมาจาก Python
            summary_json = json.dumps(summary)
            js_code = """
                <script>
                    document.addEventListener('DOMContentLoaded', function() {{
                        const data = {0};
                        const table = document.querySelector('#result_list');
                        if (table) {{
                            const tfoot = document.createElement('tfoot');
                            tfoot.innerHTML = `
                                <tr style="font-weight: bold; background: #f8f9fa; border-top: 2px solid #dee2e6;">
                                    <td style="color: #333;">ยอดรวมทั้งหมด (TOTAL)</td>
                                    <td>${{data.qty}}</td>
                                    <td>${{data.rev}}</td>
                                    <td>${{data.buy}}</td>
                                    <td>${{data.bom}}</td>
                                    <td style="color: ${{parseFloat(data.profit.replace(/,/g, '')) >= 0 ? '#28a745' : '#dc3545'}}">
                                        ${{data.profit}}
                                    </td>
                                </tr>
                            `;
                            table.appendChild(tfoot);
                        }}
                    }});
                </script>
            """.format(summary_json) # ✅ ใช้ .format แทน f-string เพื่อความชัวร์

            extra_context = extra_context or {}
            extra_context['summary_js'] = mark_safe(js_code)
            return super().changelist_view(request, extra_context)
            
        except Exception as e:
            # ถ้าเกิด Error ให้รันหน้าปกติไปก่อน ไม่ต้องค้าง
            print(f"Error in C5 Total: {e}")
            return response
        
    # --- ฟังก์ชันแสดงผลรายบรรทัด (เหมือนเดิม) --- -
    @admin.display(description="จำนวนขาย")
    def get_total_qty(self, obj): return f"{obj.total_qty or 0:,.0f} {obj.unit}"

    @admin.display(description="ยอดขายรวม")
    def get_total_revenue(self, obj): return f"{obj.total_sales_val or 0:,.2f}"

    @admin.display(description="ต้นทุนรวม (Buy)")
    def get_total_cost_buy(self, obj):
        return f"{(obj.buy_price or 0) * (obj.total_qty or 0):,.2f}"

    @admin.display(description="ต้นทุน BOM")
    def get_total_cost_bom(self, obj):
        cost = obj.total_bom_cost or 0
        return f"{float(cost):,.2f}"

    @admin.display(description="กำไร (vs Buy)")
    def get_profit_margin(self, obj):
        revenue = obj.total_sales_val or 0
        buy_cost = (obj.buy_price or 0) * (obj.total_qty or 0)
        profit = revenue - buy_cost
        color = "#28a745" if profit > 0 else "#dc3545"
        profit_display = "{:,.2f}".format(profit)
        return format_html('<b style="color: {};">{}</b>', color, profit_display)

    class Media:
        js = ('js/admin_sum_selected.js',) # เรียกไฟล์ JS มาใช้งาน

# 2. ตั้งค่า Admin ตัวเดียวจบ
@admin.register(ShipmentAccounting)
class ShipmentAccountingAdmin(ExportToExcelMixin, UnfoldModelAdmin):
    # ✅ เพิ่ม Action ที่ต้องการให้โชว์แยกกันใน List นี้ครับ
    actions = [
        'confirm_selected_items',
        'confirm_revenue_only',
        'confirm_dc_only',
        'confirm_rebate_only',
        'calculate_selected_totals',
        'export_to_excel',
    ]

    list_display = (
        'short_shipped_date', 'get_so_number', 'product', 'quantity_shipped', 
        'get_revenue_no_vat', 'get_revenue_inc_vat', 
        'get_dc_value', 'get_rebate_value',
        'is_revenue_confirmed', 'is_dc_confirmed', 'is_rebate_confirmed'
    )
    
    list_filter = (
        ('shipped_date', RangeDateTimeFilter),
        ('is_revenue_confirmed', BooleanRadioFilter),
        ('is_dc_confirmed', BooleanRadioFilter),
        ('is_rebate_confirmed', BooleanRadioFilter),
        ('sales_order__customer', AutocompleteSelectMultipleFilter),
    )
    list_filter_submit = True
    
    search_fields = ('sales_order__so_number', 'product__name', 'product__barcodes__code') 
    ordering = ('-shipped_date', 'sales_order__so_number')

    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'sales_order', 'sales_order__customer', 'product'
        )

    # --- ให้การค้นหา ใช้ รูปแบบ และ หรือ ได้ ---
    def get_search_results(self, request, queryset, search_term):
        # ถ้าคนหาใช้เครื่องหมาย | ให้แยกคำแล้วใช้ Logic OR
        if '|' in search_term:
            import operator
            from django.db.models import Q
            from functools import reduce

            parts = [p.strip() for p in search_term.split('|') if p.strip()]
            # สร้าง Query แบบ (field1 OR field2) OR (field1 OR field2)
            q_objects = []
            for part in parts:
                q_part = Q()
                for field in self.search_fields:
                    q_part |= Q(**{f"{field}__icontains": part})
                q_objects.append(q_part)
            
            queryset = queryset.filter(reduce(operator.or_, q_objects)).distinct()
            return queryset, False
        
        # ถ้าไม่มี | ก็ให้ทำงานแบบปกติ (AND)
        return super().get_search_results(request, queryset, search_term)

    # --- 📅 จัดการวันที่ ---
    def short_shipped_date(self, obj):
        if obj.shipped_date:
            return obj.shipped_date.strftime('%d/%m/%y %H:%M')
        return "-"
    short_shipped_date.short_description = "วันที่ส่ง"
    short_shipped_date.admin_order_field = 'shipped_date'

    @admin.action(description="📝 สรุปยอดรวมรายการที่เลือก (เฉพาะที่ติ๊ก)")
    def calculate_selected_totals(self, request, queryset):
        total_qty = 0
        total_revenue = Decimal('0')
        total_dc = Decimal('0')
        total_rebate = Decimal('0')
        count = queryset.count()

        for obj in queryset:
            # ดึงข้อมูลสินค้าเพื่อเอาราคาขาย
            item = obj.sales_order.items.filter(product=obj.product).first()
            if item:
                qty = obj.quantity_shipped
                rev = item.sale_price * qty
                
                total_qty += qty
                total_revenue += rev
                
                # ดึง Contract เพื่อคำนวณ DC/Rebate
                from .models import CustomerProductContract
                c = CustomerProductContract.objects.filter(
                    customer=obj.sales_order.customer, 
                    product=obj.product
                ).first()
                
                if c:
                    total_dc += (rev * c.dc_percent) / Decimal('100')
                    total_rebate += (rev * c.rebate_percent) / Decimal('100')

        # แสดงผลลัพธ์เป็นข้อความ Alert สีเขียวด้านบน
        msg = (
            f"✅ สรุป {count} รายการที่เลือก: "
            f"จำนวนรวม: {total_qty:,} ชิ้น | "
            f"ยอดรวม VAT: ฿{total_revenue:,.2f} | "
            f"DC: ฿{total_dc:,.2f} | "
            f"Rebate: ฿{total_rebate:,.2f}"
        )
        self.message_user(request, msg, messages.SUCCESS)

    # --- 📊 สรุปยอดเงิน (Banner สีเหลือง) ---
    def changelist_view(self, request, extra_context=None):
        cl = self.get_changelist_instance(request)
        qs = cl.get_queryset(request)

        from django.db.models import Sum
        totals = qs.aggregate(
            sum_vat=Sum('shipment_value'),
            sum_dc=Sum('dc_amount'),
            sum_rebate=Sum('rebate_amount'),
        )
        sum_vat = totals['sum_vat'] or Decimal('0')
        sum_dc = totals['sum_dc'] or Decimal('0')
        sum_rebate = totals['sum_rebate'] or Decimal('0')

        if qs.exists():
            msg = f"📊 สรุปยอดช่วงที่เลือก: ยอดรวม ฿{sum_vat:,.2f} | DC ฿{sum_dc:,.2f} | Rebate ฿{sum_rebate:,.2f}"
            messages.info(request, msg)

        return super().changelist_view(request, extra_context=extra_context)

    # --- 💰 ฟังก์ชันคำนวณเงินต่างๆ ---
    def get_revenue_inc_vat(self, obj):
        return f"{obj.calculate_revenue_total():,.2f}" # ✅ เรียกจาก Model สั้นๆ
    get_revenue_inc_vat.short_description = "incl.VAT"

    def get_revenue_no_vat(self, obj):
        return f"{obj.shipment_value:,.2f}"
    get_revenue_no_vat.short_description = "excl.VAT"

    def get_dc_value(self, obj):
        dc_amt = obj.dc_amount or Decimal('0')
        if not dc_amt:
            return "-"
        return format_html('<b>฿{}</b>', f"{dc_amt:,.2f}")
    get_dc_value.short_description = "ยอดDC"

    def get_rebate_value(self, obj):
        reb_amt = obj.rebate_amount or Decimal('0')
        if not reb_amt:
            return "-"
        return format_html('<b>฿{}</b>', f"{reb_amt:,.2f}")
    get_rebate_value.short_description = "ยอดRebate"

    # 🎯 5. ยอด ที่ยืนยันทั้งหมด จะถูกบันทึกย้อนไปใน salesorder และ incomereport
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        from .models import SalesPayment
        if obj.is_revenue_confirmed:
            
            SalesPayment.objects.update_or_create(
                order=obj.sales_order,
                remark__icontains=f"ยอดส่งของ {obj.shipping_no}",
                defaults={
                    'amount': obj.calculate_revenue_total(), # ✅ ใช้ตัวเลขดิบๆ ไปบันทึก
                    'payment_date': obj.confirmed_date or obj.shipped_date,
                    'remark': f"✔ ยอดส่งของเลขที่ {obj.shipping_no}"
                }
            )

        # 🎯 [SECTION 2] ยืนยันยอด Rebate (รายการหัก 1)
        if obj.is_rebate_confirmed and obj.rebate_amount > 0:
            rebate_ref = f"หัก Rebate จากใบส่งของ {obj.shipping_no}"
            if not SalesPayment.objects.filter(order=obj.sales_order, remark__icontains=rebate_ref).exists():
                SalesPayment.objects.create(
                    order=obj.sales_order,
                    amount=-obj.rebate_amount, # ติดลบเพื่อหักยอด
                    payment_date=obj.confirmed_date or timezone.now(),
                    remark=f"หักค่า Rebate สินค้า {obj.product.name} [REF-ID:{obj.id}]"
                )

        # 🎯 [SECTION 3] ยืนยันยอด DC (รายการหัก 2)
        if obj.is_dc_confirmed and obj.dc_amount > 0:
            dc_ref = f"หักค่า DC จากใบส่งของ {obj.shipping_no}"
            if not SalesPayment.objects.filter(order=obj.sales_order, remark__icontains=dc_ref).exists():
                SalesPayment.objects.create(
                    order=obj.sales_order,
                    amount=-obj.dc_amount, # ติดลบเพื่อหักยอด
                    payment_date=obj.confirmed_date or timezone.now(),
                    remark=f"หักค่า DC สินค้า {obj.product.name} [REF-ID:{obj.id}]"
                )

    # --- ✅ Actions ---
    @admin.action(description="💰 ยืนยันเฉพาะยอดรับเงิน (Revenue)")
    def confirm_revenue_only(self, request, queryset):
        for obj in queryset:
            if obj.is_revenue_confirmed:
                continue
            obj.is_revenue_confirmed = True
            # 🔥 บังคับเรียก save_model เพื่อให้สร้าง SalesPaymentLog
            self.save_model(request, obj, None, True) 
        self.message_user(request, f"ยืนยันยอดรับเงิน {queryset.count()} รายการ และสร้างประวัติเงินแล้ว")

    @admin.action(description="🚚 ยืนยันเฉพาะค่า DC")
    def confirm_dc_only(self, request, queryset):
        for obj in queryset:
            if obj.is_dc_confirmed:
                continue
            obj.is_dc_confirmed = True
            # 🔥 บังคับเรียก save_model เพื่อให้สร้างรายการหักเงิน
            self.save_model(request, obj, None, True)
        self.message_user(request, f"ยืนยันยอด DC {queryset.count()} รายการ และหักยอดจ่ายแล้ว")

    @admin.action(description="🎁 ยืนยันเฉพาะยอด Rebate")
    def confirm_rebate_only(self, request, queryset):
        for obj in queryset:
            if obj.is_rebate_confirmed:
                continue
            obj.is_rebate_confirmed = True
            # 🔥 บังคับเรียก save_model เพื่อให้สร้างรายการหักเงิน
            self.save_model(request, obj, None, True)
        self.message_user(request, f"ยืนยันยอด Rebate {queryset.count()} รายการ และหักยอดจ่ายแล้ว")

    @admin.action(description="✅ ยืนยันยอดทั้งหมด (ครบทุกส่วน)")
    def confirm_selected_items(self, request, queryset):
        for obj in queryset:
            obj.is_revenue_confirmed = True
            obj.is_dc_confirmed = True
            obj.is_rebate_confirmed = True
            # สั่ง Save ทีละตัวเพื่อให้ save_model ที่เราเขียนไว้ทำงาน
            self.save_model(request, obj, None, True)
        self.message_user(request, f"ยืนยันและบันทึกประวัติการเงิน {queryset.count()} รายการแล้ว")

    def get_so_number(self, obj):
        return obj.sales_order.so_number
    get_so_number.short_description = "เลขที่ SO"

    class Media:
        js = ('js/admin_sum_selected.js',) # เรียกไฟล์ JS มาใช้งาน

@admin.register(InternationalPurchaseTracking)
class InternationalPurchaseTrackingAdmin(ExportToExcelMixin, UnfoldModelAdmin):
    # ✅ ย่อหน้า (Indent) ต้องตรงกันแบบนี้ครับ สีแดงถึงจะหาย
    actions = ['export_to_excel']
    list_display = ('po_number', 'supplier', 'status', 'payment_status', 'display_tracking_table','arrived_date')
    list_filter = (
        ('status', MultipleChoicesDropdownFilter),
        ('supplier', AutocompleteSelectMultipleFilter),
        ('order_date', DjangoDateRangeFilter),
    )
    list_filter_submit = True
    
    # ⚠️ สำคัญมาก: ใน models.py ของเปรม Supplier ใช้ชื่อฟิลด์ 'company_name' ไม่ใช่ 'name'
    search_fields = ('po_number', 'supplier__company_name') 

    def get_queryset(self, request):
        # ให้โชว์เฉพาะ Supplier ที่เป็น 'International' เท่านั้น
        return super().get_queryset(request).filter(supplier__type='International')
    
    def display_tracking_table(self, obj):
        from django.utils.safestring import mark_safe
        
        # 🎯 เตรียมข้อมูล Milestone (ชื่อ, วันที่)
        milestones = [
            ('Ordered', obj.order_date),
            ('Paid', obj.paid_date), 
            ('Loaded', obj.loaded_date),
            ('Departed', obj.departed_date),
            ('Arrived', obj.arrived_date),
            ('Received', obj.received_date),
        ]
        
        # ส่วนแสดงผล Related PO (ถ้ามี)
        rel_po_html = ""
        if obj.related_po:
            rel_po_html = f"<div style='margin-bottom:5px; color:#666;'>🔗 เชื่อมโยงกับ: <b>{obj.related_po.po_number}</b></div>"
        
        headers = "".join([f"<th style='border:1px solid #ddd; padding:4px; background:#f8f9fa; font-size:10px;'>{m[0]}</th>" for m in milestones])
        
        cells = ""
        for name, date in milestones:
            # 🛡️ ป้องกันกรณีข้อมูลไม่ใช่ Date object (เช่น เป็น String หรือ None)
            if date and hasattr(date, 'strftime'):
                date_str = date.strftime('%d/%m/%y')
            else:
                date_str = "-"
            
            # เช็กสถานะปัจจุบันเพื่อเน้นสี
            is_active = (obj.status == name)
            color = "#28a745" if is_active else "#666"
            weight = "bold" if is_active else "normal"
            bg = "#e8f5e9" if is_active else "transparent"
            
            cells += f"<td style='border:1px solid #ddd; padding:4px; color:{color}; font-weight:{weight}; background:{bg};'>{date_str}</td>"

        return mark_safe(
            f"{rel_po_html}"
            f"<table style='width:100%; text-align:center; border-collapse:collapse; font-size:11px;'>"
            f"<thead><tr>{headers}</tr></thead>"
            f"<tbody><tr>{cells}</tr></tbody></table>"
        )
    display_tracking_table.short_description = "📅 Timeline การส่งมอบ"

    def save_formset(self, request, form, formset, change):
        formset.save()
        obj = formset.instance
        if formset.model == PurchasePaymentLog:
            from django.db.models import Sum
            paid = PurchasePaymentLog.objects.filter(purchase_order=obj).aggregate(Sum('amount'))['amount__sum'] or 0
            total = obj.grand_total
            update_fields = ['payment_status']

            if paid <= 0:
                obj.payment_status = 'Unpaid'
            elif paid < total:
                obj.payment_status = 'Partial'
            else:
                obj.payment_status = 'Paid'
                latest = PurchasePaymentLog.objects.filter(purchase_order=obj).order_by('-payment_date').first()
                if latest:
                    obj.paid_date = latest.payment_date
                    update_fields.append('paid_date')

            obj.save(update_fields=update_fields)

    # ✅ 5. Actions: ขยับสถานะ Milestone แบบรวดเร็ว (ครบชุด)
    actions = ['set_paid', 'set_loaded', 'set_departed', 'set_arrived', 'set_received', 'set_closed', 'export_to_excel']

    @admin.action(description='💰 2. จ่ายเงินแล้ว (Paid)')
    def set_paid(self, request, queryset):
        from django.utils import timezone
        # ใช้ update_fields เพื่อความชัวร์ว่าลงเฉพาะจุด
        count = 0
        for obj in queryset:
            obj.status = 'Paid'
            obj.paid_date = timezone.now().date()
            obj.save()
            count += 1
        self.message_user(request, f"✅ อัปเดต 'จ่ายเงินแล้ว' {count} รายการ")

    @admin.action(description='📦 3. ขึ้นตู้แล้ว (Loaded)')
    def set_loaded(self, request, queryset):
        from django.utils import timezone
        count = 0
        for obj in queryset:
            obj.status = 'Loaded'
            obj.loaded_date = timezone.now().date()
            obj.save()
            count += 1
        self.message_user(request, f"✅ อัปเดต 'ขึ้นตู้แล้ว' {count} รายการ")

    @admin.action(description='🚢 4. ออกเดินทาง (Departed)')
    def set_departed(self, request, queryset):
        from django.utils import timezone
        count = 0
        for obj in queryset:
            obj.status = 'Departed'
            obj.departed_date = timezone.now().date()
            obj.save()
            count += 1
        self.message_user(request, f"✅ อัปเดต 'ออกเดินทางแล้ว' {count} รายการ")

    @admin.action(description='🏁 5. ถึงไทยแล้ว (Arrived)')
    def set_arrived(self, request, queryset):
        from django.utils import timezone
        count = 0
        for obj in queryset:
            obj.status = 'Arrived'
            obj.arrived_date = timezone.now().date()
            obj.save()
            count += 1
        self.message_user(request, f"✅ อัปเดต 'ถึงไทยแล้ว' {count} รายการ")

    @admin.action(description='🏢 6. ถึงโกดังแล้ว (Received)')
    def set_received(self, request, queryset):
        from django.utils import timezone
        count = 0
        for obj in queryset:
            obj.status = 'Received'
            obj.received_date = timezone.now().date()
            obj.save()
            count += 1
        self.message_user(request, f"✅ อัปเดต 'ถึงโกดังแล้ว' {count} รายการ")

    @admin.action(description='🔒 7. ปิดใบสั่งซื้อ (Closed/ซ่อนรายการ)')
    def set_closed(self, request, queryset):
        queryset.update(status='Closed')
        
class ConditionInline(UnfoldTabularInline):
    model = ContractCondition
    extra = 1
    autocomplete_fields = ['product', 'product_tag_link'] 
    extra = 1 # 🎯 ยิงบาร์โค้ดหาสินค้าได้เหมือนเดิม
    fields = ['type', 'period', 'product', 'product_tag_link', 'method', 'value']

@admin.register(SalesContract)
class SalesContractAdmin(ExportToExcelMixin, UnfoldModelAdmin):
    list_display = ('contract_name', 'customer', 'start_date', 'end_date', 'is_active')
    search_fields = ('contract_name', 'customer__company_name')
    autocomplete_fields = ['customer']
    inlines = [ConditionInline]

    actions = ['calculate_pending_rebates', 'export_to_excel']

    @admin.action(description="🔄 คำนวณยอดเงินคืนและสร้างใบสำคัญจ่าย")
    def calculate_pending_rebates(self, request, queryset):
        import calendar
        from decimal import Decimal

        created_count = 0
        skipped_count = 0

        for contract in queryset:
            conditions = contract.conditions.all()
            if not conditions.exists():
                self.message_user(request, f"⚠️ สัญญา '{contract}' ไม่มีเงื่อนไข ข้ามไป", level='warning')
                continue

            for condition in conditions:
                # --- 1. กรอง deliveries ตามลูกค้า + ช่วงสัญญา + ยังไม่ถูกนับ ---
                deliveries_qs = SalesDeliveryLog.objects.filter(
                    sales_order__customer=contract.customer,
                    shipped_date__date__gte=contract.start_date,
                    shipped_date__date__lte=contract.end_date,
                    payout_items__isnull=True,
                )

                if condition.type == 'PRODUCT_BASED':
                    if condition.product:
                        deliveries_qs = deliveries_qs.filter(product=condition.product)
                    elif condition.product_tag_link:
                        deliveries_qs = deliveries_qs.filter(product__tags=condition.product_tag_link)
                    else:
                        deliveries_qs = deliveries_qs.none()

                deliveries = list(deliveries_qs.select_related('product'))
                if not deliveries:
                    skipped_count += 1
                    continue

                # --- 2. จัดกลุ่มตาม period ---
                groups = {}
                for d in deliveries:
                    dt = d.shipped_date
                    if condition.period == 'MONTHLY':
                        key = ('M', dt.year, dt.month)
                    elif condition.period == 'QUARTERLY':
                        key = ('Q', dt.year, (dt.month - 1) // 3 + 1)
                    else:  # YEARLY / YTD
                        key = ('Y', dt.year, 0)
                    groups.setdefault(key, []).append(d)

                # --- 3. สร้าง payout ต่อกลุ่ม ---
                for key, items in groups.items():
                    period_type, year, sub = key

                    if period_type == 'M':
                        month = sub
                        period_start = datetime.date(year, month, 1)
                        period_end = datetime.date(year, month, calendar.monthrange(year, month)[1])
                        if contract.payout_trigger == 'END_OF_CONTRACT':
                            payout_date = contract.end_date
                        elif contract.payout_delay == 'NEXT_PERIOD':
                            nm = month % 12 + 1
                            ny = year + (month // 12)
                            payout_date = datetime.date(ny, nm, min(contract.payout_day, calendar.monthrange(ny, nm)[1]))
                        else:
                            payout_date = datetime.date(year, month, min(contract.payout_day, calendar.monthrange(year, month)[1]))

                    elif period_type == 'Q':
                        q = sub
                        sm = (q - 1) * 3 + 1
                        em = q * 3
                        period_start = datetime.date(year, sm, 1)
                        period_end = datetime.date(year, em, calendar.monthrange(year, em)[1])
                        payout_date = contract.end_date

                    else:  # YEARLY / YTD
                        period_start = datetime.date(year, 1, 1)
                        period_end = datetime.date(year, 12, 31)
                        payout_date = contract.end_date

                    # ตรวจซ้ำ
                    if RebatePayout.objects.filter(
                        contract=contract,
                        period_start=period_start,
                        period_end=period_end,
                    ).exists():
                        skipped_count += 1
                        continue

                    # --- 4. คำนวณยอด ---
                    total_sales = sum(i.shipment_value for i in items)
                    if condition.method == 'PERCENT_SALES':
                        rebate = total_sales * (condition.value / Decimal('100'))
                    else:  # AMOUNT_PER_QTY
                        total_qty = sum(i.quantity_shipped for i in items)
                        rebate = Decimal(str(total_qty)) * condition.value

                    # --- 5. สร้าง RebatePayout + Items ---
                    payout = RebatePayout.objects.create(
                        contract=contract,
                        period_start=period_start,
                        period_end=period_end,
                        payout_date=payout_date,
                        total_sales_amount=total_sales,
                        rebate_amount=rebate,
                    )
                    for delivery in items:
                        RebatePayoutItem.objects.create(
                            payout=payout,
                            delivery=delivery,
                            shipped_date=delivery.shipped_date,
                        )
                    created_count += 1

        msg = f"✅ สร้างใบสรุป Rebate ใหม่ {created_count} รายการ"
        if skipped_count:
            msg += f" (ข้าม {skipped_count} รายการ: ซ้ำหรือไม่มีข้อมูล)"
        self.message_user(request, msg)

class RebatePayoutItemInline(UnfoldTabularInline):
    model = RebatePayoutItem
    extra = 0
    can_delete = True
    fields = ('get_product', 'shipped_date', 'get_qty', 'get_value', 'get_rebate')
    readonly_fields = ('get_product', 'get_qty', 'get_value', 'get_rebate')
    ordering = ('shipped_date',)

    def get_product(self, obj):
        return obj.delivery.product
    get_product.short_description = "สินค้า"

    def get_qty(self, obj):
        return obj.delivery.quantity_shipped
    get_qty.short_description = "จำนวน"

    def get_value(self, obj):
        return f"{obj.delivery.shipment_value:,.2f}"
    get_value.short_description = "มูลค่า (บาท)"

    def get_rebate(self, obj):
        return f"{obj.delivery.rebate_amount:,.2f}"
    get_rebate.short_description = "Rebate (บาท)"


@admin.register(RebatePayout)
class RebatePayoutAdmin(ExportToExcelMixin, UnfoldModelAdmin):
    list_display = ('contract', 'period_start', 'period_end', 'payout_date', 'total_sales_amount', 'rebate_amount', 'status', 'ref_invoice')
    list_filter = (
        ('status', MultipleChoicesDropdownFilter),
        ('contract__customer', AutocompleteSelectMultipleFilter),
    )
    list_filter_submit = True
    search_fields = ('contract__contract_name', 'contract__customer__company_name', 'ref_invoice')
    readonly_fields = ('contract', 'period_start', 'period_end', 'total_sales_amount', 'rebate_amount')
    list_display_links = ('contract',)
    ordering = ('-payout_date',)
    inlines = [RebatePayoutItemInline]


# ============================================================
# B5. ใบเสนอราคาซื้อ (Purchase Quotation)
# ============================================================
class PurchaseQuotationItemInline(UnfoldTabularInline):
    model = PurchaseQuotationItem
    autocomplete_fields = ['product']
    extra = 1
    fields = ['product', 'current_price_display', 'new_price']
    readonly_fields = ['current_price_display']

    def current_price_display(self, obj):
        if not (obj and obj.pk and obj.product_id and obj.quotation_id):
            return '-'
        ps = ProductSupplier.objects.filter(
            product_id=obj.product_id, supplier_id=obj.quotation.supplier_id
        ).first()
        return ps.latest_buy_price if ps and ps.latest_buy_price else 0
    current_price_display.short_description = "ราคาปัจจุบัน (Supplier)"


@admin.register(PurchaseQuotation)
class PurchaseQuotationAdmin(UnfoldModelAdmin):
    list_display = ('pq_number', 'supplier', 'quote_date', 'item_count', 'created_by')
    list_filter = (
        ('supplier', AutocompleteSelectMultipleFilter),
        ('quote_date', DjangoDateRangeFilter),
    )
    list_filter_submit = True
    search_fields = ('pq_number', 'supplier__company_name', 'items__product__name')
    readonly_fields = ('created_by',)
    inlines = [PurchaseQuotationItemInline]
    actions = ['sync_to_supplier_price']

    class Media:
        js = ('js/quotation_price_autofill.js',)

    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related('items')

    def item_count(self, obj):
        return obj.items.count()
    item_count.short_description = "จำนวนรายการ"

    def save_model(self, request, obj, form, change):
        if not obj.pk and not obj.created_by_id:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

    @admin.action(description="💰 อัพเดทราคาสัญญาซื้อ (ลงในราคา Supplier)")
    def sync_to_supplier_price(self, request, queryset):
        created, updated = 0, 0
        for quotation in queryset.prefetch_related('items__product'):
            for item in quotation.items.all():
                obj, is_created = ProductSupplier.objects.get_or_create(
                    product=item.product,
                    supplier=quotation.supplier,
                    defaults={'latest_buy_price': item.new_price},
                )
                if is_created:
                    created += 1
                else:
                    obj.latest_buy_price = item.new_price
                    obj.save(update_fields=['latest_buy_price'])
                    updated += 1
        self.message_user(
            request,
            f"อัพเดทราคาสัญญาซื้อสำเร็จ — เพิ่มใหม่ {created} รายการ, อัพเดทราคา {updated} รายการ"
        )


# ============================================================
# B6. ใบเสนอราคาขาย (Sales Quotation)
# ============================================================
class SalesQuotationItemInline(UnfoldTabularInline):
    model = SalesQuotationItem
    autocomplete_fields = ['product', 'barcode']
    extra = 1
    fields = ['product', 'barcode', 'current_price_display', 'new_price']
    readonly_fields = ['current_price_display']

    def current_price_display(self, obj):
        if not (obj and obj.pk and obj.product_id and obj.quotation_id):
            return '-'
        contract = CustomerProductContract.objects.filter(
            customer_id=obj.quotation.customer_id, product_id=obj.product_id
        ).first()
        return contract.contract_price if contract and contract.contract_price else 0
    current_price_display.short_description = "ราคาปัจจุบัน (Contract)"


@admin.register(SalesQuotation)
class SalesQuotationAdmin(UnfoldModelAdmin):
    list_display = ('sq_number', 'customer', 'quote_date', 'item_count', 'created_by')
    list_filter = (
        ('customer', AutocompleteSelectMultipleFilter),
        ('quote_date', DjangoDateRangeFilter),
    )
    list_filter_submit = True
    search_fields = ('sq_number', 'customer__company_name', 'items__product__name')
    readonly_fields = ('created_by',)
    inlines = [SalesQuotationItemInline]
    actions = ['sync_to_customer_contract']

    class Media:
        js = ('js/quotation_price_autofill.js',)

    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related('items')

    def item_count(self, obj):
        return obj.items.count()
    item_count.short_description = "จำนวนรายการ"

    def save_model(self, request, obj, form, change):
        if not obj.pk and not obj.created_by_id:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

    @admin.action(description="💰 อัพเดทราคาสัญญาขาย (ลงในสัญญาลูกค้า)")
    def sync_to_customer_contract(self, request, queryset):
        created, updated = 0, 0
        for quotation in queryset.prefetch_related('items__product', 'items__barcode'):
            for item in quotation.items.all():
                contract = CustomerProductContract.objects.filter(
                    customer=quotation.customer,
                    product=item.product,
                ).first()
                if contract:
                    contract.contract_price = item.new_price
                    contract.save(update_fields=['contract_price'])
                    updated += 1
                else:
                    barcode = item.barcode or item.product.barcodes.first()
                    CustomerProductContract.objects.create(
                        customer=quotation.customer,
                        product=item.product,
                        barcode=barcode,
                        contract_price=item.new_price,
                    )
                    created += 1
        self.message_user(
            request,
            f"อัพเดทราคาสัญญาขายสำเร็จ — เพิ่มใหม่ {created} รายการ, อัพเดทราคา {updated} รายการ"
        )
    actions = ['export_to_excel']